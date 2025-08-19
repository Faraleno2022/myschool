from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from datetime import date
import os
from .models import Eleve, Responsable, Classe, Ecole, HistoriqueEleve
from .forms import EleveForm, ResponsableForm, RechercheEleveForm, ClasseForm
from utilisateurs.models import JournalActivite
from utilisateurs.utils import user_is_admin, filter_by_user_school, user_school

# ReportLab pour génération PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Utilitaire PDF partagé (filigrane)
from ecole_moderne.pdf_utils import draw_logo_watermark

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    get_column_letter = None

@login_required
def liste_eleves(request):
    """Vue pour afficher la liste des élèves avec recherche et filtres"""
    form_recherche = RechercheEleveForm(request.GET or None)
    
    # Requête de base avec optimisation
    eleves = Eleve.objects.select_related(
        'classe', 'classe__ecole', 'responsable_principal', 'responsable_secondaire'
    ).prefetch_related('paiements')
    # Filtrage par école pour non-admin
    if not user_is_admin(request.user):
        eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    
    # Application du filtre simple (recherche globale multi-critères)
    if form_recherche.is_valid():
        recherche = form_recherche.cleaned_data.get('recherche')
        if recherche:
            eleves = eleves.filter(
                Q(matricule__icontains=recherche) |
                Q(nom__icontains=recherche) |
                Q(prenom__icontains=recherche) |
                Q(classe__nom__icontains=recherche) |
                Q(classe__ecole__nom__icontains=recherche) |
                Q(responsable_principal__nom__icontains=recherche) |
                Q(responsable_principal__prenom__icontains=recherche) |
                Q(responsable_secondaire__nom__icontains=recherche) |
                Q(responsable_secondaire__prenom__icontains=recherche)
            )
    
    # Filtre par classe via paramètre GET (classe_id)
    classe_id = request.GET.get('classe_id')
    if classe_id:
        try:
            # Sécuriser la conversion en entier; si invalide, ignorer
            int(classe_id)
            eleves = eleves.filter(classe_id=classe_id)
        except (TypeError, ValueError):
            classe_id = None
    
    # Sécurisation anti-doublons (au cas de jointures inattendues)
    eleves = eleves.distinct()
    
    # Statistiques
    stats = {
        'total_eleves': eleves.count(),
        'eleves_actifs': eleves.filter(statut='ACTIF').count(),
        'eleves_suspendus': eleves.filter(statut='SUSPENDU').count(),
    }
    
    # Pagination
    paginator = Paginator(eleves.order_by('nom', 'prenom'), 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Log de l'activité
    JournalActivite.objects.create(
        user=request.user,
        action='CONSULTATION',
        type_objet='ELEVE',
        description=f"Consultation de la liste des élèves (page {page_number or 1})",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    # Liste des classes pour export (restreinte si besoin)
    if user_is_admin(request.user):
        classes = Classe.objects.select_related('ecole').order_by('ecole__nom', 'niveau', 'nom')
    else:
        classes = Classe.objects.select_related('ecole').filter(ecole=user_school(request.user)).order_by('niveau', 'nom')

    context = {
        'page_obj': page_obj,
        'form_recherche': form_recherche,
        'stats': stats,
        'titre_page': 'Gestion des Élèves',
        'classes': classes,
        # Conserver la sélection actuelle de classe dans l'UI
        'selected_classe_id': str(classe_id) if classe_id else '',
    }

    # Rendu partiel pour la recherche dynamique
    if request.GET.get('partial') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        response = render(request, 'eleves/partials/_liste_eleves_zone.html', context)
        # Eviter le cache sur le fragment
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    response = render(request, 'eleves/liste_eleves.html', context)
    
    # Headers anti-cache pour s'assurer que la liste est toujours fraîche
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response

@login_required
def detail_eleve(request, eleve_id):
    """Vue pour afficher les détails d'un élève"""
    qs = Eleve.objects.select_related(
        'classe', 'classe__ecole', 'responsable_principal', 'responsable_secondaire'
    ).prefetch_related('paiements', 'historique')
    if not user_is_admin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')
    eleve = get_object_or_404(qs, id=eleve_id)
    
    # Statistiques des paiements
    paiements_stats = {
        'total_paiements': eleve.paiements.count(),
        'paiements_valides': eleve.paiements.filter(statut='VALIDE').count(),
        'montant_total': sum(p.montant for p in eleve.paiements.filter(statut='VALIDE')),
    }
    
    # Historique récent
    historique_recent = eleve.historique.all()[:10]
    
    # Log de l'activité
    JournalActivite.objects.create(
        user=request.user,
        action='CONSULTATION',
        type_objet='ELEVE',
        objet_id=eleve.id,
        description=f"Consultation du profil de {eleve.nom_complet}",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    context = {
        'eleve': eleve,
        'paiements_stats': paiements_stats,
        'historique_recent': historique_recent,
        'titre_page': f'Profil de {eleve.nom_complet}'
    }
    
    return render(request, 'eleves/detail_eleve.html', context)

@login_required
def ajouter_eleve(request):
    """Vue pour ajouter un nouvel élève"""
    # Si l'utilisateur n'est pas admin et n'a pas d'école associée, refuser l'accès
    if not user_is_admin(request.user) and user_school(request.user) is None:
        return render(request, 'utilisateurs/acces_refuse_ecole.html', status=403)
    if request.method == 'POST':
        form = EleveForm(request.POST, request.FILES)
        # Restreindre les classes au périmètre de l'école pour non-admin
        if not user_is_admin(request.user):
            try:
                form.fields['classe'].queryset = Classe.objects.filter(ecole=user_school(request.user))
            except Exception:
                pass
        
        # Gestion des nouveaux responsables
        responsable_principal_form = None
        responsable_secondaire_form = None
        
        # Vérifier si on doit créer de nouveaux responsables
        creer_resp_principal = request.POST.get('responsable_principal_nouveau') == 'on'
        creer_resp_secondaire = request.POST.get('responsable_secondaire_nouveau') == 'on'
        
        if creer_resp_principal:
            responsable_principal_form = ResponsableForm(request.POST, prefix='resp_principal')
        
        if creer_resp_secondaire:
            responsable_secondaire_form = ResponsableForm(request.POST, prefix='resp_secondaire')
        
        # Validation du formulaire principal
        form_valide = form.is_valid()
        
        # Validation des formulaires de responsables
        resp_principal_valide = True
        resp_secondaire_valide = True
        
        if responsable_principal_form:
            resp_principal_valide = responsable_principal_form.is_valid()
        
        if responsable_secondaire_form:
            resp_secondaire_valide = responsable_secondaire_form.is_valid()
        
        # Vérifier qu'un responsable principal est fourni (seulement si le formulaire principal est valide)
        if form_valide:
            if not creer_resp_principal and not form.cleaned_data.get('responsable_principal'):
                form.add_error('responsable_principal', 'Un responsable principal est obligatoire.')
                form_valide = False
        
        if form_valide and resp_principal_valide and resp_secondaire_valide:
            try:
                # Créer les responsables si nécessaire
                if responsable_principal_form:
                    responsable_principal = responsable_principal_form.save()
                    form.instance.responsable_principal = responsable_principal
                
                if responsable_secondaire_form:
                    responsable_secondaire = responsable_secondaire_form.save()
                    form.instance.responsable_secondaire = responsable_secondaire
                
                # Sauvegarder l'élève
                eleve = form.save(commit=False)
                eleve.cree_par = request.user
                eleve.save()
                
                # Créer l'historique
                HistoriqueEleve.objects.create(
                    eleve=eleve,
                    action='CREATION',
                    description=f"Création du profil de {eleve.prenom} {eleve.nom}",
                    utilisateur=request.user
                )
                
                # Log de l'activité
                JournalActivite.objects.create(
                    user=request.user,
                    action='CREATION',
                    type_objet='ELEVE',
                    objet_id=eleve.id,
                    description=f"Création de l'élève {eleve.prenom} {eleve.nom} (matricule: {eleve.matricule})",
                    adresse_ip=request.META.get('REMOTE_ADDR', ''),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                messages.success(request, f"L'élève {eleve.prenom} {eleve.nom} a été ajouté avec succès.")
                return redirect('eleves:detail_eleve', eleve_id=eleve.id)
                
            except Exception as e:
                messages.error(request, f"Erreur lors de l'enregistrement : {str(e)}")
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = EleveForm()
        if not user_is_admin(request.user):
            try:
                form.fields['classe'].queryset = Classe.objects.filter(ecole=user_school(request.user))
            except Exception:
                pass
        responsable_principal_form = ResponsableForm(prefix='resp_principal')
        responsable_secondaire_form = ResponsableForm(prefix='resp_secondaire')
    
    # Statistiques pour l'affichage (tous les élèves selon les permissions)
    eleves = Eleve.objects.all()
    # Filtrage par école pour non-admin
    if not user_is_admin(request.user):
        eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    
    stats = {
        'total_eleves': eleves.count(),
        'eleves_actifs': eleves.filter(statut='ACTIF').count(),
        'eleves_exclus': eleves.filter(statut='EXCLU').count(),
    }
    
    context = {
        'form': form,
        'responsable_principal_form': responsable_principal_form,
        'responsable_secondaire_form': responsable_secondaire_form,
        'stats': stats,
        'titre_page': 'Ajouter un Élève',
        'action': 'Ajouter'
    }
    
    return render(request, 'eleves/ajouter_eleve.html', context)

@login_required
def modifier_eleve(request, eleve_id):
    """Vue pour modifier un élève existant"""
    qs = Eleve.objects.all()
    if not user_is_admin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')
    eleve = get_object_or_404(qs, id=eleve_id)
    
    if request.method == 'POST':
        print(f"POST data received: {request.POST}")  # Debug
        form = EleveForm(request.POST, request.FILES, instance=eleve)
        if not user_is_admin(request.user):
            try:
                form.fields['classe'].queryset = Classe.objects.filter(ecole=user_school(request.user))
            except Exception:
                pass
        
        print(f"Form is valid: {form.is_valid()}")  # Debug
        if not form.is_valid():
            print(f"Form errors: {form.errors}")  # Debug
            
        if form.is_valid():
            print("Form validation passed, saving...")  # Debug
            # Détecter les changements
            changements = []
            for field in form.changed_data:
                if field in form.fields:
                    ancien_val = getattr(eleve, field, '')
                    nouveau_val = form.cleaned_data[field]
                    changements.append(f"{form.fields[field].label}: {ancien_val} → {nouveau_val}")
            
            eleve = form.save()
            print(f"Eleve saved successfully: {eleve}")  # Debug
            
            # Créer l'historique si des changements ont été effectués
            if changements:
                try:
                    HistoriqueEleve.objects.create(
                        eleve=eleve,
                        action='MODIFICATION',
                        description=f"Modification: {', '.join(changements)}",
                        utilisateur=request.user
                    )
                except Exception as e:
                    print(f"Error creating history: {e}")
                
                # Log de l'activité
                try:
                    JournalActivite.objects.create(
                        user=request.user,
                        action='MODIFICATION',
                        type_objet='ELEVE',
                        objet_id=eleve.id,
                        description=f"Modification de l'élève {eleve.nom_complet}: {', '.join(changements)}",
                        adresse_ip=request.META.get('REMOTE_ADDR', ''),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')
                    )
                except Exception as e:
                    print(f"Error creating activity log: {e}")
            
            # Message de succès détaillé
            if changements:
                nb_changements = len(changements)
                message_changements = f" ({nb_changements} modification{'s' if nb_changements > 1 else ''} effectuée{'s' if nb_changements > 1 else ''})"
                messages.success(
                    request, 
                    f"✅ Les informations de {eleve.prenom} {eleve.nom} ont été mises à jour avec succès{message_changements}."
                )
            else:
                messages.success(request, f"✅ Les informations de {eleve.prenom} {eleve.nom} ont été sauvegardées.")
            
            # Rediriger vers la page de modification pour voir les messages
            return redirect('eleves:modifier_eleve', eleve_id=eleve.id)
        else:
            # Formulaire invalide: informer l'utilisateur des erreurs
            # Construire un résumé concis des erreurs (limité pour l'UI)
            erreurs = []
            try:
                for champ, msgs in list(form.errors.items())[:5]:
                    libelle = form.fields.get(champ).label if champ in form.fields else champ
                    erreurs.append(f"{libelle}: {', '.join([str(m) for m in msgs])}")
            except Exception:
                pass
            if erreurs:
                messages.error(request, "Le formulaire contient des erreurs: " + " | ".join(erreurs))
            else:
                messages.error(request, "Le formulaire est invalide. Veuillez corriger les erreurs et réessayer.")
    else:
        form = EleveForm(instance=eleve)
        if not user_is_admin(request.user):
            try:
                form.fields['classe'].queryset = Classe.objects.filter(ecole=user_school(request.user))
            except Exception:
                pass
    
    context = {
        'form': form,
        'eleve': eleve,
        'titre_page': f'Modifier {eleve.nom_complet}',
        'action': 'Modifier'
    }
    
    return render(request, 'eleves/modifier_eleve_simple.html', context)

@login_required
def _get_classe_or_403(request, classe_id):
    qs = Classe.objects.select_related('ecole')
    if not user_is_admin(request.user):
        qs = qs.filter(ecole=user_school(request.user))
    return get_object_or_404(qs, id=classe_id)

@login_required
def export_eleves_classe_pdf(request, classe_id):
    """Exporte la liste des élèves d'une classe en PDF."""
    classe = _get_classe_or_403(request, classe_id)
    eleves = Eleve.objects.select_related('classe', 'responsable_principal').filter(classe=classe).order_by('nom', 'prenom')

    # Log activité
    JournalActivite.objects.create(
        user=request.user,
        action='EXPORT',
        type_objet='ELEVE',
        description=f"Export PDF élèves - Classe {classe.nom} ({classe.ecole.nom})",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )

    response = HttpResponse(content_type='application/pdf')
    filename = f"eleves_{slugify(classe.ecole.nom)}_{slugify(classe.nom)}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)
    c.setPageCompression(1)

    # Polices (Calibri/Arial si dispo, sinon Helvetica par défaut)
    font_name = 'Helvetica'
    font_bold = 'Helvetica-Bold'
    
    try:
        calibri_path = 'C:/Windows/Fonts/calibri.ttf'
        calibri_bold_path = 'C:/Windows/Fonts/calibrib.ttf'
        if os.path.exists(calibri_path) and os.path.exists(calibri_bold_path):
            pdfmetrics.registerFont(TTFont('MainFont', calibri_path))
            pdfmetrics.registerFont(TTFont('MainFont-Bold', calibri_bold_path))
            font_name = 'MainFont'
            font_bold = 'MainFont-Bold'
        else:
            arial_path = 'C:/Windows/Fonts/arial.ttf'
            arial_bold_path = 'C:/Windows/Fonts/arialbd.ttf'
            if os.path.exists(arial_path) and os.path.exists(arial_bold_path):
                pdfmetrics.registerFont(TTFont('MainFont', arial_path))
                pdfmetrics.registerFont(TTFont('MainFont-Bold', arial_bold_path))
                font_name = 'MainFont'
                font_bold = 'MainFont-Bold'
    except Exception:
        # Utiliser les polices par défaut de ReportLab
        pass

    # Filigrane standardisé
    draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)

    margin = 2*cm
    y = height - margin

    # En-tête
    c.setFont(font_bold, 16)
    c.drawString(margin, y, f"Liste des élèves - {classe.ecole.nom}")
    y -= 18
    c.setFont(font_name, 12)
    c.drawString(margin, y, f"Classe: {classe.nom}")
    y -= 10
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 18

    # En-têtes du tableau
    headers = ["Matricule", "Nom", "Sexe", "Date Naissance", "Responsable", "Téléphone"]
    col_widths = [3.0*cm, 7.0*cm, 2.0*cm, 3.0*cm, 5.0*cm, 3.5*cm]
    c.setFont(font_bold, 11)
    x = margin
    for i, htxt in enumerate(headers):
        c.drawString(x, y, htxt)
        x += col_widths[i]
    y -= 14
    c.setFillColor(colors.lightgrey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 8

    c.setFont(font_name, 10)
    for e in eleves:
        # Saut de page si nécessaire
        if y < margin + 40:
            c.showPage()
            # Filigrane sur chaque nouvelle page
            draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
            y = height - margin
            c.setFont(font_bold, 11)
            x = margin
            for i, htxt in enumerate(headers):
                c.drawString(x, y, htxt)
                x += col_widths[i]
            y -= 18
            c.setFont(font_name, 10)

        x = margin
        values = [
            e.matricule or '',
            f"{e.nom} {e.prenom}",
            e.get_sexe_display() if hasattr(e, 'get_sexe_display') else getattr(e, 'sexe', ''),
            e.date_naissance.strftime('%d/%m/%Y') if getattr(e, 'date_naissance', None) else '',
            e.responsable_principal.nom_complet if e.responsable_principal else '',
            e.responsable_principal.telephone if e.responsable_principal else '',
        ]
        for i, val in enumerate(values):
            c.drawString(x, y, str(val))
            x += col_widths[i]
        y -= 14

    c.showPage()
    c.save()
    return response

@login_required
def export_eleves_classe_excel(request, classe_id):
    """Exporte la liste des élèves d'une classe en Excel (.xlsx)."""
    if Workbook is None or get_column_letter is None:
        return HttpResponse("Erreur: openpyxl n'est pas installé sur le serveur.", status=500)

    classe = _get_classe_or_403(request, classe_id)
    eleves = Eleve.objects.select_related('classe', 'responsable_principal').filter(classe=classe).order_by('nom', 'prenom')

    # Log activité
    JournalActivite.objects.create(
        user=request.user,
        action='EXPORT',
        type_objet='ELEVE',
        description=f"Export Excel élèves - Classe {classe.nom} ({classe.ecole.nom})",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Élèves"

        # En-têtes avec style
        headers = ["Matricule", "Nom complet", "Sexe", "Date de naissance", "Responsable principal", "Téléphone"]
        ws.append(headers)
        
        # Style pour les en-têtes
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Données des élèves
        for e in eleves:
            ws.append([
                e.matricule or '',
                f"{e.nom} {e.prenom}",
                e.get_sexe_display() if hasattr(e, 'get_sexe_display') else getattr(e, 'sexe', ''),
                e.date_naissance.strftime('%d/%m/%Y') if getattr(e, 'date_naissance', None) else '',
                e.responsable_principal.nom_complet if e.responsable_principal else '',
                e.responsable_principal.telephone if e.responsable_principal else '',
            ])

        # Largeur des colonnes optimisée
        widths = [15, 30, 10, 18, 30, 18]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        # Réponse HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"eleves_{slugify(classe.ecole.nom)}_{slugify(classe.nom)}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du fichier Excel: {str(e)}", status=500)

@login_required
def export_tous_eleves_pdf(request):
    """Exporte la liste de tous les élèves en PDF."""
    # Filtrer selon les permissions
    if user_is_admin(request.user):
        eleves = Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal').all()
    else:
        eleves = Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal').filter(
            classe__ecole=user_school(request.user)
        )
    
    eleves = eleves.order_by('classe__ecole__nom', 'classe__nom', 'nom', 'prenom')

    # Log activité
    JournalActivite.objects.create(
        user=request.user,
        action='EXPORT',
        type_objet='ELEVE',
        description="Export PDF - Tous les élèves",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )

    response = HttpResponse(content_type='application/pdf')
    filename = "tous_les_eleves.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    try:
        c = canvas.Canvas(response, pagesize=landscape(A4))
        width, height = landscape(A4)
        c.setPageCompression(1)

        # Configuration des polices
        font_name = 'Helvetica'
        font_bold = 'Helvetica-Bold'
        
        try:
            calibri_path = 'C:/Windows/Fonts/calibri.ttf'
            calibri_bold_path = 'C:/Windows/Fonts/calibrib.ttf'
            if os.path.exists(calibri_path) and os.path.exists(calibri_bold_path):
                pdfmetrics.registerFont(TTFont('MainFont', calibri_path))
                pdfmetrics.registerFont(TTFont('MainFont-Bold', calibri_bold_path))
                font_name = 'MainFont'
                font_bold = 'MainFont-Bold'
        except Exception:
            pass

        # Filigrane standardisé (logo centré, rotation, opacité faible)
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)

        margin = 2*cm
        y = height - margin

        # En-tête principal
        c.setFont(font_bold, 18)
        c.drawString(margin, y, "Liste complète des élèves")
        y -= 25
        
        c.setFont(font_name, 12)
        from datetime import datetime
        c.drawString(margin, y, f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
        y -= 15
        
        c.setFillColor(colors.grey)
        c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
        c.setFillColor(colors.black)
        y -= 25

        # En-têtes du tableau
        headers = ["École", "Classe", "Matricule", "Nom", "Responsable"]
        col_widths = [4.5*cm, 3*cm, 3*cm, 6*cm, 5*cm]
        
        current_ecole = None
        
        for eleve in eleves:
            # Nouvelle école
            if current_ecole != eleve.classe.ecole.nom:
                if y < margin + 80:
                    c.showPage()
                    # Filigrane sur chaque nouvelle page
                    draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
                    y = height - margin
                
                current_ecole = eleve.classe.ecole.nom
                
                # Titre de l'école
                c.setFont(font_bold, 14)
                c.drawString(margin, y, f"École: {current_ecole}")
                y -= 20
                
                # En-têtes du tableau
                c.setFont(font_bold, 10)
                x = margin
                for i, header in enumerate(headers[1:]):  # Skip "École" pour cette section
                    c.drawString(x, y, header)
                    x += col_widths[i+1]
                y -= 15
                
                c.setFillColor(colors.lightgrey)
                c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
                c.setFillColor(colors.black)
                y -= 8
            
            # Vérifier l'espace pour une nouvelle ligne
            if y < margin + 40:
                c.showPage()
                # Filigrane sur chaque nouvelle page
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
                y = height - margin
                
                # Répéter le titre de l'école et les en-têtes
                c.setFont(font_bold, 14)
                c.drawString(margin, y, f"École: {current_ecole} (suite)")
                y -= 20
                
                c.setFont(font_bold, 10)
                x = margin
                for i, header in enumerate(headers[1:]):
                    c.drawString(x, y, header)
                    x += col_widths[i+1]
                y -= 18
            
            # Ligne de données
            c.setFont(font_name, 9)
            x = margin
            values = [
                eleve.classe.nom,
                eleve.matricule or '',
                f"{eleve.nom} {eleve.prenom}",
                eleve.responsable_principal.nom_complet if eleve.responsable_principal else '',
            ]
            
            for i, val in enumerate(values):
                # Tronquer si trop long
                text = str(val)[:25] + '...' if len(str(val)) > 25 else str(val)
                c.drawString(x, y, text)
                x += col_widths[i+1]
            y -= 12

        c.showPage()
        c.save()
        return response
        
    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du PDF: {str(e)}", status=500)

@login_required
def export_tous_eleves_excel(request):
    """Exporte la liste de tous les élèves en Excel (.xlsx)."""
    if Workbook is None or get_column_letter is None:
        return HttpResponse("Erreur: openpyxl n'est pas installé sur le serveur.", status=500)

    # Filtrer selon les permissions
    if user_is_admin(request.user):
        eleves = Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal').all()
    else:
        eleves = Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal').filter(
            classe__ecole=user_school(request.user)
        )
    
    eleves = eleves.order_by('classe__ecole__nom', 'classe__nom', 'nom', 'prenom')

    # Log activité
    JournalActivite.objects.create(
        user=request.user,
        action='EXPORT',
        type_objet='ELEVE',
        description="Export Excel - Tous les élèves",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tous les élèves"

        # En-têtes avec style
        headers = ["École", "Classe", "Matricule", "Nom complet", "Sexe", "Date de naissance", "Responsable principal", "Téléphone"]
        ws.append(headers)
        
        # Style pour les en-têtes
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Données des élèves
        for eleve in eleves:
            ws.append([
                eleve.classe.ecole.nom,
                eleve.classe.nom,
                eleve.matricule or '',
                f"{eleve.nom} {eleve.prenom}",
                eleve.get_sexe_display() if hasattr(eleve, 'get_sexe_display') else getattr(eleve, 'sexe', ''),
                eleve.date_naissance.strftime('%d/%m/%Y') if getattr(eleve, 'date_naissance', None) else '',
                eleve.responsable_principal.nom_complet if eleve.responsable_principal else '',
                eleve.responsable_principal.telephone if eleve.responsable_principal else '',
            ])

        # Largeur des colonnes optimisée
        widths = [20, 15, 15, 25, 10, 15, 25, 15]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        # Réponse HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = "tous_les_eleves.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du fichier Excel: {str(e)}", status=500)

@login_required
@require_http_methods(["POST"])
def supprimer_eleve(request, eleve_id):
    """Vue pour supprimer un élève (soft delete)"""
    qs = Eleve.objects.all()
    if not user_is_admin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')
    eleve = get_object_or_404(qs, id=eleve_id)
    
    # Vérifier s'il y a des paiements associés
    if eleve.paiements.exists():
        messages.error(request, "Impossible de supprimer cet élève car il a des paiements associés.")
        return redirect('eleves:detail_eleve', eleve_id=eleve.id)
    
    nom_complet = f"{eleve.prenom} {eleve.nom}"
    matricule = eleve.matricule
    
    # Soft delete - changer le statut au lieu de supprimer
    eleve.statut = 'EXCLU'
    eleve.save()
    
    # Créer l'historique
    HistoriqueEleve.objects.create(
        eleve=eleve,
        action='EXCLUSION',
        description=f"Exclusion de l'élève {nom_complet}",
        utilisateur=request.user
    )
    
    # Log de l'activité
    JournalActivite.objects.create(
        user=request.user,
        action='SUPPRESSION',
        type_objet='ELEVE',
        objet_id=eleve.id,
        description=f"Exclusion de l'élève {nom_complet} (matricule: {matricule})",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    messages.success(request, f"L'élève {nom_complet} a été exclu.")
    return redirect('eleves:liste_eleves')

@login_required
def gestion_classes(request):
    """Vue pour gérer les classes"""
    classes = Classe.objects.select_related('ecole').annotate(
        nombre_eleves=Count('eleves')
    ).order_by('ecole__nom', 'niveau', 'nom')
    if not user_is_admin(request.user):
        classes = classes.filter(ecole=user_school(request.user))
    
    # Statistiques
    stats = {
        'total_classes': classes.count(),
        'total_eleves': sum(c.nombre_eleves for c in classes),
        'classes_par_ecole': {}
    }
    
    ecoles_iter = Ecole.objects.all()
    if not user_is_admin(request.user):
        ecoles_iter = ecoles_iter.filter(id=user_school(request.user).id)
    for ecole in ecoles_iter:
        classes_ecole = classes.filter(ecole=ecole)
        stats['classes_par_ecole'][ecole.nom] = {
            'nombre_classes': classes_ecole.count(),
            'nombre_eleves': sum(c.nombre_eleves for c in classes_ecole)
        }
    
    context = {
        'classes': classes,
        'stats': stats,
        'titre_page': 'Gestion des Classes'
    }
    
    return render(request, 'eleves/gestion_classes.html', context)

@login_required
def ajax_classes_par_ecole(request, ecole_id):
    """Vue AJAX pour récupérer les classes d'une école"""
    try:
        # Non-admin: ne peut demander que sa propre école
        if not user_is_admin(request.user):
            if str(user_school(request.user).id) != str(ecole_id):
                return JsonResponse({'success': False, 'error': "Accès non autorisé à cette école."}, status=403)
        ecole = get_object_or_404(Ecole, id=ecole_id)
        classes = Classe.objects.filter(ecole=ecole).values('id', 'nom')
        return JsonResponse({
            'success': True,
            'classes': list(classes)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def ajax_statistiques_eleves(request):
    """Vue AJAX pour récupérer les statistiques des élèves"""
    try:
        if user_is_admin(request.user):
            eleves = Eleve.objects.all()
        else:
            eleves = Eleve.objects.filter(classe__ecole=user_school(request.user))
        stats = {
            'total_eleves': eleves.count(),
            'eleves_actifs': eleves.filter(statut='ACTIF').count(),
            'eleves_exclus': eleves.filter(statut='EXCLU').count(),
        }
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def statistiques_eleves(request):
    """Vue pour afficher les statistiques complètes des élèves"""
    from django.db.models import Sum, Avg, Max, Min
    from datetime import datetime, date
    from dateutil.relativedelta import relativedelta
    
    # 1. STATISTIQUES GÉNÉRALES
    if user_is_admin(request.user):
        eleves_base = Eleve.objects.all()
        classes_base = Classe.objects.all()
        responsables_base = Responsable.objects.all()
        ecoles_base = Ecole.objects.all()
    else:
        ecole_u = user_school(request.user)
        eleves_base = Eleve.objects.filter(classe__ecole=ecole_u)
        classes_base = Classe.objects.filter(ecole=ecole_u)
        responsables_base = Responsable.objects.all()
        ecoles_base = Ecole.objects.filter(id=ecole_u.id)
    total_eleves = eleves_base.count()
    stats_generales = {
        'total_eleves': total_eleves,
        'eleves_actifs': eleves_base.filter(statut='ACTIF').count(),
        'eleves_suspendus': eleves_base.filter(statut='SUSPENDU').count(),
        'eleves_exclus': eleves_base.filter(statut='EXCLU').count(),
        'eleves_transferes': eleves_base.filter(statut='TRANSFERE').count(),
        'eleves_diplomes': eleves_base.filter(statut='DIPLOME').count(),
        'total_ecoles': ecoles_base.count(),
        'total_classes': classes_base.count(),
        'total_responsables': responsables_base.count(),
    }
    
    # 2. STATISTIQUES DÉMOGRAPHIQUES
    stats_demographiques = {
        'garcons': eleves_base.filter(sexe='M').count(),
        'filles': eleves_base.filter(sexe='F').count(),
        'pourcentage_garcons': 0,
        'pourcentage_filles': 0,
    }
    
    if total_eleves > 0:
        stats_demographiques['pourcentage_garcons'] = round((stats_demographiques['garcons'] / total_eleves) * 100, 1)
        stats_demographiques['pourcentage_filles'] = round((stats_demographiques['filles'] / total_eleves) * 100, 1)
    
    # 3. STATISTIQUES D'ÂGE
    eleves_avec_age = eleves_base.exclude(date_naissance__isnull=True)
    ages = []
    for eleve in eleves_avec_age:
        age = relativedelta(date.today(), eleve.date_naissance).years
        ages.append(age)
    
    stats_age = {
        'age_moyen': round(sum(ages) / len(ages), 1) if ages else 0,
        'age_min': min(ages) if ages else 0,
        'age_max': max(ages) if ages else 0,
        'eleves_moins_10': len([a for a in ages if a < 10]),
        'eleves_10_15': len([a for a in ages if 10 <= a <= 15]),
        'eleves_plus_15': len([a for a in ages if a > 15]),
    }
    
    # 4. RÉPARTITION PAR ÉCOLE (détaillée)
    stats_par_ecole = []
    for ecole in ecoles_base:
        eleves_ecole = eleves_base.filter(classe__ecole=ecole)
        classes_ecole = classes_base.filter(ecole=ecole)
        
        ecole_stats = {
            'ecole': ecole,
            'total_eleves': eleves_ecole.count(),
            'eleves_actifs': eleves_ecole.filter(statut='ACTIF').count(),
            'garcons': eleves_ecole.filter(sexe='M').count(),
            'filles': eleves_ecole.filter(sexe='F').count(),
            'total_classes': classes_ecole.count(),
            'classes_actives': classes_ecole.filter(eleves__isnull=False).distinct().count(),
            'moyenne_eleves_par_classe': 0,
        }
        
        if ecole_stats['total_classes'] > 0:
            ecole_stats['moyenne_eleves_par_classe'] = round(ecole_stats['total_eleves'] / ecole_stats['total_classes'], 1)
        
        # Pourcentages pour cette école
        if ecole_stats['total_eleves'] > 0:
            ecole_stats['pourcentage_garcons'] = round((ecole_stats['garcons'] / ecole_stats['total_eleves']) * 100, 1)
            ecole_stats['pourcentage_filles'] = round((ecole_stats['filles'] / ecole_stats['total_eleves']) * 100, 1)
        else:
            ecole_stats['pourcentage_garcons'] = 0
            ecole_stats['pourcentage_filles'] = 0
        
        stats_par_ecole.append(ecole_stats)
    
    # 5. RÉPARTITION PAR NIVEAU (détaillée)
    stats_par_niveau = []
    total_pour_pourcentage = total_eleves if total_eleves > 0 else 1
    
    for niveau_code, niveau_nom in Classe.NIVEAUX_CHOICES:
        eleves_niveau = eleves_base.filter(classe__niveau=niveau_code)
        count = eleves_niveau.count()
        
        if count > 0:
            niveau_stats = {
                'niveau_code': niveau_code,
                'niveau_nom': niveau_nom,
                'total_eleves': count,
                'garcons': eleves_niveau.filter(sexe='M').count(),
                'filles': eleves_niveau.filter(sexe='F').count(),
                'actifs': eleves_niveau.filter(statut='ACTIF').count(),
                'pourcentage': round((count / total_pour_pourcentage) * 100, 1),
                'classes': Classe.objects.filter(niveau=niveau_code, eleves__isnull=False).distinct().count(),
            }
            stats_par_niveau.append(niveau_stats)
    
    # 6. STATISTIQUES PAR CLASSE (top 10)
    stats_par_classe = []
    classes_avec_eleves = classes_base.annotate(
        nb_eleves=Count('eleves')
    ).filter(nb_eleves__gt=0).order_by('-nb_eleves')[:10]
    
    for classe in classes_avec_eleves:
        eleves_classe = eleves_base.filter(classe=classe)
        classe_stats = {
            'classe': classe,
            'total_eleves': eleves_classe.count(),
            'garcons': eleves_classe.filter(sexe='M').count(),
            'filles': eleves_classe.filter(sexe='F').count(),
            'actifs': eleves_classe.filter(statut='ACTIF').count(),
        }
        stats_par_classe.append(classe_stats)
    
    # 7. STATISTIQUES TEMPORELLES
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    stats_temporelles = {
        'inscriptions_cette_annee': Eleve.objects.filter(date_inscription__year=current_year).count(),
        'inscriptions_ce_mois': Eleve.objects.filter(
            date_inscription__year=current_year,
            date_inscription__month=current_month
        ).count(),
        'inscriptions_cette_semaine': Eleve.objects.filter(
            date_inscription__gte=date.today() - relativedelta(days=7)
        ).count(),
    }
    
    # Évolution mensuelle (6 derniers mois)
    evolution_mensuelle = []
    for i in range(6):
        mois_date = date.today() - relativedelta(months=i)
        nb_inscriptions = Eleve.objects.filter(
            date_inscription__year=mois_date.year,
            date_inscription__month=mois_date.month
        ).count()
        
        evolution_mensuelle.append({
            'mois': mois_date.strftime('%B %Y'),
            'mois_court': mois_date.strftime('%b'),
            'inscriptions': nb_inscriptions
        })
    
    evolution_mensuelle.reverse()  # Du plus ancien au plus récent
    
    # 8. STATISTIQUES DE RESPONSABLES
    stats_responsables = {
        'total_responsables': responsables_base.count(),
        'responsables_principaux': eleves_base.values('responsable_principal').distinct().count(),
        'responsables_secondaires': eleves_base.filter(responsable_secondaire__isnull=False).values('responsable_secondaire').distinct().count(),
        'eleves_avec_deux_responsables': eleves_base.filter(responsable_secondaire__isnull=False).count(),
        'eleves_avec_un_responsable': eleves_base.filter(responsable_secondaire__isnull=True).count(),
    }
    
    # Répartition par relation
    relations_stats = []
    for relation_code, relation_nom in Responsable.RELATION_CHOICES:
        count = Responsable.objects.filter(relation=relation_code).count()
        if count > 0:
            relations_stats.append({
                'relation': relation_nom,
                'count': count,
                'pourcentage': round((count / stats_responsables['total_responsables']) * 100, 1) if stats_responsables['total_responsables'] > 0 else 0
            })
    
    # 9. STATISTIQUES FINANCIÈRES (basiques)
    from paiements.models import Paiement
    
    paiements_qs = Paiement.objects.all()
    if not user_is_admin(request.user):
        paiements_qs = paiements_qs.filter(eleve__classe__ecole=user_school(request.user))
    stats_financieres = {
        'eleves_avec_paiements': eleves_base.filter(paiements__isnull=False).distinct().count(),
        'eleves_sans_paiements': eleves_base.filter(paiements__isnull=True).count(),
        'total_paiements': paiements_qs.count(),
        'paiements_valides': paiements_qs.filter(statut='VALIDE').count(),
        'paiements_en_attente': paiements_qs.filter(statut='EN_ATTENTE').count(),
    }
    
    if stats_financieres['total_paiements'] > 0:
        stats_financieres['taux_validation'] = round(
            (stats_financieres['paiements_valides'] / stats_financieres['total_paiements']) * 100, 1
        )
    else:
        stats_financieres['taux_validation'] = 0
    
    # 10. INDICATEURS DE PERFORMANCE
    indicateurs = {
        'taux_activite': round((stats_generales['eleves_actifs'] / total_eleves) * 100, 1) if total_eleves > 0 else 0,
        'taux_retention': round(((total_eleves - stats_generales['eleves_exclus'] - stats_generales['eleves_transferes']) / total_eleves) * 100, 1) if total_eleves > 0 else 0,
        'ratio_eleves_classes': round(total_eleves / stats_generales['total_classes'], 1) if stats_generales['total_classes'] > 0 else 0,
        'ratio_eleves_responsables': round(total_eleves / stats_responsables['total_responsables'], 1) if stats_responsables['total_responsables'] > 0 else 0,
    }
    
    context = {
        'stats_generales': stats_generales,
        'stats_demographiques': stats_demographiques,
        'stats_age': stats_age,
        'stats_par_ecole': stats_par_ecole,
        'stats_par_niveau': stats_par_niveau,
        'stats_par_classe': stats_par_classe,
        'stats_temporelles': stats_temporelles,
        'evolution_mensuelle': evolution_mensuelle,
        'stats_responsables': stats_responsables,
        'relations_stats': relations_stats,
        'stats_financieres': stats_financieres,
        'indicateurs': indicateurs,
        'titre_page': 'Statistiques Complètes des Élèves'
    }
    
    return render(request, 'eleves/statistiques.html', context)

@login_required
def fiche_inscription_pdf(request, eleve_id):
    """Génère la fiche d'inscription d'un élève en PDF"""
    qs = Eleve.objects.select_related(
        'classe', 'classe__ecole', 'responsable_principal', 'responsable_secondaire'
    )
    if not user_is_admin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')
    eleve = get_object_or_404(qs, id=eleve_id)
    
    # Log de l'activité
    JournalActivite.objects.create(
        user=request.user,
        action='IMPRESSION',
        type_objet='ELEVE',
        objet_id=eleve.id,
        description=f"Impression fiche d'inscription PDF de {eleve.nom_complet}",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    # Créer la réponse HTTP pour le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="fiche_inscription_{eleve.matricule}.pdf"'
    
    # Créer le PDF
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
    # Configuration des polices avec détection cross-platform
    main_font_registered = False
    try:
        # Chemins possibles pour les polices selon l'OS
        font_paths = []
        
        # Windows
        if os.name == 'nt':
            font_paths.extend([
                ('C:/Windows/Fonts/calibri.ttf', 'C:/Windows/Fonts/calibrib.ttf'),
                ('C:/Windows/Fonts/arial.ttf', 'C:/Windows/Fonts/arialbd.ttf'),
            ])
        
        # Linux/Unix (PythonAnywhere, Ubuntu, etc.)
        else:
            font_paths.extend([
                ('/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf', 
                 '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf'),
                ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 
                 '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
                ('/System/Library/Fonts/Arial.ttf', '/System/Library/Fonts/Arial Bold.ttf'),  # macOS
            ])
        
        # Essayer chaque paire de polices
        for regular_path, bold_path in font_paths:
            if os.path.exists(regular_path) and os.path.exists(bold_path):
                try:
                    pdfmetrics.registerFont(TTFont('MainFont', regular_path))
                    pdfmetrics.registerFont(TTFont('MainFont-Bold', bold_path))
                    main_font_registered = True
                    break
                except Exception:
                    continue
        
        # Si aucune police système trouvée, utiliser les polices par défaut de ReportLab
        if not main_font_registered:
            # Les polices Helvetica sont déjà disponibles par défaut dans ReportLab
            # On n'a pas besoin de les enregistrer, juste de s'assurer qu'elles existent
            pass
            
    except Exception:
        # En cas d'erreur, s'assurer que les alias existent
        main_font_registered = False
    
    # Compression PDF pour meilleure qualité
    c.setPageCompression(1)
    
    # Filigrane avec logo de l'école (même taille que les autres exports PDF)
    c.saveState()
    try:
        # Chemin vers le logo
        logo_path = os.path.join('static', 'logos', 'logo.png')
        if os.path.exists(logo_path):
            # Taille ~150% de la largeur de page (comme dans rapports/utils.py)
            wm_width = width * 1.5
            wm_height = wm_width  # carré approximatif, preserveAspectRatio activera le ratio réel
            wm_x = (width - wm_width) / 2
            wm_y = (height - wm_height) / 2
            
            # Opacité visible mais discrète (comme dans les reçus de paiement)
            try:
                c.setFillAlpha(0.15)
            except Exception:
                pass
            
            # Légère rotation pour l'effet filigrane
            c.translate(width / 2.0, height / 2.0)
            c.rotate(30)
            c.translate(-width / 2.0, -height / 2.0)
            
            c.drawImage(logo_path, wm_x, wm_y, width=wm_width, height=wm_height, preserveAspectRatio=True, mask='auto')
        else:
            # Fallback vers texte si logo non trouvé
            c.setFillAlpha(0.04)
            try:
                c.setFont("MainFont-Bold", 60)
            except:
                c.setFont("Helvetica-Bold", 60)
            c.rotate(45)
            c.drawString(200, -100, eleve.classe.ecole.nom.upper())
            c.rotate(-45)
    finally:
        c.restoreState()
    
    # En-tête
    y = height - 2*cm
    try:
        c.setFont("MainFont-Bold", 16)
    except:
        c.setFont("Helvetica-Bold", 16)
    
    # Centrer le texte manuellement avec taille réduite
    text = eleve.classe.ecole.nom.upper()
    text_width = c.stringWidth(text, "MainFont-Bold", 16) if "MainFont-Bold" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica-Bold", 16)
    c.drawString((width - text_width) / 2, y, text)
    y -= 0.8*cm
    
    try:
        c.setFont("MainFont-Bold", 16)
    except:
        c.setFont("Helvetica-Bold", 16)
    
    text = "FICHE D'INSCRIPTION"
    text_width = c.stringWidth(text, "MainFont-Bold", 16) if "MainFont-Bold" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica-Bold", 16)
    c.drawString((width - text_width) / 2, y, text)
    
    y -= 1.5*cm
    
    # Photo de l'élève (si disponible)
    photo_x = width - 4*cm
    photo_y = y - 3*cm
    photo_width = 3*cm
    photo_height = 4*cm
    
    if eleve.photo:
        try:
            c.drawImage(eleve.photo.path, photo_x, photo_y, width=photo_width, height=photo_height)
        except Exception:
            # Placeholder si l'image ne peut pas être chargée
            c.rect(photo_x, photo_y, photo_width, photo_height)
            try:
                c.setFont("MainFont", 12)
            except:
                c.setFont("Helvetica", 12)
            text = "Photo"
            text_width = c.stringWidth(text, "MainFont", 12) if "MainFont" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica", 12)
            c.drawString(photo_x + (photo_width - text_width)/2, photo_y + photo_height/2, text)
    else:
        # Placeholder pour photo
        c.rect(photo_x, photo_y, photo_width, photo_height)
        try:
            c.setFont("MainFont", 12)
        except:
            c.setFont("Helvetica", 12)
        text = "Photo"
        text_width = c.stringWidth(text, "MainFont", 12) if "MainFont" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica", 12)
        c.drawString(photo_x + (photo_width - text_width)/2, photo_y + photo_height/2, text)
    
    # Informations de l'élève
    try:
        c.setFont("MainFont-Bold", 14)
    except:
        c.setFont("Helvetica-Bold", 14)
    
    # Section Informations personnelles
    c.drawString(2*cm, y, "INFORMATIONS PERSONNELLES")
    y -= 0.8*cm
    
    try:
        c.setFont("MainFont", 12)
    except:
        c.setFont("Helvetica", 12)
    
    # Colonne gauche
    left_col = 2*cm
    right_col = 10*cm
    line_height = 0.6*cm
    
    c.drawString(left_col, y, f"Matricule: {eleve.matricule}")
    c.drawString(right_col, y, f"Date d'inscription: {eleve.date_inscription.strftime('%d/%m/%Y')}")
    y -= line_height
    
    c.drawString(left_col, y, f"Nom: {eleve.nom}")
    c.drawString(right_col, y, f"Prénom: {eleve.prenom}")
    y -= line_height
    
    c.drawString(left_col, y, f"Sexe: {eleve.get_sexe_display()}")
    c.drawString(right_col, y, f"Date de naissance: {eleve.date_naissance.strftime('%d/%m/%Y')}")
    y -= line_height
    
    c.drawString(left_col, y, f"Lieu de naissance: {eleve.lieu_naissance}")
    c.drawString(right_col, y, f"Âge: {eleve.age} ans")
    y -= line_height
    
    c.drawString(left_col, y, f"Statut: {eleve.get_statut_display()}")
    y -= line_height * 1.5
    
    # Section Informations scolaires
    try:
        c.setFont("MainFont-Bold", 14)
    except:
        c.setFont("Helvetica-Bold", 14)
    
    c.drawString(left_col, y, "INFORMATIONS SCOLAIRES")
    y -= 0.8*cm
    
    try:
        c.setFont("MainFont", 12)
    except:
        c.setFont("Helvetica", 12)
    
    c.drawString(left_col, y, f"École: {eleve.classe.ecole.nom}")
    y -= line_height
    
    c.drawString(left_col, y, f"Classe: {eleve.classe.nom}")
    c.drawString(right_col, y, f"Niveau: {eleve.classe.get_niveau_display()}")
    y -= line_height
    
    c.drawString(left_col, y, f"Année scolaire: {eleve.classe.annee_scolaire}")
    y -= line_height * 1.5
    
    # Section Responsable principal
    try:
        c.setFont("MainFont-Bold", 14)
    except:
        c.setFont("Helvetica-Bold", 14)
    
    c.drawString(left_col, y, "RESPONSABLE PRINCIPAL")
    y -= 0.8*cm
    
    try:
        c.setFont("MainFont", 12)
    except:
        c.setFont("Helvetica", 12)
    
    resp_principal = eleve.responsable_principal
    c.drawString(left_col, y, f"Nom complet: {resp_principal.nom_complet}")
    c.drawString(right_col, y, f"Relation: {resp_principal.get_relation_display()}")
    y -= line_height
    
    c.drawString(left_col, y, f"Téléphone: {resp_principal.telephone}")
    if resp_principal.email:
        c.drawString(right_col, y, f"Email: {resp_principal.email}")
    y -= line_height
    
    if resp_principal.profession:
        c.drawString(left_col, y, f"Profession: {resp_principal.profession}")
        y -= line_height
    
    if resp_principal.adresse:
        c.drawString(left_col, y, f"Adresse: {resp_principal.adresse}")
        y -= line_height
    
    # Section Responsable secondaire (si existe)
    if eleve.responsable_secondaire:
        y -= line_height * 0.5
        
        try:
            c.setFont("MainFont-Bold", 14)
        except:
            c.setFont("Helvetica-Bold", 14)
        
        c.drawString(left_col, y, "RESPONSABLE SECONDAIRE")
        y -= 0.8*cm
        
        try:
            c.setFont("MainFont", 12)
        except:
            c.setFont("Helvetica", 12)
        
        resp_secondaire = eleve.responsable_secondaire
        c.drawString(left_col, y, f"Nom complet: {resp_secondaire.nom_complet}")
        c.drawString(right_col, y, f"Relation: {resp_secondaire.get_relation_display()}")
        y -= line_height
        
        c.drawString(left_col, y, f"Téléphone: {resp_secondaire.telephone}")
        if resp_secondaire.email:
            c.drawString(right_col, y, f"Email: {resp_secondaire.email}")
        y -= line_height
        
        if resp_secondaire.profession:
            c.drawString(left_col, y, f"Profession: {resp_secondaire.profession}")
            y -= line_height
        
        if resp_secondaire.adresse:
            c.drawString(left_col, y, f"Adresse: {resp_secondaire.adresse}")
            y -= line_height
    
    # Pied de page
    y = 3*cm
    try:
        c.setFont("MainFont", 10)
    except:
        c.setFont("Helvetica", 10)
    
    text = f"Fiche générée le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    text_width = c.stringWidth(text, "MainFont", 10) if "MainFont" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica", 10)
    c.drawString((width - text_width) / 2, y, text)
    y -= 0.5*cm
    
    text = "Système de Gestion Scolaire - École Moderne HADJA KANFING DIANÉ"
    text_width = c.stringWidth(text, "MainFont", 10) if "MainFont" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica", 10)
    c.drawString((width - text_width) / 2, y, text)
    
    # Finaliser le PDF
    c.save()
    
    return response
