from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from eleves.models import Classe
from utilisateurs.utils import filter_by_user_school, user_school
from ecole_moderne.security_decorators import admin_required, require_school_object
from .forms import ClasseNotesForm, MatiereClasseForm, EvaluationForm, NotesBulkForm
from .models import MatiereClasse, Evaluation, Note
from eleves.models import Eleve
from decimal import Decimal
from django.http import HttpResponse
import os
from datetime import datetime


# Groupes de niveaux pour l'affichage
PRIMAIRE = {
    'PRIMAIRE_1', 'PRIMAIRE_2', 'PRIMAIRE_3', 'PRIMAIRE_4', 'PRIMAIRE_5', 'PRIMAIRE_6'
}
COLLEGE = {
    'COLLEGE_7', 'COLLEGE_8', 'COLLEGE_9', 'COLLEGE_10'
}
LYCEE = {
    'LYCEE_11', 'LYCEE_12', 'TERMINALE'
}


def _draw_school_header(c, ecole, *, y_start, margin, page_width):
    """Dessine un en-tête officiel (centré) avec logo, nom en MAJUSCULES, coordonnées et encadré.
    Retourne la nouvelle coordonnée y après dessin."""
    from reportlab.lib import colors
    y = y_start
    # En-tête national
    center_x = page_width / 2
    c.setFont('Helvetica-Bold', 18)
    c.drawCentredString(center_x, y, "République de Guinée")
    y -= 12
    c.setFont('Helvetica-Oblique', 10)
    # Dessiner la devise avec couleurs par mot: Travail (rouge), Justice (jaune), Solidarité (vert)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.lib import colors
    parts = [
        ("Travail", colors.red),
        (" - ", colors.black),
        ("Justice", colors.yellow),
        (" - ", colors.black),
        ("Solidarité", colors.green),
    ]
    total_w = sum(pdfmetrics.stringWidth(t, 'Helvetica-Oblique', 10) for t, _ in parts)
    start_x = center_x - (total_w / 2)
    x = start_x
    for text, col in parts:
        c.setFillColor(col)
        c.drawString(x, y, text)
        x += pdfmetrics.stringWidth(text, 'Helvetica-Oblique', 10)
    c.setFillColor(colors.black)
    y -= 12
    c.setFont('Helvetica', 10)
    c.drawCentredString(center_x, y, "Ministère de l’Enseignement Pré-Universitaire et de l’Alphabétisation")
    y -= 12
    # Abréviations sur 3 lignes (centrées)
    c.setFont('Helvetica-Bold', 10)
    y -= 6
    c.drawCentredString(center_x, y, "IRE:")
    y -= 12
    c.drawCentredString(center_x, y, "DPE:")
    y -= 12
    c.drawCentredString(center_x, y, "DESEE:")
    y -= 16
    # Espace supplémentaire pour descendre le premier cadre du bulletin
    y -= 30

    # Mémoriser la position du haut du cadre pour le dessiner après le contenu
    frame_top = y
    box_height = 60

    # Logo (gauche) si disponible
    logo_path = None
    try:
        if hasattr(ecole, 'logo') and getattr(ecole.logo, 'path', None) and os.path.exists(ecole.logo.path):
            logo_path = ecole.logo.path
    except Exception:
        logo_path = None
    if logo_path:
        try:
            c.drawImage(logo_path, margin + 8, y - 62, width=54, height=54, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    # Texte centré
    top_line_y = y + 12
    school_name = (getattr(ecole, 'nom', '') or 'ÉCOLE').upper()
    c.setFont('Helvetica-Bold', 18)
    c.drawCentredString(center_x, top_line_y, school_name)

    c.setFont('Helvetica', 10)
    adresse = getattr(ecole, 'adresse', None) or ''
    telephone = getattr(ecole, 'telephone', None) or ''
    email = getattr(ecole, 'email', None) or ''
    directeur = getattr(ecole, 'directeur', None) or ''

    # Helper: wrap centered text within available width
    from reportlab.pdfbase import pdfmetrics
    def draw_wrapped_centered(text, y_pos, max_width, line_height=12):
        words = text.split()
        lines = []
        cur = ''
        for w in words:
            test = (cur + ' ' + w).strip()
            if pdfmetrics.stringWidth(test, 'Helvetica', 10) <= max_width:
                cur = test
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        for ln in lines:
            c.drawCentredString(center_x, y_pos, ln)
            y_pos -= line_height
        return y_pos

    # Centrer les informations (adresse/contacts/directeur) au milieu du cadre
    line_y = top_line_y - 30
    if adresse:
        # keep inside box: reduce available width a bit
        avail_w = page_width - 2*margin - 20
        line_y = draw_wrapped_centered(f"Adresse: {adresse}", line_y, avail_w)
    contacts = []
    if telephone:
        contacts.append(f"Tél: {telephone}")
    if email:
        contacts.append(f"Email: {email}")
    if contacts:
        avail_w = page_width - 2*margin - 20
        line_y = draw_wrapped_centered("  |  ".join(contacts), line_y, avail_w)
    if directeur:
        c.drawCentredString(center_x, line_y, f"Directeur: {directeur}")

    # Dessiner le cadre maintenant que le contenu est placé
    # Le cadre commence au-dessus du nom de l'école pour avoir plus d'espace en haut
    frame_start_y = top_line_y - 8  # En dessous du nom de l'école
    adjusted_box_height = box_height  # Utiliser la hauteur fixe définie
    c.setLineWidth(1)
    c.setStrokeColor(colors.black)
    c.roundRect(margin, frame_start_y - adjusted_box_height, page_width - 2*margin, adjusted_box_height, 6, stroke=1, fill=0)

    # Retourner y en dessous du cadre
    y = y - box_height - 8
    # Ligne séparatrice légère
    c.setFillColor(colors.grey)
    c.rect(margin, y, page_width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 8
    return y

@login_required
def tableau_bord(request):
    """Tableau de bord des notes: liste les classes par groupe de niveaux.
    Filtré par l'école de l'utilisateur (sauf admin).
    """
    classes_qs = filter_by_user_school(Classe.objects.all().order_by('niveau', 'nom'), request.user, 'ecole')

    def group_classes(qs):
        primaire, college, lycee = [], [], []
        for c in qs:
            if c.niveau in PRIMAIRE:
                primaire.append(c)
            elif c.niveau in COLLEGE:
                college.append(c)
            elif c.niveau in LYCEE:
                lycee.append(c)
        return primaire, college, lycee

    primaire, college, lycee = group_classes(classes_qs)

    context = {
        'classes_primaire': primaire,
        'classes_college': college,
        'classes_lycee': lycee,
    }
    return render(request, 'notes/dashboard.html', context)


@admin_required
def creer_classe(request, niveau):
    """Créer une classe pour un niveau donné dans l'école de l'utilisateur."""
    if request.method == 'POST':
        form = ClasseNotesForm(request.POST, niveau_initial=niveau)
        if form.is_valid():
            classe = form.save(commit=False)
            classe.ecole = user_school(request.user)
            if classe.ecole is None:
                messages.error(request, "Aucune école associée à votre compte.")
                return redirect('notes:tableau_bord')
            classe.save()
            messages.success(request, f"Classe '{classe.nom}' créée avec succès.")
            return redirect('notes:tableau_bord')
    else:
        form = ClasseNotesForm(niveau_initial=niveau)
    return render(request, 'notes/classe_form.html', {'form': form, 'niveau': niveau})


@admin_required
def supprimer_classe(request, classe_id):
    """Supprimer une classe si elle appartient à l'école de l'utilisateur et qu'elle est vide."""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    if request.method == 'POST':
        if hasattr(classe, 'eleves') and classe.eleves.exists():
            messages.error(request, "Impossible de supprimer une classe qui contient des élèves.")
            return redirect('notes:tableau_bord')
        classe.delete()
        messages.success(request, "Classe supprimée avec succès.")
        return redirect('notes:tableau_bord')
    return render(request, 'notes/confirm_delete.html', {
        'objet': classe,
        'message': "Confirmez-vous la suppression de cette classe ? (Cette action est irréversible)",
        'action_url': reverse('notes:supprimer_classe', args=[classe.id])
    })


@admin_required
def matieres_classe(request, classe_id):
    """Liste et gestion des matières d'une classe."""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    matieres = MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole).order_by('nom')
    return render(request, 'notes/matieres_classe.html', {
        'classe': classe,
        'matieres': matieres,
    })


@admin_required
def creer_matiere(request, classe_id):
    """Créer une matière pour une classe donnée."""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    if request.method == 'POST':
        form = MatiereClasseForm(request.POST)
        if form.is_valid():
            mat = form.save(commit=False)
            mat.classe = classe
            mat.ecole = classe.ecole
            try:
                mat.save()
                messages.success(request, f"Matière '{mat.nom}' ajoutée.")
                return redirect('notes:matieres_classe', classe.id)
            except Exception as e:
                messages.error(request, f"Erreur lors de la création: {e}")
    else:
        form = MatiereClasseForm()
    return render(request, 'notes/matiere_form.html', {'form': form, 'classe': classe})


@admin_required
def supprimer_matiere(request, pk):
    """Supprimer une matière de classe."""
    matiere = get_object_or_404(filter_by_user_school(MatiereClasse.objects.select_related('classe', 'ecole'), request.user, 'ecole'), pk=pk)
    if request.method == 'POST':
        classe_id = matiere.classe_id
        matiere.delete()
        messages.success(request, "Matière supprimée.")
        return redirect('notes:matieres_classe', classe_id)
    return render(request, 'notes/confirm_delete.html', {
        'objet': matiere,
        'message': "Confirmez-vous la suppression de cette matière ?",
        'action_url': reverse('notes:supprimer_matiere', args=[matiere.id])
    })


@admin_required
def creer_evaluation(request, classe_id, matiere_id):
    """Créer une évaluation pour une classe/matière donnée."""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    matiere = get_object_or_404(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole), pk=matiere_id)
    if request.method == 'POST':
        form = EvaluationForm(request.POST)
        if form.is_valid():
            ev = form.save(commit=False)
            ev.ecole = classe.ecole
            ev.classe = classe
            ev.matiere = matiere
            ev.annee_scolaire = getattr(classe, 'annee_scolaire', None)
            ev.cree_par = request.user
            ev.save()
            messages.success(request, f"Évaluation '{ev.titre}' créée pour {classe.nom} — {matiere.nom}.")
            return redirect('notes:saisie_notes', evaluation_id=ev.id)
    else:
        form = EvaluationForm()
    return render(request, 'notes/evaluation_form.html', {
        'form': form,
        'classe': classe,
        'matiere': matiere,
    })


@admin_required
def saisie_notes(request, evaluation_id):
    """Saisie en masse des notes par matricule pour une évaluation.
    Format par ligne: MATRICULE;NOTE
    """
    # Récupération évaluation dans le périmètre école
    evaluation = get_object_or_404(
        filter_by_user_school(Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'),
        pk=evaluation_id
    )
    if request.method == 'POST':
        form = NotesBulkForm(request.POST)
        if form.is_valid():
            donnees = form.cleaned_data['donnees']
            lignes = [l.strip() for l in donnees.splitlines() if l.strip()]
            ok, erreurs, maj, crees = 0, [], 0, 0
            # Restreindre aux élèves de la classe + école de l'évaluation
            eleves_qs = Eleve.objects.select_related('classe', 'classe__ecole').filter(classe=evaluation.classe)
            eleves_qs = filter_by_user_school(eleves_qs, request.user, 'classe__ecole')
            # Index par matricule (upper)
            index_mat = { (e.matricule or '').strip().upper(): e for e in eleves_qs }
            for i, ligne in enumerate(lignes, start=1):
                parts = [p.strip() for p in ligne.split(';')]
                if len(parts) < 2:
                    erreurs.append(f"Ligne {i}: format invalide (attendu MATRICULE;NOTE)")
                    continue
                matricule, note_txt = parts[0].upper(), parts[1].replace(',', '.').strip()
                obs = ''  # Pas d'observation requise
                if matricule not in index_mat:
                    erreurs.append(f"Ligne {i}: matricule inconnu pour la classe ({matricule})")
                    continue
                # Parse note 0..20
                try:
                    val = Decimal(note_txt)
                except Exception:
                    erreurs.append(f"Ligne {i}: note invalide '{note_txt}'")
                    continue
                if val < 0 or val > 20:
                    erreurs.append(f"Ligne {i}: la note doit être entre 0 et 20 (reçu {val})")
                    continue
                eleve = index_mat[matricule]
                # Créer/mettre à jour la note
                obj, created = Note.objects.update_or_create(
                    evaluation=evaluation,
                    eleve=eleve,
                    defaults={
                        'ecole': evaluation.ecole,
                        'classe': evaluation.classe,
                        'matiere': evaluation.matiere,
                        'matricule': eleve.matricule or matricule,
                        'note': val,
                        'observation': obs or None,
                        'saisie_par': request.user,
                    }
                )
                if created:
                    crees += 1
                else:
                    maj += 1
                ok += 1
            if ok:
                messages.success(request, f"{ok} note(s) traitée(s) — {crees} créée(s), {maj} mise(s) à jour.")
            if erreurs:
                messages.warning(request, "\n".join(erreurs[:10]) + ("\n…" if len(erreurs) > 10 else ''))
            return redirect('notes:saisie_notes', evaluation_id=evaluation.id)
    else:
        form = NotesBulkForm()
    # Préparer un export des élèves de la classe avec matricules pour aide à la saisie
    eleves = Eleve.objects.select_related('classe').filter(classe=evaluation.classe).order_by('nom', 'prenom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    # Notes existantes pour cette évaluation
    notes_existantes = evaluation.notes.select_related('eleve').order_by('eleve__nom', 'eleve__prenom')
    return render(request, 'notes/saisie_notes.html', {
        'evaluation': evaluation,
        'form': form,
        'eleves': eleves,
        'notes_existantes': notes_existantes,
    })


@admin_required
def evaluations_matiere(request, classe_id, matiere_id):
    """Liste des évaluations d'une matière pour une classe, avec accès rapide à la saisie et à l'affichage des notes."""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    matiere = get_object_or_404(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole), pk=matiere_id)
    evaluations = (
        Evaluation.objects.filter(classe=classe, matiere=matiere)
        .order_by('-date', '-id')
    )
    return render(request, 'notes/evaluations_matiere.html', {
        'classe': classe,
        'matiere': matiere,
        'evaluations': evaluations,
    })


@admin_required
def evaluation_detail(request, evaluation_id):
    """Affiche un tableau des élèves de la classe avec leurs notes (ou vide si non saisie)."""
    evaluation = get_object_or_404(
        filter_by_user_school(Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'),
        pk=evaluation_id
    )
    # Élèves de la classe
    eleves = Eleve.objects.select_related('classe').filter(classe=evaluation.classe).order_by('nom', 'prenom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    # Index des notes
    notes_map = {n.eleve_id: n for n in evaluation.notes.select_related('eleve')}
    rows = []
    for e in eleves:
        n = notes_map.get(e.id)
        rows.append({
            'eleve': e,
            'matricule': e.matricule,
            'note': getattr(n, 'note', None),
            'observation': getattr(n, 'observation', ''),
        })
    return render(request, 'notes/evaluation_detail.html', {
        'evaluation': evaluation,
        'rows': rows,
    })


@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_pdf(request, classe_id: int, eleve_id: int, trimestre: str = "T1"):
    """Génère un bulletin de notes PDF pour un élève donné et un trimestre (T1/T2/T3).
    Calcule la moyenne par matière (pondérée par coefficient d'évaluation) et la moyenne générale (pondérée par coefficient de matière).
    """
    # Sécuriser l'accès à la classe / élève
    classe = get_object_or_404(filter_by_user_school(Classe.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleve = get_object_or_404(filter_by_user_school(Eleve.objects.select_related('classe', 'classe__ecole'), request.user, 'classe__ecole'), pk=eleve_id, classe=classe)

    # Matières définies pour la classe
    matieres = MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom')

    # Récupérer les évaluations de la classe/matière pour le trimestre
    evals_by_matiere = {}
    for mat in matieres:
        evals = Evaluation.objects.filter(classe=classe, matiere=mat, trimestre=trimestre).order_by('date', 'id')
        evals_by_matiere[mat.id] = list(evals)

    # Récupérer les notes de l'élève pour ces évaluations
    notes_by_eval = {n.evaluation_id: n for n in Note.objects.filter(eleve=eleve, evaluation__classe=classe, evaluation__trimestre=trimestre).select_related('evaluation', 'evaluation__matiere')}

    # Calculs des moyennes par matière
    lignes = []
    somme_moyennes_coef = Decimal('0')
    somme_coef_matieres = Decimal('0')

    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        if not evals:
            moy_mat = None
        else:
            num = Decimal('0')
            den = Decimal('0')
            for ev in evals:
                n = notes_by_eval.get(ev.id)
                if n is None or n.note is None:
                    continue
                c = Decimal(ev.coefficient or 1)
                num += Decimal(n.note) * c
                den += c
            moy_mat = (num / den).quantize(Decimal('0.01')) if den > 0 else None

        if moy_mat is not None:
            somme_moyennes_coef += moy_mat * Decimal(mat.coefficient or 1)
            somme_coef_matieres += Decimal(mat.coefficient or 1)

        lignes.append({
            'matiere': mat.nom,
            'coef_matiere': mat.coefficient,
            'moyenne': moy_mat,
        })

    moyenne_generale = None
    if somme_coef_matieres > 0:
        moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01'))

    # Moyennes de classe par matière (pondérées par coeffs d'évaluations)
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        total_num = Decimal('0'); total_den = Decimal('0')
        for ev in evals:
            # toutes les notes de l'évaluation pour la classe
            for n in Note.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                total_num += Decimal(n.note) * cc
                total_den += cc
        moyennes_classe_par_matiere[mat.id] = (total_num / total_den).quantize(Decimal('0.01')) if total_den > 0 else None

    # Classement (rang): calculer la moyenne générale de tous les élèves
    eleves_classe = Eleve.objects.filter(classe=classe).only('id')
    moyennes_generales = []  # list of (eleve_id, moyenne_generale)
    for e in eleves_classe:
        notes_by_eval_e = {n.evaluation_id: n for n in Note.objects.filter(eleve=e, evaluation__classe=classe, evaluation__trimestre=trimestre)}
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            # moyenne matière élève
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_by_eval_e.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            if den > 0:
                moy_mat_e = (num / den)
                s_num += moy_mat_e * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        if s_den > 0:
            mg = (s_num / s_den)
        else:
            mg = None
        if mg is not None:
            moyennes_generales.append((e.id, mg))
    # Trier desc (meilleure note première), calcul rang de l'élève
    moyennes_generales.sort(key=lambda t: t[1], reverse=True)
    rang = None
    total_eleves_ayant_moyenne = len(moyennes_generales)
    for idx, (eid, mg) in enumerate(moyennes_generales, start=1):
        if eid == eleve.id:
            rang = idx
            break

    # Mention selon barème simple (modifiable)
    def mention_for(avg: Decimal | None) -> str:
        if avg is None:
            return ""
        if avg >= Decimal('16'):
            return "Très Bien"
        if avg >= Decimal('14'):
            return "Bien"
        if avg >= Decimal('12'):
            return "Assez Bien"
        if avg >= Decimal('10'):
            return "Passable"
        return "Insuffisant"
    mention = mention_for(moyenne_generale)

    # Génération PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletin_{eleve.matricule}_{trimestre}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Filigrane standard si disponible
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
    except Exception:
        pass

    margin = 2 * cm
    y = height - margin

    # En-tête avec logo et coordonnées
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin de notes — {trimestre}"); y -= 40
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Élève: {eleve.nom} {eleve.prenom}  (Matricule: {eleve.matricule or '-'} )")
    y -= 16
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}")
    y -= 12
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 16

    # Tableau entêtes
    c.setFont('Helvetica-Bold', 12)
    headers = ["Matière", "Coef.", "Moyenne /20", "Moy. classe"]
    colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
    x = margin
    for i, htxt in enumerate(headers):
        c.drawString(x, y, htxt)
        x += colw[i]
    y -= 14
    c.setFillColor(colors.lightgrey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 10

    c.setFont('Helvetica', 11)
    for row in lignes:
        if y < margin + 60:
            c.showPage()
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
            except Exception:
                pass
            y = height - margin
        x = margin
        c.drawString(x, y, str(row['matiere'])); x += colw[0]
        c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
        moy_txt = '-' if row['moyenne'] is None else f"{row['moyenne']}"
        c.drawString(x, y, moy_txt); x += colw[2]
        # moyenne de classe
        mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
        mc_txt = '-' if mc is None else f"{mc}"
        c.drawString(x, y, mc_txt)
        y -= 14

    # Séparateur
    y -= 6
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 16

    # Moyenne générale + Rang + Mention
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margin, y, f"Moyenne générale: {moyenne_generale if moyenne_generale is not None else '-'} / 20")
    y -= 16
    c.setFont('Helvetica', 12)
    if rang is not None:
        c.drawString(margin, y, f"Rang: {rang} / {total_eleves_ayant_moyenne}")
        y -= 14
    if mention:
        c.drawString(margin, y, f"Mention: {mention}")
        y -= 18

    # Pied de page
    # Signatures
    c.setFont('Helvetica', 11)
    sig_y = margin + 50
    c.drawString(margin, sig_y, "Prof. principal:")
    c.line(margin + 120, sig_y-2, margin + 250, sig_y-2)
    c.drawString(margin + 280, sig_y, "Chef d'établ.:")
    c.line(margin + 380, sig_y-2, margin + 510, sig_y-2)
    c.drawString(margin, sig_y - 28, "Parent/Tuteur:")
    c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)

    c.setFont('Helvetica-Oblique', 10)
    c.setFillColor(colors.darkgrey)
    from datetime import datetime
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage(); c.save()
    return response


@admin_required
def bulletins_classe_pdf(request, classe_id: int, trimestre: str = "T1"):
    """Génère en un seul PDF les bulletins de tous les élèves d'une classe pour un trimestre."""
    # Sécuriser la classe
    classe = get_object_or_404(filter_by_user_school(Classe.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    # Élèves de la classe (dans le périmètre utilisateur)
    eleves = Eleve.objects.select_related('classe').filter(classe=classe).order_by('nom', 'prenom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')

    # Précharger matières et évaluations du trimestre
    matieres = list(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom'))
    evals_by_matiere = {}
    for mat in matieres:
        evals_by_matiere[mat.id] = list(Evaluation.objects.filter(classe=classe, matiere=mat, trimestre=trimestre).order_by('date', 'id'))

    # Init PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletins_classe_{classe.nom}_{trimestre}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Pré-calcul des moyennes de classe par matière
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        total_num = Decimal('0'); total_den = Decimal('0')
        for ev in evals:
            for n in Note.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                total_num += Decimal(n.note) * cc
                total_den += cc
        moyennes_classe_par_matiere[mat.id] = (total_num / total_den).quantize(Decimal('0.01')) if total_den > 0 else None

    # Pré-calcul des moyennes générales par élève (pour classement/rang)
    moyenne_generale_map: dict[int, Decimal] = {}
    for e in eleves:
        notes_by_eval_e = {n.evaluation_id: n for n in Note.objects.filter(eleve=e, evaluation__classe=classe, evaluation__trimestre=trimestre)}
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_by_eval_e.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            if den > 0:
                moy_mat_e = (num / den)
                s_num += moy_mat_e * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        if s_den > 0:
            moyenne_generale_map[e.id] = (s_num / s_den).quantize(Decimal('0.01'))

    # Classement
    classement = sorted(moyenne_generale_map.items(), key=lambda t: t[1], reverse=True)
    rang_map: dict[int, int] = {eid: idx for idx, (eid, _) in enumerate(classement, start=1)}
    total_eleves_ayant_moyenne = len(classement)

    def mention_for(avg: Decimal | None) -> str:
        if avg is None:
            return ""
        if avg >= Decimal('16'):
            return "Très Bien"
        if avg >= Decimal('14'):
            return "Bien"
        if avg >= Decimal('12'):
            return "Assez Bien"
        if avg >= Decimal('10'):
            return "Passable"
        return "Insuffisant"

    def draw_bulletin_for_student(eleve):
        # Calcul des moyennes pour l'élève
        notes_by_eval = {n.evaluation_id: n for n in Note.objects.filter(eleve=eleve, evaluation__classe=classe, evaluation__trimestre=trimestre).select_related('evaluation', 'evaluation__matiere')}
        lignes = []
        somme_moyennes_coef = Decimal('0')
        somme_coef_matieres = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            if not evals:
                moy_mat = None
            else:
                num = Decimal('0'); den = Decimal('0')
                for ev in evals:
                    n = notes_by_eval.get(ev.id)
                    if n is None or n.note is None:
                        continue
                    cc = Decimal(ev.coefficient or 1)
                    num += Decimal(n.note) * cc
                    den += cc
                moy_mat = (num / den).quantize(Decimal('0.01')) if den > 0 else None
            if moy_mat is not None:
                somme_moyennes_coef += moy_mat * Decimal(mat.coefficient or 1)
                somme_coef_matieres += Decimal(mat.coefficient or 1)
            lignes.append({
                'matiere': mat.nom,
                'coef_matiere': mat.coefficient,
                'moyenne': moy_mat,
            })
        moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

        # Dessiner la page
        try:
            from ecole_moderne.pdf_utils import draw_logo_watermark
            draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
        except Exception:
            pass

        margin = 2 * cm
        y = height - margin
        if getattr(classe, 'ecole', None):
            y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
        y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin de notes — {trimestre}"); y -= 40
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Élève: {eleve.nom} {eleve.prenom}  (Matricule: {eleve.matricule or '-'} )")
        y -= 16
        c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}")
        y -= 12
        c.setFillColor(colors.grey)
        c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
        c.setFillColor(colors.black)
        y -= 16

        c.setFont('Helvetica-Bold', 12)
        headers = ["Matière", "Coef.", "Moyenne /20", "Moy. classe"]
        colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
        x = margin
        for i, htxt in enumerate(headers):
            c.drawString(x, y, htxt); x += colw[i]
        y -= 14
        c.setFillColor(colors.lightgrey)
        c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
        c.setFillColor(colors.black)
        y -= 10
        c.setFont('Helvetica', 11)
        for row in lignes:
            if y < margin + 60:
                c.showPage()
                try:
                    from ecole_moderne.pdf_utils import draw_logo_watermark
                    draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
                except Exception:
                    pass
                y = height - margin
            x = margin
            c.drawString(x, y, str(row['matiere'])); x += colw[0]
            c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
            moy_txt = '-' if row['moyenne'] is None else f"{row['moyenne']}"
            c.drawString(x, y, moy_txt); x += colw[2]
            mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
            mc_txt = '-' if mc is None else f"{mc}"
            c.drawString(x, y, mc_txt)
            y -= 14

        y -= 6
        c.setFillColor(colors.grey)
        c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
        c.setFillColor(colors.black)
        y -= 16
        c.setFont('Helvetica-Bold', 13)
        c.drawString(margin, y, f"Moyenne générale: {moyenne_generale if moyenne_generale is not None else '-'} / 20")
        y -= 16
        # Rang + Mention
        rg = rang_map.get(eleve.id)
        if rg is not None:
            c.setFont('Helvetica', 12)
            c.drawString(margin, y, f"Rang: {rg} / {total_eleves_ayant_moyenne}")
            y -= 14
        men = mention_for(moyenne_generale)
        if men:
            c.setFont('Helvetica', 12)
            c.drawString(margin, y, f"Mention: {men}")
            y -= 16
        # Signatures
        c.setFont('Helvetica', 11)
        sig_y = margin + 50
        c.drawString(margin, sig_y, "Professeur principal:")
        c.line(margin + 150, sig_y-2, margin + 320, sig_y-2)
        c.drawString(margin + 350, sig_y, "Chef d’établissement:")
        c.line(margin + 520, sig_y-2, margin + 700, sig_y-2)
        c.drawString(margin, sig_y - 28, "Parent/Tuteur:")
        c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)
        c.showPage()

    # Dessiner pour chaque élève
    for e in eleves:
        draw_bulletin_for_student(e)

    c.save()
    return response


@admin_required
def export_notes_excel(request, classe_id: int, matiere_id: int, trimestre: str = "T1"):
    """Export Excel des notes d'une classe pour une matière et un trimestre.
    Colonnes: Matricule, Élève, [colonnes de chaque évaluation], Moyenne matière.
    """
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    matiere = get_object_or_404(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole), pk=matiere_id)

    # Évaluations du trimestre pour cette matière
    evaluations = list(Evaluation.objects.filter(classe=classe, matiere=matiere, trimestre=trimestre).order_by('date', 'id'))
    # Élèves
    eleves = Eleve.objects.filter(classe=classe).order_by('nom', 'prenom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')

    # Notes indexées par (eleve_id, evaluation_id)
    notes = Note.objects.filter(evaluation__in=evaluations, eleve__in=eleves)
    notes_map = {(n.eleve_id, n.evaluation_id): n for n in notes}

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse("openpyxl requis (pip install openpyxl)", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{matiere.nom} {trimestre}"

    # En-tête
    headers = ["Matricule", "Élève"] + [ev.titre or f"Eval {i+1}" for i, ev in enumerate(evaluations)] + ["Moyenne /20"]
    ws.append(headers)

    # Lignes
    from decimal import Decimal as D
    for e in eleves:
        row = [e.matricule, f"{e.nom} {e.prenom}"]
        num = D('0'); den = D('0')
        for ev in evaluations:
            n = notes_map.get((e.id, ev.id))
            if n and n.note is not None:
                row.append(float(n.note))
                c = D(ev.coefficient or 1)
                num += D(n.note) * c
                den += c
            else:
                row.append(None)
        moy = float((num/den)) if den > 0 else None
        row.append(moy)
        ws.append(row)

    # Styles simples: largeur colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col <= 2 else 12

    # Réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"notes_{classe.nom}_{matiere.nom}_{trimestre}.xlsx".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _collect_evals_all_trimestres(classe, matieres):
    """Retourne un dict {matiere_id: [evaluations sur T1+T2+T3]} triées par date."""
    evals_by_matiere = {}
    for mat in matieres:
        evals_by_matiere[mat.id] = list(Evaluation.objects.filter(classe=classe, matiere=mat, trimestre__in=["T1", "T2", "T3"]).order_by('date', 'id'))
    return evals_by_matiere


@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_annuel_pdf(request, classe_id: int, eleve_id: int):
    """Bulletin annuel PDF (T1+T2+T3 cumulés) avec moyennes par matière, moyenne générale, rang, mention, signatures."""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleve = get_object_or_404(filter_by_user_school(Eleve.objects.select_related('classe', 'classe__ecole'), request.user, 'classe__ecole'), pk=eleve_id, classe=classe)

    matieres = list(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom'))
    evals_by_matiere = _collect_evals_all_trimestres(classe, matieres)

    # Calculs élève
    lignes = []
    somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
    notes_by_eval = {n.evaluation_id: n for n in Note.objects.filter(eleve=eleve, evaluation__classe=classe, evaluation__trimestre__in=["T1","T2","T3"])}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        num = Decimal('0'); den = Decimal('0')
        for ev in evals:
            n = notes_by_eval.get(ev.id)
            if not n or n.note is None:
                continue
            cc = Decimal(ev.coefficient or 1)
            num += Decimal(n.note) * cc
            den += cc
        moy_mat = (num/den).quantize(Decimal('0.01')) if den > 0 else None
        if moy_mat is not None:
            somme_moyennes_coef += moy_mat * Decimal(mat.coefficient or 1)
            somme_coef_matieres += Decimal(mat.coefficient or 1)
        lignes.append({'matiere': mat.nom, 'coef_matiere': mat.coefficient, 'moyenne': moy_mat})
    moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

    # Moyennes de classe par matière
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        tnum = Decimal('0'); tden = Decimal('0')
        for ev in evals:
            for n in Note.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                tnum += Decimal(n.note) * cc
                tden += cc
        moyennes_classe_par_matiere[mat.id] = (tnum/tden).quantize(Decimal('0.01')) if tden > 0 else None

    # Classement annuel
    eleves = filter_by_user_school(Eleve.objects.filter(classe=classe), request.user, 'classe__ecole')
    moyenne_generale_map: dict[int, Decimal] = {}
    for e in eleves:
        notes_e = {n.evaluation_id: n for n in Note.objects.filter(eleve=e, evaluation__classe=classe, evaluation__trimestre__in=["T1","T2","T3"])}
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_e.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            if den > 0:
                moy_mat_e = (num/den)
                s_num += moy_mat_e * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        if s_den > 0:
            moyenne_generale_map[e.id] = (s_num / s_den).quantize(Decimal('0.01'))

    classement = sorted(moyenne_generale_map.items(), key=lambda t: t[1], reverse=True)
    rang_map: dict[int, int] = {eid: idx for idx, (eid, _) in enumerate(classement, start=1)}
    rang = rang_map.get(eleve.id)

    def mention_for(avg: Decimal | None) -> str:
        if avg is None:
            return ""
        if avg >= Decimal('16'): return "Très Bien"
        if avg >= Decimal('14'): return "Bien"
        if avg >= Decimal('12'): return "Assez Bien"
        if avg >= Decimal('10'): return "Passable"
        return "Insuffisant"
    mention = mention_for(moyenne_generale)

    # PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletin_annuel_{eleve.matricule}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
    except Exception:
        pass
    margin = 2*cm; y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, "Bulletin de notes — Annuel"); y -= 40
    c.setFont('Helvetica', 12); c.drawString(margin, y, f"Élève: {eleve.nom} {eleve.prenom} (Matricule: {eleve.matricule or '-'})"); y -= 16
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}"); y -= 12
    c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    c.setFont('Helvetica-Bold', 12)
    headers = ["Matière", "Coef.", "Moy. annuelle", "Moy. classe"]
    colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
    x = margin
    for i, h in enumerate(headers): c.drawString(x, y, h); x += colw[i]
    y -= 14; c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10
    c.setFont('Helvetica', 11)
    for row in lignes:
        if y < margin + 60:
            c.showPage();
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
            except Exception:
                pass
            y = height - margin
        x = margin
        c.drawString(x, y, row['matiere']); x += colw[0]
        c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
        c.drawString(x, y, '-' if row['moyenne'] is None else f"{row['moyenne']}"); x += colw[2]
        mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
        c.drawString(x, y, '-' if mc is None else f"{mc}")
        y -= 14

    y -= 6
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 16
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margin, y, f"Moyenne générale annuelle: {moyenne_generale if moyenne_generale is not None else '-'} / 20")
    y -= 16
    if rang is not None:
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Rang annuel: {rang} / {len(classement)}")
        y -= 14
    men = mention
    if men:
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Mention: {men}")
        y -= 16
    # Signatures
    c.setFont('Helvetica', 11); sig_y = margin + 50
    c.drawString(margin, sig_y, "Professeur principal:"); c.line(margin + 150, sig_y-2, margin + 320, sig_y-2)
    c.drawString(margin + 350, sig_y, "Chef d’établissement:"); c.line(margin + 520, sig_y-2, margin + 700, sig_y-2)
    c.drawString(margin, sig_y - 28, "Parent/Tuteur:"); c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)
    from datetime import datetime
    c.setFont('Helvetica-Oblique', 10); c.setFillColor(colors.darkgrey); c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.showPage(); c.save(); return response


@admin_required
def bulletins_annuels_classe_pdf(request, classe_id: int):
    """Bulletins annuels (T1+T2+T3) pour tous les élèves d'une classe en un seul PDF."""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleves = filter_by_user_school(Eleve.objects.filter(classe=classe).order_by('nom','prenom'), request.user, 'classe__ecole')
    matieres = list(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom'))
    evals_by_matiere = _collect_evals_all_trimestres(classe, matieres)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletins_annuels_{classe.nom}.pdf".replace(' ','_')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Pré-calcul moyennes classe par matière
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        tnum = Decimal('0'); tden = Decimal('0')
        for ev in evals:
            for n in Note.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                tnum += Decimal(n.note) * cc
                tden += cc
        moyennes_classe_par_matiere[mat.id] = (tnum/tden).quantize(Decimal('0.01')) if tden > 0 else None

    # Classement annuel
    moyenne_generale_map: dict[int, Decimal] = {}
    for e in eleves:
        notes_e = {n.evaluation_id: n for n in Note.objects.filter(eleve=e, evaluation__classe=classe, evaluation__trimestre__in=["T1","T2","T3"])}
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_e.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            if den > 0:
                moy_mat_e = (num/den)
                s_num += moy_mat_e * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        if s_den > 0:
            moyenne_generale_map[e.id] = (s_num / s_den).quantize(Decimal('0.01'))
    classement = sorted(moyenne_generale_map.items(), key=lambda t: t[1], reverse=True)
    rang_map: dict[int, int] = {eid: idx for idx, (eid, _) in enumerate(classement, start=1)}

    def mention_for(avg: Decimal | None) -> str:
        if avg is None: return ""
        if avg >= Decimal('16'): return "Très Bien"
        if avg >= Decimal('14'): return "Bien"
        if avg >= Decimal('12'): return "Assez Bien"
        if avg >= Decimal('10'): return "Passable"
        return "Insuffisant"

    def draw_for_student(eleve):
        try:
            from ecole_moderne.pdf_utils import draw_logo_watermark
            draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
        except Exception:
            pass
        margin = 2*cm; y = height - margin
        if getattr(classe, 'ecole', None):
            y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
        y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, "Bulletin de notes — Annuel"); y -= 40
        c.setFont('Helvetica', 12); c.drawString(margin, y, f"Élève: {eleve.nom} {eleve.prenom} (Matricule: {eleve.matricule or '-'})"); y -= 16
        c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}"); y -= 12
        c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

        c.setFont('Helvetica-Bold', 12)
        headers = ["Matière", "Coef.", "Moy. annuelle", "Moy. classe"]
        colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
        x = margin
        for i, h in enumerate(headers): c.drawString(x, y, h); x += colw[i]
        y -= 14; c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10
        c.setFont('Helvetica', 11)

        # Lignes
        lignes = []
        somme_moy_coef = Decimal('0'); somme_coef = Decimal('0')
        notes_by_eval = {n.evaluation_id: n for n in Note.objects.filter(eleve=eleve, evaluation__classe=classe, evaluation__trimestre__in=["T1","T2","T3"])}
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_by_eval.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            moy_mat = (num/den).quantize(Decimal('0.01')) if den > 0 else None
            if moy_mat is not None:
                somme_moy_coef += moy_mat * Decimal(mat.coefficient or 1)
                somme_coef += Decimal(mat.coefficient or 1)
            lignes.append({'matiere': mat.nom, 'coef_matiere': mat.coefficient, 'moyenne': moy_mat})

        for row in lignes:
            if y < margin + 60:
                c.showPage();
                try:
                    from ecole_moderne.pdf_utils import draw_logo_watermark
                    draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
                except Exception:
                    pass
                y = height - margin
            x = margin
            c.drawString(x, y, row['matiere']); x += colw[0]
            c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
            c.drawString(x, y, '-' if row['moyenne'] is None else f"{row['moyenne']}"); x += colw[2]
            mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
            c.drawString(x, y, '-' if mc is None else f"{mc}")
            y -= 14

        y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16
        mg = (somme_moy_coef / somme_coef).quantize(Decimal('0.01')) if somme_coef > 0 else None
        c.setFont('Helvetica-Bold', 13); c.drawString(margin, y, f"Moyenne générale annuelle: {mg if mg is not None else '-'} / 20"); y -= 16
        rg = rang_map.get(eleve.id)
        if rg is not None: c.setFont('Helvetica', 12); c.drawString(margin, y, f"Rang annuel: {rg} / {len(classement)}"); y -= 14
        men = mention_for(mg)
        if men: c.setFont('Helvetica', 12); c.drawString(margin, y, f"Mention: {men}"); y -= 16
        # Signatures
        c.setFont('Helvetica', 11); sig_y = margin + 50
        c.drawString(margin, sig_y, "Professeur principal:"); c.line(margin + 150, sig_y-2, margin + 320, sig_y-2)
        c.drawString(margin + 350, sig_y, "Chef d’établissement:"); c.line(margin + 520, sig_y-2, margin + 700, sig_y-2)
        c.drawString(margin, sig_y - 28, "Parent/Tuteur:"); c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)
        c.showPage()

    for e in eleves:
        draw_for_student(e)

    c.save(); return response


@login_required
@require_school_object(model=Classe, pk_kwarg='classe_id', field_path='ecole')
def classement_classe(request, classe_id: int, trimestre: str = "T1"):
    """Affiche le classement des élèves d'une classe pour un trimestre donné."""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Calculer le classement
    eleves = classe.eleves.filter(statut='actif').order_by('nom', 'prenom')
    classement = []
    
    for eleve in eleves:
        notes = Note.objects.filter(
            eleve=eleve,
            evaluation__classe=classe,
            evaluation__trimestre=trimestre
        ).select_related('evaluation__matiere')
        
        if not notes.exists():
            continue
            
        somme_moy_coef = Decimal('0')
        somme_coef = Decimal('0')
        
        matieres_notes = {}
        for note in notes:
            matiere = note.evaluation.matiere
            if matiere.id not in matieres_notes:
                matieres_notes[matiere.id] = {
                    'matiere': matiere.nom,
                    'coef': matiere.coefficient,
                    'notes': []
                }
            matieres_notes[matiere.id]['notes'].append(note.valeur)
        
        for matiere_data in matieres_notes.values():
            if matiere_data['notes']:
                moyenne_matiere = sum(matiere_data['notes']) / len(matiere_data['notes'])
                somme_moy_coef += Decimal(str(moyenne_matiere)) * Decimal(str(matiere_data['coef']))
                somme_coef += Decimal(str(matiere_data['coef']))
        
        if somme_coef > 0:
            moyenne_generale = (somme_moy_coef / somme_coef).quantize(Decimal('0.01'))
            classement.append({
                'eleve': eleve,
                'moyenne': moyenne_generale,
                'mention': mention_for(moyenne_generale)
            })
    
    # Trier par moyenne décroissante
    classement.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Ajouter les rangs
    for i, item in enumerate(classement):
        item['rang'] = i + 1
    
    context = {
        'classe': classe,
        'trimestre': trimestre,
        'classement': classement,
        'total_eleves': len(classement)
    }
    
    return render(request, 'notes/classement_classe.html', context)


@admin_required
def classement_classe_pdf(request, classe_id: int, trimestre: str = "T1"):
    """Export PDF du classement d'une classe."""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Récupérer le classement (même logique que la vue HTML)
    eleves = classe.eleves.filter(statut='actif').order_by('nom', 'prenom')
    classement = []
    
    for eleve in eleves:
        notes = Note.objects.filter(
            eleve=eleve,
            evaluation__classe=classe,
            evaluation__trimestre=trimestre
        ).select_related('evaluation__matiere')
        
        if not notes.exists():
            continue
            
        somme_moy_coef = Decimal('0')
        somme_coef = Decimal('0')
        
        matieres_notes = {}
        for note in notes:
            matiere = note.evaluation.matiere
            if matiere.id not in matieres_notes:
                matieres_notes[matiere.id] = {
                    'matiere': matiere.nom,
                    'coef': matiere.coefficient,
                    'notes': []
                }
            matieres_notes[matiere.id]['notes'].append(note.valeur)
        
        for matiere_data in matieres_notes.values():
            if matiere_data['notes']:
                moyenne_matiere = sum(matiere_data['notes']) / len(matiere_data['notes'])
                somme_moy_coef += Decimal(str(moyenne_matiere)) * Decimal(str(matiere_data['coef']))
                somme_coef += Decimal(str(matiere_data['coef']))
        
        if somme_coef > 0:
            moyenne_generale = (somme_moy_coef / somme_coef).quantize(Decimal('0.01'))
            classement.append({
                'eleve': eleve,
                'moyenne': moyenne_generale,
                'mention': mention_for(moyenne_generale)
            })
    
    # Trier par moyenne décroissante
    classement.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="classement_{classe.nom}_{trimestre}.pdf"'
    
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
    except Exception:
        pass
    
    margin = 2 * cm
    y = height - margin
    
    # En-tête avec logo et coordonnées
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    
    y -= 20
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(width/2, y, f"Classement de la classe {classe.nom} - {trimestre}")
    y -= 40
    
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Année scolaire: {getattr(classe, 'annee_scolaire', '')}")
    y -= 16
    c.drawString(margin, y, f"Total d'élèves classés: {len(classement)}")
    y -= 20
    
    # En-têtes du tableau
    c.setFont('Helvetica-Bold', 12)
    headers = ["Rang", "Nom et Prénom", "Matricule", "Moyenne", "Mention"]
    colw = [2*cm, 6*cm, 3*cm, 2.5*cm, 3*cm]
    x = margin
    for i, h in enumerate(headers):
        c.drawString(x, y, h)
        x += colw[i]
    
    y -= 14
    c.setFillColor(colors.lightgrey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 10
    
    # Données du classement
    c.setFont('Helvetica', 11)
    for i, item in enumerate(classement):
        if y < margin + 60:
            c.showPage()
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
            except Exception:
                pass
            y = height - margin
        
        x = margin
        c.drawString(x, y, str(i + 1))  # Rang
        x += colw[0]
        c.drawString(x, y, f"{item['eleve'].nom} {item['eleve'].prenom}")  # Nom
        x += colw[1]
        c.drawString(x, y, item['eleve'].matricule or '-')  # Matricule
        x += colw[2]
        c.drawString(x, y, f"{item['moyenne']}")  # Moyenne
        x += colw[3]
        c.drawString(x, y, item['mention'] or '-')  # Mention
        
        y -= 14
    
    # Pied de page
    c.setFont('Helvetica-Oblique', 10)
    c.setFillColor(colors.darkgrey)
    from datetime import datetime
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    c.save()
    return response


@admin_required
def classement_classe_excel(request, classe_id: int, trimestre: str = "T1"):
    """Export Excel du classement d'une classe."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Récupérer le classement (même logique que les autres vues)
    eleves = classe.eleves.filter(statut='actif').order_by('nom', 'prenom')
    classement = []
    
    for eleve in eleves:
        notes = Note.objects.filter(
            eleve=eleve,
            evaluation__classe=classe,
            evaluation__trimestre=trimestre
        ).select_related('evaluation__matiere')
        
        if not notes.exists():
            continue
            
        somme_moy_coef = Decimal('0')
        somme_coef = Decimal('0')
        
        matieres_notes = {}
        for note in notes:
            matiere = note.evaluation.matiere
            if matiere.id not in matieres_notes:
                matieres_notes[matiere.id] = {
                    'matiere': matiere.nom,
                    'coef': matiere.coefficient,
                    'notes': []
                }
            matieres_notes[matiere.id]['notes'].append(note.valeur)
        
        for matiere_data in matieres_notes.values():
            if matiere_data['notes']:
                moyenne_matiere = sum(matiere_data['notes']) / len(matiere_data['notes'])
                somme_moy_coef += Decimal(str(moyenne_matiere)) * Decimal(str(matiere_data['coef']))
                somme_coef += Decimal(str(matiere_data['coef']))
        
        if somme_coef > 0:
            moyenne_generale = (somme_moy_coef / somme_coef).quantize(Decimal('0.01'))
            classement.append({
                'eleve': eleve,
                'moyenne': moyenne_generale,
                'mention': mention_for(moyenne_generale)
            })
    
    # Trier par moyenne décroissante
    classement.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Classement {classe.nom} {trimestre}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = ["Rang", "Nom", "Prénom", "Matricule", "Moyenne", "Mention"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Données
    for row, item in enumerate(classement, 2):
        ws.cell(row=row, column=1, value=row-1)  # Rang
        ws.cell(row=row, column=2, value=item['eleve'].nom)
        ws.cell(row=row, column=3, value=item['eleve'].prenom)
        ws.cell(row=row, column=4, value=item['eleve'].matricule or '-')
        ws.cell(row=row, column=5, value=float(item['moyenne']))
        ws.cell(row=row, column=6, value=item['mention'] or '-')
    
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="classement_{classe.nom}_{trimestre}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@require_school_object(model=Classe, pk_kwarg='classe_id', field_path='ecole')
def cartes_scolaires_classe(request, classe_id):
    """Interface pour générer les cartes scolaires d'une classe"""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Filtrage par école pour non-admin
    if not request.user.is_superuser:
        ecole_user = user_school(request.user)
        if ecole_user and classe.ecole != ecole_user:
            messages.error(request, "Accès non autorisé à cette classe.")
            return redirect('notes:tableau_bord')
    
    # Récupérer tous les élèves de la classe
    eleves = classe.eleves.filter(statut='ACTIF').order_by('nom', 'prenom')
    
    context = {
        'classe': classe,
        'eleves': eleves,
        'nb_eleves': eleves.count(),
    }
    
    return render(request, 'notes/cartes_scolaires.html', context)


@login_required
@require_school_object(model=Classe, pk_kwarg='classe_id', field_path='ecole')
def cartes_scolaires_pdf(request, classe_id):
    """Génère les cartes scolaires PDF pour une classe"""
    classe = get_object_or_404(filter_by_user_school(Classe.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Filtrage par école pour non-admin
    if not request.user.is_superuser:
        ecole_user = user_school(request.user)
        if ecole_user and classe.ecole != ecole_user:
            messages.error(request, "Accès non autorisé à cette classe.")
            return redirect('notes:tableau_bord')
    
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        import io
    except ImportError:
        messages.error(request, "ReportLab requis pour générer le PDF")
        return redirect('notes:cartes_scolaires_classe', classe_id=classe_id)
    
    # Récupérer les élèves
    eleves = classe.eleves.filter(statut='ACTIF').order_by('nom', 'prenom')
    
    if not eleves.exists():
        messages.warning(request, "Aucun élève actif dans cette classe.")
        return redirect('notes:cartes_scolaires_classe', classe_id=classe_id)
    
    # Créer le PDF
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Filigrane standardisé (logo centré, rotation légère, opacité 4%)
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
    except Exception:
        pass
    
    # Configuration des cartes (10 cartes par page - 2x5)
    card_width = 8.5 * cm
    card_height = 5.4 * cm  # Format carte de crédit
    margin = 0.3 * cm
    spacing_x = 0.2 * cm  # Espacement horizontal réduit
    spacing_y = 0.15 * cm  # Espacement vertical réduit
    
    # Position des cartes sur la page (2 colonnes x 5 lignes)
    positions = [
        # Colonne gauche
        (margin, height - margin - card_height),  # Ligne 1
        (margin, height - margin - 2 * card_height - spacing_y),  # Ligne 2
        (margin, height - margin - 3 * card_height - 2 * spacing_y),  # Ligne 3
        (margin, height - margin - 4 * card_height - 3 * spacing_y),  # Ligne 4
        (margin, height - margin - 5 * card_height - 4 * spacing_y),  # Ligne 5
        # Colonne droite
        (margin + card_width + spacing_x, height - margin - card_height),  # Ligne 1
        (margin + card_width + spacing_x, height - margin - 2 * card_height - spacing_y),  # Ligne 2
        (margin + card_width + spacing_x, height - margin - 3 * card_height - 2 * spacing_y),  # Ligne 3
        (margin + card_width + spacing_x, height - margin - 4 * card_height - 3 * spacing_y),  # Ligne 4
        (margin + card_width + spacing_x, height - margin - 5 * card_height - 4 * spacing_y),  # Ligne 5
    ]
    
    card_count = 0
    
    for eleve in eleves:
        # Nouvelle page si nécessaire
        if card_count > 0 and card_count % 10 == 0:
            c.showPage()
            # Filigrane standardisé sur nouvelle page
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
            except Exception:
                pass
        
        # Position de la carte actuelle
        pos_x, pos_y = positions[card_count % 10]
        
        # Dessiner l'arrière-plan blanc de la carte d'abord
        c.setFillColor(colors.white)
        c.rect(pos_x, pos_y, card_width, card_height, fill=1, stroke=0)
        
        # Ajouter un filigrane sur chaque carte individuelle
        c.saveState()
        try:
            from django.contrib.staticfiles import finders
            import os
            from django.conf import settings
            
            # Chercher le logo pour le filigrane
            logo_path = finders.find('logos/logo.png')
            if not logo_path:
                logo_path = os.path.join(settings.BASE_DIR, 'static', 'logos', 'logo.png')
            
            if logo_path and os.path.exists(logo_path):
                # Calculer la position centrale de la carte
                center_x = pos_x + card_width / 2
                center_y = pos_y + card_height / 2
                
                # Appliquer la transformation (rotation et opacité)
                c.translate(center_x, center_y)
                c.rotate(30)  # Rotation de 30 degrés
                
                # Taille du filigrane (plus petit pour s'adapter à la carte)
                watermark_size = min(card_width, card_height) * 0.6
                
                # Dessiner le logo en filigrane avec opacité réduite
                c.setFillAlpha(0.08)  # Opacité très faible
                c.drawImage(logo_path, -watermark_size/2, -watermark_size/2, 
                          watermark_size, watermark_size, preserveAspectRatio=True)
        except Exception:
            pass
        c.restoreState()
        
        # Dessiner le cadre de la carte
        c.setStrokeColor(colors.black)
        c.setLineWidth(2)
        c.rect(pos_x, pos_y, card_width, card_height, fill=0, stroke=1)
        
        # Logo de l'école (en haut à droite)
        logo_size = 1.2 * cm
        logo_x = pos_x + card_width - logo_size - 0.2 * cm
        logo_y = pos_y + card_height - logo_size - 0.1 * cm
        
        try:
            import os
            from django.conf import settings
            from django.contrib.staticfiles import finders
            
            # Le logo sera dessiné directement sur le fond blanc de la carte
            # Pas besoin d'arrière-plan supplémentaire
            
            # Chercher le logo dans static/logos/
            logo_path = finders.find('logos/logo.png')
            
            if logo_path and os.path.exists(logo_path):
                # Dessiner le logo depuis static/logos/
                c.drawImage(logo_path, logo_x, logo_y, logo_size, logo_size, preserveAspectRatio=True)
            else:
                # Logo static introuvable - essayer le chemin direct
                static_logo_path = os.path.join(settings.BASE_DIR, 'static', 'logos', 'logo.png')
                if os.path.exists(static_logo_path):
                    c.drawImage(static_logo_path, logo_x, logo_y, logo_size, logo_size, preserveAspectRatio=True)
                else:
                    # Aucun logo trouvé
                    c.setFont('Helvetica', 6)
                    c.setFillColor(colors.grey)
                    c.drawCentredString(logo_x + logo_size/2, logo_y + logo_size/2, "LOGO")
        except Exception as e:
            # En cas d'erreur, afficher simplement le texte d'erreur
            c.setFont('Helvetica', 6)
            c.setFillColor(colors.red)
            c.drawCentredString(logo_x + logo_size/2, logo_y + logo_size/2, "ERREUR")
        
        # En-tête école (ajusté pour laisser place au logo)
        c.setFillColor(colors.darkblue)
        ecole_nom = classe.ecole.nom.upper() if classe.ecole else "ÉCOLE"
        
        # Ajuster la taille de police selon la longueur du nom
        if len(ecole_nom) > 40:
            c.setFont('Helvetica-Bold', 6)
        elif len(ecole_nom) > 30:
            c.setFont('Helvetica-Bold', 7)
        else:
            c.setFont('Helvetica-Bold', 8)
        
        # Centrer le nom en tenant compte du logo
        text_width = card_width - logo_size - 0.4 * cm
        c.drawCentredString(pos_x + text_width/2, pos_y + card_height - 0.8*cm, ecole_nom)
        
        # Titre "CARTE SCOLAIRE"
        c.setFont('Helvetica-Bold', 8)
        c.setFillColor(colors.red)
        c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 1.3*cm, "CARTE SCOLAIRE")
        
        # Année scolaire
        annee_actuelle = datetime.now().year
        annee_scolaire = f"{annee_actuelle}-{annee_actuelle + 1}"
        c.setFont('Helvetica', 7)
        c.setFillColor(colors.black)
        c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 1.7*cm, f"Année: {annee_scolaire}")
        
        # Photo de l'élève (côté gauche)
        photo_size = 2.2 * cm
        photo_x = pos_x + 0.3 * cm
        photo_y = pos_y + 0.5 * cm
        
        # Dessiner le cadre de la photo
        c.setStrokeColor(colors.grey)
        c.setLineWidth(1)
        c.rect(photo_x, photo_y, photo_size, photo_size)
        
        # Afficher la photo de l'élève si elle existe
        if eleve.photo and hasattr(eleve.photo, 'path'):
            try:
                import os
                from reportlab.lib.utils import ImageReader
                from PIL import Image
                
                # Vérifier que le fichier photo existe
                if os.path.exists(eleve.photo.path):
                    # Ouvrir et redimensionner l'image
                    with Image.open(eleve.photo.path) as img:
                        # Convertir en RGB si nécessaire
                        if img.mode != 'RGB':
                            img = img.convert('RGB')
                        
                        # Calculer les dimensions pour maintenir le ratio
                        img_width, img_height = img.size
                        ratio = min(photo_size / (img_width * 72/96), photo_size / (img_height * 72/96))
                        
                        new_width = img_width * ratio * 72/96
                        new_height = img_height * ratio * 72/96
                        
                        # Centrer l'image dans le cadre
                        img_x = photo_x + (photo_size - new_width) / 2
                        img_y = photo_y + (photo_size - new_height) / 2
                        
                        # Dessiner l'image
                        c.drawImage(ImageReader(img), img_x, img_y, new_width, new_height)
                else:
                    # Fichier photo introuvable
                    c.setFont('Helvetica', 7)
                    c.setFillColor(colors.red)
                    c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.1*cm, "PHOTO")
                    c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.3*cm, "MANQUANTE")
            except Exception as e:
                # Erreur lors du traitement de l'image
                c.setFont('Helvetica', 7)
                c.setFillColor(colors.red)
                c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.1*cm, "ERREUR")
                c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.3*cm, "PHOTO")
        else:
            # Pas de photo définie
            c.setFont('Helvetica', 8)
            c.setFillColor(colors.grey)
            c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.2*cm, "PHOTO")
        
        # Informations élève (côté droit)
        info_x = pos_x + 3.2 * cm
        info_y = pos_y + card_height - 2.2 * cm
        
        c.setFillColor(colors.black)
        
        # Nom et prénom
        c.setFont('Helvetica-Bold', 9)
        nom_complet = f"{eleve.nom} {eleve.prenom}".upper()
        if len(nom_complet) > 25:
            nom_complet = nom_complet[:25] + "..."
        c.drawString(info_x, info_y, nom_complet)
        info_y -= 0.4 * cm
        
        # Matricule
        c.setFont('Helvetica', 7)
        c.drawString(info_x, info_y, f"Matricule: {eleve.matricule}")
        info_y -= 0.3 * cm
        
        # Classe
        c.drawString(info_x, info_y, f"Classe: {classe.nom}")
        info_y -= 0.3 * cm
        
        # Date de naissance
        if eleve.date_naissance:
            date_naiss = eleve.date_naissance.strftime('%d/%m/%Y')
            c.drawString(info_x, info_y, f"Né(e) le: {date_naiss}")
            info_y -= 0.3 * cm
        
        # Contact responsable
        if hasattr(eleve, 'responsable_principal') and eleve.responsable_principal:
            resp = eleve.responsable_principal
            if resp.telephone:
                c.drawString(info_x, info_y, f"Contact: {resp.telephone}")
        
        # Pied de carte
        c.setFont('Helvetica', 6)
        c.setFillColor(colors.grey)
        c.drawString(pos_x + 0.2*cm, pos_y + 0.2*cm, "Cette carte est strictement personnelle")
        
        card_count += 1
    
    c.save()
    
    # Préparer la réponse
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"cartes_scolaires_{classe.nom.replace(' ', '_')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    
    return response


@login_required
@require_school_object(model=Eleve, pk_kwarg='matricule', field_path='classe__ecole')
def carte_eleve_pdf(request, matricule):
    """Génère la carte scolaire d'un élève spécifique par son matricule"""
    from django.contrib import messages
    from eleves.models import Eleve
    
    # Récupérer l'élève par matricule
    try:
        eleve = Eleve.objects.get(matricule=matricule, statut='actif')
        
        # Vérifier les permissions (sauf pour superuser)
        if not request.user.is_superuser:
            if hasattr(request.user, 'profil') and request.user.profil.ecole:
                if eleve.classe.ecole != request.user.profil.ecole:
                    messages.error(request, "Vous n'avez pas accès à cet élève.")
                    return redirect('notes:tableau_bord')
            else:
                messages.error(request, "Accès non autorisé.")
                return redirect('notes:tableau_bord')
                
    except Eleve.DoesNotExist:
        messages.error(request, f"Aucun élève actif trouvé avec le matricule: {matricule}")
        return redirect('notes:tableau_bord')
    
    # Créer le PDF avec une seule carte
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Filigrane standardisé (logo centré, rotation légère, opacité 4%)
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
    except Exception:
        pass
    
    # Configuration de la carte (centrée sur la page)
    card_width = 8.5 * cm
    card_height = 5.4 * cm  # Format carte de crédit
    
    # Position centrée sur la page
    pos_x = (width - card_width) / 2
    pos_y = (height - card_height) / 2
    
    # Dessiner l'arrière-plan blanc de la carte d'abord
    c.setFillColor(colors.white)
    c.rect(pos_x, pos_y, card_width, card_height, fill=1, stroke=0)
    
    # Ajouter un filigrane sur la carte individuelle
    c.saveState()
    try:
        from django.contrib.staticfiles import finders
        import os
        from django.conf import settings
        
        # Chercher le logo pour le filigrane
        logo_path = finders.find('logos/logo.png')
        if not logo_path:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'logos', 'logo.png')
        
        if logo_path and os.path.exists(logo_path):
            # Calculer la position centrale de la carte
            center_x = pos_x + card_width / 2
            center_y = pos_y + card_height / 2
            
            # Appliquer la transformation (rotation et opacité)
            c.translate(center_x, center_y)
            c.rotate(30)  # Rotation de 30 degrés
            
            # Taille du filigrane (plus petit pour s'adapter à la carte)
            watermark_size = min(card_width, card_height) * 0.6
            
            # Dessiner le logo en filigrane avec opacité réduite
            c.setFillAlpha(0.08)  # Opacité très faible
            c.drawImage(logo_path, -watermark_size/2, -watermark_size/2, 
                      watermark_size, watermark_size, preserveAspectRatio=True)
    except Exception:
        pass
    c.restoreState()
    
    # Dessiner le cadre de la carte
    c.setStrokeColor(colors.black)
    c.setLineWidth(2)
    c.rect(pos_x, pos_y, card_width, card_height, fill=0, stroke=1)
    
    # Logo de l'école (en haut à droite)
    logo_size = 1.2 * cm
    logo_x = pos_x + card_width - logo_size - 0.2 * cm
    logo_y = pos_y + card_height - logo_size - 0.1 * cm
    
    # Ajouter un rectangle blanc derrière le logo pour contraste
    c.setFillColor(colors.white)
    c.rect(logo_x - 0.05 * cm, logo_y - 0.05 * cm, 
           logo_size + 0.1 * cm, logo_size + 0.1 * cm, fill=1, stroke=0)
    
    try:
        from django.contrib.staticfiles import finders
        import os
        from django.conf import settings
        
        # Chercher le logo
        logo_path = finders.find('logos/logo.png')
        if not logo_path:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'logos', 'logo.png')
        
        if logo_path and os.path.exists(logo_path):
            c.drawImage(logo_path, logo_x, logo_y, logo_size, logo_size, preserveAspectRatio=True)
        else:
            # Fallback: texte si pas de logo
            c.setFont("Helvetica-Bold", 8)
            c.setFillColor(colors.black)
            c.drawString(logo_x, logo_y + logo_size/2, "LOGO")
    except Exception:
        # Fallback: texte si erreur
        c.setFont("Helvetica-Bold", 8)
        c.setFillColor(colors.black)
        c.drawString(logo_x, logo_y + logo_size/2, "LOGO")
    
    # Nom de l'école (en haut à gauche)
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(colors.black)
    school_name = eleve.classe.ecole.nom if eleve.classe.ecole else "École"
    c.drawString(pos_x + 0.2 * cm, pos_y + card_height - 0.4 * cm, school_name[:25])
    
    # Année scolaire (sous le nom de l'école)
    c.setFont("Helvetica", 8)
    current_year = timezone.now().year
    next_year = current_year + 1
    annee_scolaire = f"Année {current_year}-{next_year}"
    c.drawString(pos_x + 0.2 * cm, pos_y + card_height - 0.7 * cm, annee_scolaire)
    
    # Photo de l'élève (à gauche)
    photo_size = 1.8 * cm
    photo_x = pos_x + 0.2 * cm
    photo_y = pos_y + 0.3 * cm
    
    # Cadre pour la photo
    c.setStrokeColor(colors.gray)
    c.setLineWidth(1)
    c.rect(photo_x, photo_y, photo_size, photo_size, fill=0, stroke=1)
    
    # Afficher la photo ou placeholder
    try:
        if eleve.photo and hasattr(eleve.photo, 'path') and os.path.exists(eleve.photo.path):
            from PIL import Image
            
            # Ouvrir et redimensionner l'image
            img = Image.open(eleve.photo.path)
            
            # Convertir en RGB si nécessaire
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Redimensionner en gardant les proportions
            img.thumbnail((int(photo_size * 2.83), int(photo_size * 2.83)), Image.Resampling.LANCZOS)
            
            # Calculer la position pour centrer l'image
            img_width, img_height = img.size
            img_width_cm = img_width / 2.83  # Conversion pixels vers cm
            img_height_cm = img_height / 2.83
            
            center_x = photo_x + (photo_size - img_width_cm) / 2
            center_y = photo_y + (photo_size - img_height_cm) / 2
            
            # Sauvegarder temporairement l'image
            temp_path = f"/tmp/temp_photo_{eleve.id}.jpg"
            img.save(temp_path, "JPEG", quality=85)
            
            # Dessiner l'image redimensionnée et centrée
            c.drawImage(temp_path, center_x, center_y, img_width_cm, img_height_cm)
            
            # Nettoyer le fichier temporaire
            try:
                os.remove(temp_path)
            except:
                pass
                
        else:
            # Placeholder si pas de photo
            c.setFont("Helvetica", 8)
            c.setFillColor(colors.gray)
            c.drawCentredText(photo_x + photo_size/2, photo_y + photo_size/2, "PHOTO")
            
    except Exception as e:
        # Placeholder en cas d'erreur
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.gray)
        c.drawCentredText(photo_x + photo_size/2, photo_y + photo_size/2, "PHOTO")
    
    # Informations de l'élève (à droite de la photo)
    info_x = photo_x + photo_size + 0.3 * cm
    info_y_start = pos_y + card_height - 1.2 * cm
    line_height = 0.35 * cm
    
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(colors.black)
    
    # Nom complet (en majuscules)
    nom_complet = f"{eleve.prenom} {eleve.nom}".upper()
    c.drawString(info_x, info_y_start, nom_complet[:20])
    
    # Matricule
    c.setFont("Helvetica", 9)
    c.drawString(info_x, info_y_start - line_height, f"Mat: {eleve.matricule}")
    
    # Classe
    c.drawString(info_x, info_y_start - 2 * line_height, f"Classe: {eleve.classe.nom}")
    
    # Date de naissance
    if eleve.date_naissance:
        date_naiss = eleve.date_naissance.strftime("%d/%m/%Y")
        c.drawString(info_x, info_y_start - 3 * line_height, f"Né(e): {date_naiss}")
    
    # Téléphone responsable
    if eleve.telephone_responsable:
        tel = eleve.telephone_responsable[:12]  # Limiter la longueur
        c.drawString(info_x, info_y_start - 4 * line_height, f"Tél: {tel}")
    
    # Note légale (en bas)
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.gray)
    c.drawCentredText(pos_x + card_width/2, pos_y + 0.1 * cm, 
                     "Cette carte est strictement personnelle")
    
    # Finaliser le PDF
    c.showPage()
    c.save()
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"carte_scolaire_{eleve.matricule}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
