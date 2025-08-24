from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta, date
from django.utils import timezone as django_timezone
from decimal import Decimal
import json
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from .models import Rapport, TypeRapport, ExportProgramme
from .utils import collecter_donnees_periode, generer_pdf_periode, _draw_header_and_watermark
from eleves.models import Eleve, Ecole
from paiements.models import Paiement, PaiementRemise, EcheancierPaiement, TypePaiement
from bus.models import AbonnementBus
from depenses.models import Depense
from salaires.models import Enseignant, EtatSalaire
from utilisateurs.utils import user_is_admin, user_school
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Décorateur d'accès admin uniquement
admin_required = user_passes_test(user_is_admin)

@login_required
@admin_required
def tableau_bord(request):
    """Vue principale du module Rapports"""
    context = {
        'rapports_recents': Rapport.objects.filter(
            genere_par=request.user
        ).order_by('-date_generation')[:10],
        'types_rapports': TypeRapport.objects.filter(actif=True),
        'exports_programmes': ExportProgramme.objects.filter(
            cree_par=request.user,
            statut='ACTIF'
        )[:5],
        'today': date.today()
    }
    return render(request, 'rapports/tableau_bord.html', context)

@login_required
@admin_required
def generer_rapport_journalier(request):
    """Génère un rapport journalier automatique"""
    date_rapport = date.today()
    if request.GET.get('date'):
        date_rapport = datetime.strptime(request.GET.get('date'), '%Y-%m-%d').date()
    
    # Collecte des données journalières
    donnees = collecter_donnees_journalieres(date_rapport, user=request.user)
    
    # Génération du PDF
    pdf_buffer = generer_pdf_journalier(donnees, date_rapport)
    
    # Sauvegarde du rapport
    rapport = Rapport.objects.create(
        type_rapport=get_or_create_type_rapport('JOURNALIER'),
        titre=f"Rapport Journalier - {date_rapport.strftime('%d/%m/%Y')}",
        periode_debut=date_rapport,
        periode_fin=date_rapport,
        format_rapport='PDF',
        statut='TERMINE',
        genere_par=request.user,
        parametres=json.dumps(donnees, default=str)
    )
    
    # Sauvegarde du fichier PDF
    rapport.fichier.save(
        f'rapport_journalier_{date_rapport.strftime("%Y%m%d")}.pdf',
        pdf_buffer
    )
    
    return HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')

@login_required
@admin_required
def export_rapport_annuel_excel(request):
    """Export Excel du rapport annuel."""
    aujourd_hui = date.today()
    debut_annee = date(aujourd_hui.year, 1, 1)
    fin_annee = date(aujourd_hui.year, 12, 31)
    if request.GET.get('annee'):
        annee = int(request.GET.get('annee'))
        debut_annee = date(annee, 1, 1)
        fin_annee = date(annee, 12, 31)
    debut_dt = django_timezone.make_aware(datetime.combine(debut_annee, datetime.min.time()))
    fin_dt = django_timezone.make_aware(datetime.combine(fin_annee, datetime.max.time()))

    donnees = collecter_donnees_periode(debut_dt, fin_dt, 'ANNUEL', user=request.user)
    wb = _build_excel_from_donnees(donnees, titre=f"Rapport Annuel - {debut_annee.year}")
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rapport_annuel_{debut_annee.year}.xlsx"'
    wb.save(response)
    return response

@login_required
@admin_required
def export_rapport_mensuel_excel(request):
    """Export Excel du rapport mensuel."""
    aujourd_hui = date.today()
    debut_mois = aujourd_hui.replace(day=1)
    fin_mois = (debut_mois + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    if request.GET.get('mois') and request.GET.get('annee'):
        mois = int(request.GET.get('mois'))
        annee = int(request.GET.get('annee'))
        debut_mois = date(annee, mois, 1)
        fin_mois = (debut_mois + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    debut_dt = django_timezone.make_aware(datetime.combine(debut_mois, datetime.min.time()))
    fin_dt = django_timezone.make_aware(datetime.combine(fin_mois, datetime.max.time()))

    donnees = collecter_donnees_periode(debut_dt, fin_dt, 'MENSUEL', user=request.user)
    wb = _build_excel_from_donnees(donnees, titre=f"Rapport Mensuel - {debut_mois.strftime('%B %Y')}")
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rapport_mensuel_{debut_mois.strftime("%Y%m")}.xlsx"'
    wb.save(response)
    return response

@login_required
@admin_required
def export_rapport_hebdomadaire_excel(request):
    """Export Excel du rapport hebdomadaire (lundi-dimanche)."""
    aujourd_hui = date.today()
    debut_semaine = aujourd_hui - timedelta(days=aujourd_hui.weekday())
    fin_semaine = debut_semaine + timedelta(days=6)
    if request.GET.get('debut'):
        debut_semaine = datetime.strptime(request.GET.get('debut'), '%Y-%m-%d').date()
        fin_semaine = debut_semaine + timedelta(days=6)
    debut_dt = django_timezone.make_aware(datetime.combine(debut_semaine, datetime.min.time()))
    fin_dt = django_timezone.make_aware(datetime.combine(fin_semaine, datetime.max.time()))

    donnees = collecter_donnees_periode(debut_dt, fin_dt, 'HEBDOMADAIRE', user=request.user)
    wb = _build_excel_from_donnees(donnees, titre=f"Rapport Hebdomadaire - {debut_semaine.strftime('%d/%m')} au {fin_semaine.strftime('%d/%m/%Y')}")
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rapport_hebdomadaire_{debut_semaine.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
@admin_required
def export_rapport_journalier_excel(request):
    """Export Excel du rapport journalier (mêmes données que le PDF)."""
    date_rapport = date.today()
    if request.GET.get('date'):
        date_rapport = datetime.strptime(request.GET.get('date'), '%Y-%m-%d').date()

    donnees = collecter_donnees_journalieres(date_rapport, user=request.user)
    wb = _build_excel_from_donnees(donnees, titre=f"Rapport Journalier - {date_rapport.strftime('%d/%m/%Y')}")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rapport_journalier_{date_rapport.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
@admin_required
def generer_rapport_hebdomadaire(request):
    """Génère un rapport hebdomadaire"""
    # Calcul de la semaine (lundi à dimanche)
    aujourd_hui = date.today()
    debut_semaine = aujourd_hui - timedelta(days=aujourd_hui.weekday())
    fin_semaine = debut_semaine + timedelta(days=6)
    
    if request.GET.get('debut'):
        debut_semaine = datetime.strptime(request.GET.get('debut'), '%Y-%m-%d').date()
        fin_semaine = debut_semaine + timedelta(days=6)
    
    # Convertir les dates en datetime timezone-aware pour éviter les warnings
    debut_dt = django_timezone.make_aware(datetime.combine(debut_semaine, datetime.min.time()))
    fin_dt = django_timezone.make_aware(datetime.combine(fin_semaine, datetime.max.time()))
    
    donnees = collecter_donnees_periode(debut_dt, fin_dt, 'HEBDOMADAIRE', user=request.user)
    pdf_buffer = generer_pdf_periode(donnees, debut_semaine, fin_semaine, 'HEBDOMADAIRE')
    
    # Sauvegarde
    rapport = Rapport.objects.create(
        type_rapport=get_or_create_type_rapport('HEBDOMADAIRE'),
        titre=f"Rapport Hebdomadaire - {debut_semaine.strftime('%d/%m')} au {fin_semaine.strftime('%d/%m/%Y')}",
        periode_debut=datetime.combine(debut_semaine, datetime.min.time()),
        periode_fin=datetime.combine(fin_semaine, datetime.max.time()),
        format_rapport='PDF',
        statut='TERMINE',
        genere_par=request.user,
        parametres=json.dumps(donnees, default=str)
    )
    
    rapport.fichier.save(
        f'rapport_hebdomadaire_{debut_semaine.strftime("%Y%m%d")}.pdf',
        pdf_buffer
    )
    
    return HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')

@login_required
@admin_required
def generer_rapport_mensuel(request):
    """Génère un rapport mensuel"""
    aujourd_hui = date.today()
    debut_mois = aujourd_hui.replace(day=1)
    fin_mois = (debut_mois + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    if request.GET.get('mois') and request.GET.get('annee'):
        mois = int(request.GET.get('mois'))
        annee = int(request.GET.get('annee'))
        debut_mois = date(annee, mois, 1)
        fin_mois = (debut_mois + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # Convertir les dates en datetime timezone-aware pour éviter les warnings
    debut_dt = django_timezone.make_aware(datetime.combine(debut_mois, datetime.min.time()))
    fin_dt = django_timezone.make_aware(datetime.combine(fin_mois, datetime.max.time()))
    
    donnees = collecter_donnees_periode(debut_dt, fin_dt, 'MENSUEL', user=request.user)
    pdf_buffer = generer_pdf_periode(donnees, debut_mois, fin_mois, 'MENSUEL')
    
    # Sauvegarde
    rapport = Rapport.objects.create(
        type_rapport=get_or_create_type_rapport('MENSUEL'),
        titre=f"Rapport Mensuel - {debut_mois.strftime('%B %Y')}",
        periode_debut=datetime.combine(debut_mois, datetime.min.time()),
        periode_fin=datetime.combine(fin_mois, datetime.max.time()),
        format_rapport='PDF',
        statut='TERMINE',
        genere_par=request.user,
        parametres=json.dumps(donnees, default=str)
    )
    
    rapport.fichier.save(
        f'rapport_mensuel_{debut_mois.strftime("%Y%m")}.pdf',
        pdf_buffer
    )
    
    return HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')

@login_required
@admin_required
def generer_rapport_annuel(request):
    """Génère un rapport annuel"""
    aujourd_hui = date.today()
    debut_annee = date(aujourd_hui.year, 1, 1)
    fin_annee = date(aujourd_hui.year, 12, 31)
    
    if request.GET.get('annee'):
        annee = int(request.GET.get('annee'))
        debut_annee = date(annee, 1, 1)
        fin_annee = date(annee, 12, 31)
    
    # Convertir les dates en datetime timezone-aware pour éviter les warnings
    debut_dt = django_timezone.make_aware(datetime.combine(debut_annee, datetime.min.time()))
    fin_dt = django_timezone.make_aware(datetime.combine(fin_annee, datetime.max.time()))
    
    donnees = collecter_donnees_periode(debut_dt, fin_dt, 'ANNUEL', user=request.user)
    pdf_buffer = generer_pdf_periode(donnees, debut_annee, fin_annee, 'ANNUEL')
    
    # Sauvegarde
    rapport = Rapport.objects.create(
        type_rapport=get_or_create_type_rapport('ANNUEL'),
        titre=f"Rapport Annuel - {debut_annee.year}",
        periode_debut=datetime.combine(debut_annee, datetime.min.time()),
        periode_fin=datetime.combine(fin_annee, datetime.max.time()),
        format_rapport='PDF',
        statut='TERMINE',
        genere_par=request.user,
        parametres=json.dumps(donnees, default=str)
    )
    
    rapport.fichier.save(
        f'rapport_annuel_{debut_annee.year}.pdf',
        pdf_buffer
    )
    
    return HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')

@login_required
@admin_required
def liste_rapports(request):
    """Liste tous les rapports générés"""
    rapports = Rapport.objects.filter(
        genere_par=request.user
    ).order_by('-date_generation')
    
    context = {
        'rapports': rapports
    }
    return render(request, 'rapports/liste_rapports.html', context)


@login_required
@admin_required
def rapport_transport_scolaire(request):
    """Tableau Transport scolaire par classe: Classe | Nombre d'abonnés | Total payé | Reste à payer

    Logique:
    - Nombre d'abonnés: nombre d'abonnements bus ACTIFS par classe (\n pas les élèves uniques si plusieurs abonnements).
    - Total dû: somme des montants d'abonnements ACTIFS par classe.
    - Total payé: si un TypePaiement dont le nom contient 'bus' ou 'transport' existe,
      on additionne les paiements VALIDÉS de ces types par classe.
    - Reste à payer: max(Total dû - Total payé, 0).
    """
    # Restreindre par école selon l'utilisateur
    ecole_user = user_school(request.user)
    abonnements_qs = AbonnementBus.objects.select_related('eleve', 'eleve__classe')
    paiements_qs = Paiement.objects.select_related('eleve', 'eleve__classe', 'type_paiement')

    if ecole_user:
        abonnements_qs = abonnements_qs.filter(eleve__classe__ecole=ecole_user)
        paiements_qs = paiements_qs.filter(eleve__classe__ecole=ecole_user)

    # Abonnements actifs uniquement
    abonnements_qs = abonnements_qs.filter(statut=AbonnementBus.Statut.ACTIF)

    # Agrégation abonnements par classe
    from django.db.models import F
    abonnements_par_classe = abonnements_qs.values(
        'eleve__classe__id', 'eleve__classe__nom'
    ).annotate(
        nb_abonnes=Count('id'),
        total_du=Sum('montant')
    )

    # Détecter TypePaiement "Transport/Bus"
    transport_types = TypePaiement.objects.filter(
        Q(nom__icontains='transport') | Q(nom__icontains='bus')
    )

    total_paye_par_classe = {}
    if transport_types.exists():
        paiements_transport = paiements_qs.filter(
            type_paiement__in=transport_types,
            statut='VALIDE'
        )
        paiements_group = paiements_transport.values(
            'eleve__classe__id'
        ).annotate(
            total_paye=Sum('montant')
        )
        total_paye_par_classe = {row['eleve__classe__id']: row['total_paye'] or 0 for row in paiements_group}

    lignes = []
    totals = {
        'nb_abonnes': 0,
        'total_du': 0,
        'total_paye': 0,
        'reste': 0,
    }

    for row in abonnements_par_classe:
        classe_id = row['eleve__classe__id']
        classe_nom = row['eleve__classe__nom'] or 'Classe'
        nb_abonnes = row['nb_abonnes'] or 0
        total_du = row['total_du'] or 0
        total_paye = total_paye_par_classe.get(classe_id, 0)
        reste = total_du - total_paye
        if reste < 0:
            reste = 0

        lignes.append({
            'classe': classe_nom,
            'nb_abonnes': nb_abonnes,
            'total_paye': total_paye,
            'reste': reste,
            'total_du': total_du,
        })

        totals['nb_abonnes'] += nb_abonnes
        totals['total_du'] += total_du
        totals['total_paye'] += total_paye
        totals['reste'] += reste

    # Ordonner par nom de classe
    lignes = sorted(lignes, key=lambda x: x['classe'])

    context = {
        'lignes': lignes,
        'totals': totals,
        'has_payment_types': transport_types.exists(),
        'ecole': getattr(ecole_user, 'nom', None),
    }

    return render(request, 'rapports/transport_scolaire.html', context)

def get_or_create_type_rapport(nom):
    """Récupère ou crée un type de rapport"""
    type_rapport, created = TypeRapport.objects.get_or_create(
        nom=nom,
        defaults={
            'description': f'Rapport {nom.lower()}',
            'categorie': 'FINANCIER',
            'template_path': f'rapports/{nom.lower()}.html',
            'actif': True
        }
    )
    return type_rapport

def _build_excel_from_donnees(donnees, titre):
    """Construit un classeur Excel à partir de la structure de données des rapports.
    Feuille 1: Synthèse par école.
    Feuille 2: Répartition par classe (toutes écoles).
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Synthèse'

    header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='DDDDDD')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titre
    ws1.append([titre])
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal='center')

    headers1 = [
        'École', 'Nouveaux élèves', 'Nb paiements', 'Scolarité normale', "Scolarité payée",
        "Frais d'inscription", 'Reste à payer', 'Montant original', 'Remises', 'Net encaissé',
        'Nb dépenses', 'Total dépenses', 'États salaires', 'Total salaires'
    ]
    ws1.append(headers1)
    for col in range(1, len(headers1) + 1):
        c = ws1.cell(row=2, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        c.border = border_all

    row = 3
    for _, e in donnees.get('ecoles', {}).items():
        vals = [
            e.get('nom', ''),
            e.get('nouveaux_eleves', 0),
            e['paiements'].get('nombre', 0),
            float(e['paiements'].get('total_du_concernes', 0) or 0),
            float(e['paiements'].get('scolarite', 0) or 0),
            float(e['paiements'].get('frais_inscription', 0) or 0),
            float(e['paiements'].get('reste_a_payer', 0) or 0),
            float(e['paiements'].get('montant_original', 0) or 0),
            float(e['paiements'].get('total_remises', 0) or 0),
            float(e['paiements'].get('montant_total', 0) or 0),
            e['depenses'].get('nombre', 0),
            float(donnees.get('depenses_globales', {}).get('montant_total', 0) or 0) if False else float(e['depenses'].get('montant_total', 0) or 0),
            e['salaires'].get('etats_valides', 0),
            float(e['salaires'].get('montant_total', 0) or 0),
        ]
        ws1.append(vals)
        for col in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row, column=col)
            cell.border = border_all
            if col >= 4 and col != 11 and col != 13:  # colonnes montants
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        row += 1

    # Largeurs
    widths = [26, 16, 14, 18, 18, 18, 16, 18, 16, 16, 14, 16, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Feuille 2: par classe
    ws2 = wb.create_sheet('Par classe')
    headers2 = ['École', 'Classe', 'Effectif', 'Total dû', 'Total payé', 'Remises', 'Reste à payer']
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        c = ws2.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        c.border = border_all

    r = 2
    for _, e in donnees.get('ecoles', {}).items():
        ecole_nom = e.get('nom', '')
        for c in e.get('classes', []) or []:
            ws2.append([
                ecole_nom,
                c.get('classe', ''),
                c.get('effectif', 0),
                float(c.get('total_du', 0) or 0),
                float(c.get('total_paye', 0) or 0),
                float(c.get('remises', 0) or 0),
                float(c.get('reste', 0) or 0),
            ])
            for col in range(1, len(headers2) + 1):
                cell = ws2.cell(row=r, column=col)
                cell.border = border_all
                if col >= 4:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            r += 1

    for i, w in enumerate([22, 18, 12, 16, 16, 16, 16], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    return wb

def collecter_donnees_journalieres(date_rapport, user=None):
    """Collecte toutes les données importantes pour le rapport journalier"""
    donnees = {
        'date': date_rapport,
        'ecoles': {},
        # Dépenses non reliées à l'école: globales pour la journée
        'depenses_globales': {
            'nombre': 0,  # initialisé, sera remplacé par des entiers
            'montant_total': Decimal('0')
        }
    }

    # Restreindre le périmètre des écoles selon l'utilisateur
    ecoles_qs = Ecole.objects.all()
    if user is not None and not user_is_admin(user):
        ecole_user = user_school(user)
        ecoles_qs = ecoles_qs.filter(id=getattr(ecole_user, 'id', None)) if ecole_user else Ecole.objects.none()

    # Dépenses du jour (GLOBAL - pas de relation à Ecole)
    depenses_jour_global = Depense.objects.filter(
        date_facture=date_rapport,
        statut='VALIDEE'
    )
    donnees['depenses_globales']['nombre'] = depenses_jour_global.count()
    donnees['depenses_globales']['montant_total'] = depenses_jour_global.aggregate(
        total=Sum('montant_ttc')
    )['total'] or Decimal('0')

    # Pour chaque école
    for ecole in ecoles_qs:
        # Nom d'école affiché (normalisation pour SONFONIA)
        _nom_affiche = ecole.nom
        _nom_upper = (ecole.nom or '').upper()
        if any(key in _nom_upper for key in ['SONFONIA', 'SONFONIE']):
            _nom_affiche = "GROUPE SCOLAIRE HADJA KANFING DIANÉ-SONFONIA"

        donnees_ecole = {
            'nom': _nom_affiche,
            'nouveaux_eleves': Eleve.objects.filter(
                classe__ecole=ecole,
                date_inscription=date_rapport
            ).count(),
            'paiements': {
                'nombre': 0,
                'montant_total': Decimal('0'),
                'frais_inscription': Decimal('0'),
                'scolarite': Decimal('0'),
                # Total dû (GNF) pour les élèves concernés ce jour (paiement/inscription)
                'total_du_concernes': Decimal('0')
            },
            # Répartition par classe (remplie plus bas)
            'classes': [],
            # Dépenses affichées par école mises à 0 pour éviter toute confusion,
            # car le modèle Depense n'est pas lié à Ecole. Un total global sera affiché.
            'depenses': {
                'nombre': 0,
                'montant_total': Decimal('0')
            },
            'salaires': {
                'etats_valides': 0,
                'montant_total': Decimal('0')
            }
        }
        
        # Paiements du jour - utiliser une approche plus flexible
        paiements_jour = Paiement.objects.filter(
            eleve__classe__ecole=ecole,
            date_paiement=date_rapport
        ).exclude(statut='ANNULE')  # Exclure seulement les annulés
        
        # Si pas de paiements ce jour, essayer avec les paiements récents (30 derniers jours)
        if not paiements_jour.exists():
            date_limite = date_rapport - timedelta(days=30)
            paiements_jour = Paiement.objects.filter(
                eleve__classe__ecole=ecole,
                date_paiement__gte=date_limite,
                statut='VALIDE'
            )
        
        donnees_ecole['paiements']['nombre'] = paiements_jour.count()
        donnees_ecole['paiements']['montant_total'] = paiements_jour.aggregate(
            total=Sum('montant')
        )['total'] or Decimal('0')
        
        # Calculer les remises appliquées
        remises_appliquees = PaiementRemise.objects.filter(
            paiement__in=paiements_jour
        ).aggregate(
            total_remises=Sum('montant_remise')
        )['total_remises'] or Decimal('0')
        
        # Calculer le montant original (avant remises)
        montant_original = donnees_ecole['paiements']['montant_total'] + remises_appliquees
        
        # Ajouter les données des remises
        donnees_ecole['paiements']['montant_original'] = montant_original
        donnees_ecole['paiements']['total_remises'] = remises_appliquees
        donnees_ecole['paiements']['montant_net'] = donnees_ecole['paiements']['montant_total']
        
        # Séparation frais d'inscription et scolarité (alignée avec rapports/utils.collecter_donnees_periode)
        frais_inscription = Decimal('0')
        scolarite = Decimal('0')
        non_categorises = Decimal('0')

        for p in paiements_jour.select_related('type_paiement'):
            montant = p.montant or Decimal('0')
            nom = (getattr(getattr(p, 'type_paiement', None), 'nom', '') or '').lower()

            has_inscription = 'inscription' in nom
            has_scolarite = ('scolar' in nom) or ('tranche' in nom) or ('1ère tranche' in nom) or ('2ème tranche' in nom) or ('3ème tranche' in nom)

            if has_inscription and has_scolarite:
                # Paiement combiné: 30 000 GNF pour inscription, reste en scolarité
                part_ins = min(Decimal('30000'), montant)
                part_sco = montant - part_ins
                frais_inscription += part_ins
                scolarite += part_sco
            elif has_inscription:
                frais_inscription += montant
            elif has_scolarite:
                scolarite += montant
            else:
                non_categorises += montant

        # Estimation/fallback: couvrir les frais d'inscription théoriques avec non catégorisés si besoin
        nb_nouveaux_eleves = donnees_ecole['nouveaux_eleves']
        theorique_insc = Decimal('30000') * nb_nouveaux_eleves
        if frais_inscription == 0 and nb_nouveaux_eleves > 0 and non_categorises > 0:
            a_affecter = min(theorique_insc, non_categorises)
            frais_inscription += a_affecter
            non_categorises -= a_affecter

        # Plafond: ne jamais dépasser 30 000 GNF par nouvel élève
        if nb_nouveaux_eleves > 0 and frais_inscription > theorique_insc:
            excedent = frais_inscription - theorique_insc
            frais_inscription = theorique_insc
            scolarite += excedent

        # Cohérence: si 0 nouveaux élèves, ne pas compter des frais d'inscription → reclasser comme scolarité
        if nb_nouveaux_eleves == 0 and frais_inscription > 0:
            scolarite += frais_inscription
            frais_inscription = Decimal('0')

        # Ajouter le reste non catégorisé à la scolarité par défaut
        scolarite += non_categorises

        # Assigner les valeurs finales
        donnees_ecole['paiements']['frais_inscription'] = frais_inscription
        donnees_ecole['paiements']['scolarite'] = scolarite
        
        # (Supprimé) Frais de scolarité annuel ne figure pas dans le rapport journalier
        
        # Calcul: Reste à payer (élèves concernés par la journée)
        # Inclut: élèves ayant payé ce jour + nouveaux inscrits ce jour
        eleves_concernes_ids = set(
            paiements_jour.values_list('eleve_id', flat=True).distinct()
        )
        nouveaux_ids = set(Eleve.objects.filter(
            classe__ecole=ecole,
            date_inscription=date_rapport
        ).values_list('id', flat=True))
        eleves_concernes_ids |= nouveaux_ids

        # Déterminer l'année scolaire (pivot: août)
        if date_rapport.month >= 8:
            annee_scolaire = f"{date_rapport.year}-{date_rapport.year + 1}"
        else:
            annee_scolaire = f"{date_rapport.year - 1}-{date_rapport.year}"

        # Répartition des remises par classe (sur les paiements de la période du jour)
        remises_par_classe_map = {}
        try:
            remises_group = PaiementRemise.objects.filter(
                paiement__in=paiements_jour
            ).values(
                'paiement__eleve__classe_id'
            ).annotate(
                total=Sum('montant_remise')
            )
            remises_par_classe_map = {row['paiement__eleve__classe_id']: (row['total'] or Decimal('0')) for row in remises_group}
        except Exception:
            remises_par_classe_map = {}

        reste_a_payer = Decimal('0')
        total_du_concernes = Decimal('0')
        if eleves_concernes_ids:
            # Répartition par classe: accumuler par classe
            par_classe = {}
            qs_ech = EcheancierPaiement.objects.filter(
                eleve_id__in=list(eleves_concernes_ids),
                annee_scolaire=annee_scolaire
            ).values(
                'eleve__classe_id', 'eleve__classe__nom',
                'frais_inscription_du', 'tranche_1_due', 'tranche_2_due', 'tranche_3_due',
                'frais_inscription_paye', 'tranche_1_payee', 'tranche_2_payee', 'tranche_3_payee'
            )
            for row in qs_ech:
                classe_id = row.get('eleve__classe_id')
                classe_nom = row.get('eleve__classe__nom') or 'Classe'
                du = (row.get('frais_inscription_du') or Decimal('0')) \
                     + (row.get('tranche_1_due') or Decimal('0')) \
                     + (row.get('tranche_2_due') or Decimal('0')) \
                     + (row.get('tranche_3_due') or Decimal('0'))
                paye = (row.get('frais_inscription_paye') or Decimal('0')) \
                       + (row.get('tranche_1_payee') or Decimal('0')) \
                       + (row.get('tranche_2_payee') or Decimal('0')) \
                       + (row.get('tranche_3_payee') or Decimal('0'))
                solde = du - paye

                # Totaux généraux
                total_du_concernes += du
                if solde > 0:
                    reste_a_payer += solde

                # Accumulation par classe
                if classe_id not in par_classe:
                    par_classe[classe_id] = {
                        'classe': classe_nom,
                        'effectif': 0,
                        'total_du': Decimal('0'),
                        'total_paye': Decimal('0'),
                        'reste': Decimal('0'),
                        'remises': Decimal('0'),
                    }
                pc = par_classe[classe_id]
                pc['effectif'] += 1
                pc['total_du'] += du
                pc['total_paye'] += paye
                if solde > 0:
                    pc['reste'] += solde
                # Ajouter remises pour cette classe (agrégées sur les paiements du jour)
                try:
                    pc['remises'] = remises_par_classe_map.get(classe_id, pc['remises'])
                except Exception:
                    pass

            # Ranger la liste ordonnée par nom de classe
            donnees_ecole['classes'] = sorted(par_classe.values(), key=lambda x: x['classe'])
        donnees_ecole['paiements']['reste_a_payer'] = reste_a_payer
        donnees_ecole['paiements']['total_du_concernes'] = total_du_concernes
        
        # Dépenses: pas de répartition par école (le modèle n'est pas rattaché à Ecole)
        # On laisse 0 au niveau de l'école et on affiche un total global dans le résumé

        # États de salaire validés
        # Convertir la date en timezone aware pour éviter les warnings
        debut_jour = django_timezone.make_aware(datetime.combine(date_rapport, datetime.min.time()))
        fin_jour = django_timezone.make_aware(datetime.combine(date_rapport, datetime.max.time()))
        
        etats_jour = EtatSalaire.objects.filter(
            enseignant__ecole=ecole,
            date_validation__range=[debut_jour, fin_jour],
            valide=True
        )
        
        donnees_ecole['salaires']['etats_valides'] = etats_jour.count()
        donnees_ecole['salaires']['montant_total'] = etats_jour.aggregate(
            total=Sum('salaire_net')
        )['total'] or Decimal('0')
        
        donnees['ecoles'][ecole.id] = donnees_ecole
    
    return donnees

def generer_pdf_journalier(donnees, date_rapport):
    """Génère le PDF du rapport journalier"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Titre
    titre_style = ParagraphStyle(
        'TitreRapport',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.darkblue,
        alignment=1  # Centré
    )
    
    story.append(Paragraph(f"RAPPORT JOURNALIER - {date_rapport.strftime('%d/%m/%Y')}", titre_style))
    story.append(Spacer(1, 20))
    
    # Pour chaque école
    for ecole_id, donnees_ecole in donnees['ecoles'].items():
        # Titre de l'école
        story.append(Paragraph(f"École: {donnees_ecole['nom']}", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        # Tableau des données
        data = [
            ['Indicateur', 'Valeur'],
            ['Nouveaux élèves inscrits', str(donnees_ecole['nouveaux_eleves'])],
            ['Nombre de paiements', str(donnees_ecole['paiements']['nombre'])],
            ["Scolarité normale", f"{donnees_ecole['paiements'].get('total_du_concernes', Decimal('0')):,} GNF".replace(',', ' ')],
            ['Scolarité payé', f"{donnees_ecole['paiements']['scolarite']:,} GNF".replace(',', ' ')],
            ["Frais d'inscription", f"{donnees_ecole['paiements']['frais_inscription']:,} GNF".replace(',', ' ')],
            ["Reste à payer", f"{donnees_ecole['paiements'].get('reste_a_payer', Decimal('0')):,} GNF".replace(',', ' ')],
            ['Montant original (avant remises)', f"{donnees_ecole['paiements']['montant_original']:,} GNF".replace(',', ' ')],
            ['Total des remises accordées', f"{donnees_ecole['paiements']['total_remises']:,} GNF".replace(',', ' ')],
            ['Montant net encaissé', f"{donnees_ecole['paiements']['montant_total']:,} GNF".replace(',', ' ')],
            ['Nombre de dépenses', str(donnees_ecole['depenses']['nombre'])],
            ['Montant total des dépenses', f"{donnees_ecole['depenses']['montant_total']:,} GNF".replace(',', ' ')],
            ['États de salaire validés', str(donnees_ecole['salaires']['etats_valides'])],
            ['Montant total des salaires', f"{donnees_ecole['salaires']['montant_total']:,} GNF".replace(',', ' ')],
        ]
        
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 16))

        # Répartition par classe
        if donnees_ecole.get('classes'):
            story.append(Paragraph("Répartition par classe", styles['Heading3']))
            story.append(Spacer(1, 6))
            class_data = [[
                'Classe', 'Effectif', 'Total dû', 'Total payé', 'Remises', 'Reste à payer'
            ]]
            for c in donnees_ecole['classes']:
                class_data.append([
                    c['classe'],
                    str(c['effectif']),
                    f"{c['total_du']:,} GNF".replace(',', ' '),
                    f"{c['total_paye']:,} GNF".replace(',', ' '),
                    f"{c.get('remises', 0):,} GNF".replace(',', ' '),
                    f"{c['reste']:,} GNF".replace(',', ' '),
                ])

            class_table = Table(class_data, colWidths=[120, 60, 90, 90, 90, 90])
            class_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
                ('ALIGN', (0,0), (0,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ]))
            story.append(class_table)
            story.append(Spacer(1, 16))
    
    # Section Dépenses globales (non rattachées à une école)
    story.append(Paragraph("DÉPENSES GLOBALES (journée)", styles['Heading2']))
    story.append(Spacer(1, 8))
    dep_global = donnees.get('depenses_globales', {'nombre': 0, 'montant_total': Decimal('0')})
    depenses_data = [
        ['Nombre de dépenses', str(dep_global.get('nombre', 0))],
        ['Montant total des dépenses', f"{dep_global.get('montant_total', Decimal('0')):,} GNF".replace(',', ' ')]
    ]
    dep_table = Table(depenses_data, colWidths=[3*inch, 2*inch])
    dep_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(dep_table)
    story.append(Spacer(1, 16))

    # Résumé global
    total_paiements = sum(d['paiements']['montant_total'] for d in donnees['ecoles'].values())
    total_depenses = donnees['depenses_globales']['montant_total']
    total_paiements_original = sum(d['paiements']['montant_original'] for d in donnees['ecoles'].values())
    total_remises = sum(d['paiements']['total_remises'] for d in donnees['ecoles'].values())
    total_salaires = sum(d['salaires']['montant_total'] for d in donnees['ecoles'].values())
    
    # Résumé global
    resume_data = [
        ['Indicateur', 'Montant'],
        ['Montant original des paiements', f"{total_paiements_original:,} GNF".replace(',', ' ')],
        ['Total des remises accordées', f"{total_remises:,} GNF".replace(',', ' ')],
        ['Montant net des paiements', f"{total_paiements:,} GNF".replace(',', ' ')],
        ['Total des dépenses', f"{total_depenses:,} GNF".replace(',', ' ')],
        ['Total des salaires', f"{total_salaires:,} GNF".replace(',', ' ')],
        ['Solde net', f"{total_paiements - total_depenses - total_salaires:,} GNF".replace(',', ' ')],
    ]
    
    resume_table = Table(resume_data, colWidths=[3*inch, 2*inch])
    resume_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(resume_table)
    
    # Ajout entête + filigrane sur toutes les pages
    doc.build(story, onFirstPage=_draw_header_and_watermark, onLaterPages=_draw_header_and_watermark)
    buffer.seek(0)
    return buffer


@login_required
@admin_required
def rapport_remises_detaille(request):
    """Rapport détaillé des remises appliquées"""
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Dates par défaut (mois en cours)
    if not date_debut:
        date_debut = date.today().replace(day=1)
    else:
        date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
    
    if not date_fin:
        date_fin = date.today()
    else:
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
    
    # Récupérer toutes les remises appliquées dans la période
    remises_appliquees = PaiementRemise.objects.filter(
        paiement__date_paiement__range=[date_debut, date_fin],
        paiement__statut='VALIDE'
    ).select_related(
        'paiement', 'paiement__eleve', 'paiement__eleve__classe', 
        'paiement__eleve__classe__ecole', 'remise'
    ).order_by('-paiement__date_paiement')
    
    # Statistiques des remises
    stats_remises = {
        'total_remises': remises_appliquees.aggregate(
            total=Sum('montant_remise')
        )['total'] or Decimal('0'),
        'nombre_paiements_avec_remise': remises_appliquees.values('paiement').distinct().count(),
        'nombre_eleves_beneficiaires': remises_appliquees.values('paiement__eleve').distinct().count(),
    }
    
    # Répartition par type de remise
    repartition_types = remises_appliquees.values(
        'remise__nom', 'remise__motif'
    ).annotate(
        total_montant=Sum('montant_remise'),
        nombre_applications=Count('id')
    ).order_by('-total_montant')
    
    # Répartition par école
    repartition_ecoles = remises_appliquees.values(
        'paiement__eleve__classe__ecole__nom'
    ).annotate(
        total_montant=Sum('montant_remise'),
        nombre_applications=Count('id')
    ).order_by('-total_montant')
    
    context = {
        'remises_appliquees': remises_appliquees,
        'stats_remises': stats_remises,
        'repartition_types': repartition_types,
        'repartition_ecoles': repartition_ecoles,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'periode_str': f"du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    }
    
    return render(request, 'rapports/rapport_remises.html', context)
