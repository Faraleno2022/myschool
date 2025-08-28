from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Count
from django.http import JsonResponse, HttpResponse, Http404
from django.shortcuts import get_object_or_404, render, redirect
from django.utils import timezone
from django.db.models import Q, Sum, Count
from django.db import transaction, IntegrityError
from django.utils import timezone
from decimal import Decimal
from datetime import date, datetime
import os
import logging
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import F, ExpressionWrapper, DecimalField, Case, When, Value, Q, Sum, Count
from django.db.models.functions import Coalesce, Least
from django.db.models.functions import Greatest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from io import BytesIO
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.utils import ImageReader
except Exception:
    canvas = None
    A4 = (595.27, 841.89)
    ImageReader = None
from ecole_moderne.pdf_utils import draw_logo_watermark

from .models import Paiement, EcheancierPaiement, TypePaiement, ModePaiement, RemiseReduction, PaiementRemise, Relance, TwilioInboundMessage
from eleves.models import Eleve, GrilleTarifaire, Classe
from .forms import PaiementForm, EcheancierForm, RechercheForm
from .remise_forms import PaiementRemiseForm, CalculateurRemiseForm
from utilisateurs.utils import user_is_admin, filter_by_user_school, user_school
from utilisateurs.permissions import has_permission, can_add_payments, can_modify_payments, can_delete_payments, can_validate_payments, can_view_reports, can_apply_discounts
from .notifications import (
    send_payment_receipt,
    send_enrollment_confirmation,
    send_relance_notification,
    send_retard_notification,
)

# ... (rest of the code remains the same)

def ensure_echeancier_for_eleve(eleve: "Eleve", *, created_by=None) -> "EcheancierPaiement":
    """Crée (silencieusement) un `EcheancierPaiement` pour l'élève s'il n'existe pas.

    - Utilise `eleves.GrilleTarifaire` pour pré-remplir les montants dus et l'année scolaire
    - Définit des dates d'échéance par défaut: inscription=today, T1=15/01, T2=15/03, T3=15/05
    - Retourne l'échéancier existant ou nouvellement créé
    """
    try:
        ech = getattr(eleve, 'echeancier', None)
    except Exception:
        ech = None
    if ech:
        return ech

    # Déterminer la meilleure grille disponible
    try:
        niveau = getattr(eleve.classe, 'niveau', None)
        ecole = getattr(eleve.classe, 'ecole', None)
        annee_classe = getattr(eleve.classe, 'annee_scolaire', None)
    except Exception:
        niveau = None
        ecole = None
        annee_classe = None

    try:
        from datetime import date as _d
        today_d = _d.today()
    except Exception:
        from datetime import date as _d
        today_d = _d.today()

    annee_scolaire_def = f"{today_d.year}-{today_d.year+1}" if today_d.month >= 9 else f"{today_d.year-1}-{today_d.year}"
    grille = None
    try:
        if ecole and niveau:
            # 1) Grille exacte sur l'année de la classe
            if annee_classe:
                grille = GrilleTarifaire.objects.filter(ecole=ecole, niveau=niveau, annee_scolaire=annee_classe).first()
            # 2) Sinon année scolaire par défaut
            if grille is None:
                grille = GrilleTarifaire.objects.filter(ecole=ecole, niveau=niveau, annee_scolaire=annee_scolaire_def).first()
            # 3) Sinon la plus récente
            if grille is None:
                grille = GrilleTarifaire.objects.filter(ecole=ecole, niveau=niveau).order_by('-annee_scolaire').first()
    except Exception:
        grille = None

    # Préparer les champs
    if grille:
        annee_scol = grille.annee_scolaire
        fi = grille.frais_inscription or 0
        t1 = grille.tranche_1 or 0
        t2 = grille.tranche_2 or 0
        t3 = grille.tranche_3 or 0
    else:
        annee_scol = annee_classe or annee_scolaire_def
        fi = 0
        t1 = 0
        t2 = 0
        t3 = 0

    # Dates d'échéance par défaut
    try:
        try:
            annee_debut = int(str(annee_scol).split('-')[0])
        except Exception:
            annee_debut = today_d.year if today_d.month >= 9 else today_d.year - 1
        annee_fin = annee_debut + 1
        from datetime import date as _d
        d_insc = today_d
        d_t1 = _d(annee_fin, 1, 15)
        d_t2 = _d(annee_fin, 3, 15)
        d_t3 = _d(annee_fin, 5, 15)
    except Exception:
        d_insc = today_d
        d_t1 = today_d
        d_t2 = today_d
        d_t3 = today_d

    with transaction.atomic():
        ech = EcheancierPaiement.objects.create(
            eleve=eleve,
            annee_scolaire=annee_scol,
            frais_inscription_du=fi,
            tranche_1_due=t1,
            tranche_2_due=t2,
            tranche_3_due=t3,
            date_echeance_inscription=d_insc,
            date_echeance_tranche_1=d_t1,
            date_echeance_tranche_2=d_t2,
            date_echeance_tranche_3=d_t3,
            cree_par=created_by if created_by and getattr(created_by, 'is_authenticated', False) else None,
        )
    return ech

def _auto_validate_echeancier_for_eleve(eleve: "Eleve") -> None:
    """Synchronise l'échéancier de l'élève avec les paiements VALIDÉS avant impression du reçu.

    Règles conservatrices:
    - Si la somme des paiements validés + remises couvre le total dû -> statut = PAYE_COMPLET
      et on aligne les champs *_payee sur les *_due pour cohérence d'affichage.
    - Si couverture = 0 -> statut = A_PAYER (pas d'allocation détaillée effectuée ici)
    - Sinon -> statut = PAYE_PARTIEL (sans répartir finement par tranche)

    Cette fonction évite les incohérences si l'allocation manuelle par tranche a été oubliée.
    """
    try:
        # Récupérer l'échéancier (sans exception si absent)
        echeancier = getattr(eleve, 'echeancier', None)
        if echeancier is None:
            echeancier = EcheancierPaiement.objects.filter(eleve=eleve).first()
        if not echeancier:
            return

        # Totaux dus
        total_du = int((echeancier.frais_inscription_du or 0)
                       + (echeancier.tranche_1_due or 0)
                       + (echeancier.tranche_2_due or 0)
                       + (echeancier.tranche_3_due or 0))

        # Paiements validés et remises appliquées sur des paiements
        aggs = (
            Paiement.objects
            .filter(eleve=eleve, statut='VALIDE')
            .aggregate(sum_montant=Sum('montant'), sum_remises=Sum('remises__montant_remise'))
        )
        sum_montant = int(aggs.get('sum_montant') or 0)
        sum_remises = int(aggs.get('sum_remises') or 0)

        couverture = max(0, sum_montant + sum_remises)

        # Déterminer le nouveau statut avec gestion du retard
        # Calcul de l'exigible (sommes dont la date d'échéance est passée ou aujourd'hui)
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
        exigible = 0
        if echeancier.date_echeance_inscription and echeancier.date_echeance_inscription <= today:
            exigible += int(echeancier.frais_inscription_du or 0)
        if echeancier.date_echeance_tranche_1 and echeancier.date_echeance_tranche_1 <= today:
            exigible += int(echeancier.tranche_1_due or 0)
        if echeancier.date_echeance_tranche_2 and echeancier.date_echeance_tranche_2 <= today:
            exigible += int(echeancier.tranche_2_due or 0)
        if echeancier.date_echeance_tranche_3 and echeancier.date_echeance_tranche_3 <= today:
            exigible += int(echeancier.tranche_3_due or 0)

        # Allocation conservatrice sur les tranches SANS jamais réduire l'existant
        # 1) total déjà indiqué comme payé dans l'échéancier
        old_insc = int(echeancier.frais_inscription_paye or 0)
        old_t1 = int(echeancier.tranche_1_payee or 0)
        old_t2 = int(echeancier.tranche_2_payee or 0)
        old_t3 = int(echeancier.tranche_3_payee or 0)
        current_total_paid = max(0, old_insc + old_t1 + old_t2 + old_t3)

        # 2) Incrément à répartir basé sur la couverture réelle
        increment = max(0, couverture - current_total_paid)
        remaining = increment

        def _alloc(due: int, paid: int, remaining_local: int):
            due_i = int(due or 0)
            paid_i = int(paid or 0)
            room = max(0, due_i - paid_i)
            take = min(room, max(0, int(remaining_local)))
            return paid_i + take, remaining_local - take

        changed = False
        if remaining > 0:
            # Ordre: inscription -> T1 -> T2 -> T3
            new_insc, remaining = _alloc(echeancier.frais_inscription_du, old_insc, remaining)
            new_t1, remaining = _alloc(echeancier.tranche_1_due, old_t1, remaining)
            new_t2, remaining = _alloc(echeancier.tranche_2_due, old_t2, remaining)
            new_t3, remaining = _alloc(echeancier.tranche_3_due, old_t3, remaining)

            if new_insc != old_insc:
                echeancier.frais_inscription_paye = new_insc
                changed = True
            if new_t1 != old_t1:
                echeancier.tranche_1_payee = new_t1
                changed = True
            if new_t2 != old_t2:
                echeancier.tranche_2_payee = new_t2
                changed = True
            if new_t3 != old_t3:
                echeancier.tranche_3_payee = new_t3
                changed = True

        # Somme payée effective bornée au total dû
        paye_effectif = min(couverture, total_du)

        if total_du <= 0:
            new_statut = 'PAYE_COMPLET'
        elif paye_effectif >= total_du:
            new_statut = 'PAYE_COMPLET'
        elif exigible > 0 and paye_effectif < exigible:
            new_statut = 'EN_RETARD'
        elif paye_effectif <= 0:
            new_statut = 'A_PAYER'
        else:
            new_statut = 'PAYE_PARTIEL'

        # Appliquer le statut et éventuellement aligner les montants payés si totalement soldé
        # 'changed' peut déjà être True si allocation ci-dessus a modifié des champs
        if echeancier.statut != new_statut:
            echeancier.statut = new_statut
            changed = True
        if new_statut == 'PAYE_COMPLET':
            # Aligner les montants payés pour refléter le soldé complet
            if echeancier.frais_inscription_paye != echeancier.frais_inscription_du:
                echeancier.frais_inscription_paye = echeancier.frais_inscription_du
                changed = True
            if echeancier.tranche_1_payee != echeancier.tranche_1_due:
                echeancier.tranche_1_payee = echeancier.tranche_1_due
                changed = True
            if echeancier.tranche_2_payee != echeancier.tranche_2_due:
                echeancier.tranche_2_payee = echeancier.tranche_2_due
                changed = True
            if echeancier.tranche_3_payee != echeancier.tranche_3_due:
                echeancier.tranche_3_payee = echeancier.tranche_3_due
                changed = True

        if changed:
            echeancier.save()
    except Exception:
        # Ne jamais bloquer l'impression du reçu à cause de cette étape
        logging.getLogger(__name__).exception("Erreur lors de la validation automatique de l'échéancier")

@csrf_exempt
@require_http_methods(["POST"])
def twilio_inbound(request):
    """Réception des messages entrants Twilio (SMS/WhatsApp).
{{ ... }}
    Journalise les données utiles et répond 200.
    """
    if not _is_valid_twilio_request(request):
        return HttpResponse("Invalid signature", status=403)
    try:
        data = request.POST.dict()
    except Exception:
        data = {}
    # Champs utiles possibles: From, To, Body, SmsSid, MessageSid, WaId, NumMedia, etc.
    logging.getLogger(__name__).info("Twilio inbound message: %s", data)
    # Persist inbound message
    try:
        from_number = (data.get('From') or '').strip()
        to_number = (data.get('To') or '').strip()
        body = data.get('Body')
        message_sid = data.get('MessageSid') or data.get('SmsSid')
        wa_id = data.get('WaId')
        try:
            num_media = int(data.get('NumMedia') or 0)
        except Exception:
            num_media = 0
        channel = 'WHATSAPP' if from_number.lower().startswith('whatsapp:') else 'SMS'
        TwilioInboundMessage.objects.update_or_create(
            message_sid=message_sid,
            defaults={
                'channel': channel,
                'from_number': from_number,
                'to_number': to_number,
                'body': body,
                'wa_id': wa_id,
                'num_media': num_media,
                'raw_data': data,
            }
        )
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de l'enregistrement du message entrant Twilio")
    return JsonResponse({"status": "ok"})

@csrf_exempt
@require_http_methods(["POST"]) 
def twilio_status_callback(request):
    """Réception des callbacks de statut Twilio (optionnel).
    Journalise l'événement et répond 200.
    """
    if not _is_valid_twilio_request(request):
        return HttpResponse("Invalid signature", status=403)
    try:
        data = request.POST.dict()
    except Exception:
        data = {}
    logging.getLogger(__name__).info("Twilio status callback: %s", data)
    # Persist status update if MessageSid is present
    try:
        message_sid = data.get('MessageSid') or data.get('SmsSid')
        if message_sid:
            status = data.get('MessageStatus') or data.get('SmsStatus')
            error_code = data.get('ErrorCode')
            error_message = data.get('ErrorMessage')
            from django.utils import timezone as _tz
            obj, created = TwilioInboundMessage.objects.get_or_create(message_sid=message_sid, defaults={'raw_data': data})
            obj.delivery_status = status
            obj.error_code = str(error_code) if error_code is not None else obj.error_code
            obj.error_message = error_message or obj.error_message
            obj.status_updated_at = _tz.now()
            # Conserver dernières données brutes utiles
            try:
                merged = obj.raw_data or {}
                merged.update(data)
                obj.raw_data = merged
            except Exception:
                obj.raw_data = data
            obj.save()
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de l'enregistrement du status callback Twilio")
    return JsonResponse({"status": "ok"})

# ---------------------------------------------------------------
# Tableau de bord Paiements – statistiques réelles + listes
# ---------------------------------------------------------------

def _compute_stats():
    """Calcule les statistiques affichées sur le tableau de bord.
    Retourne un dict: total_paiements_mois, nombre_paiements_mois, eleves_en_retard, paiements_en_attente.
    """
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
    except Exception:
        today = date.today()

    # Début du mois courant
    try:
        month_start = today.replace(day=1)
    except Exception:
        # fallback simple
        month_start = date(today.year, today.month, 1)

    # Somme des paiements validés sur le mois (DateField -> filtre inclusif par bornes)
    total_mois = (
        Paiement.objects.filter(
            statut='VALIDE',
            date_paiement__gte=month_start,
            date_paiement__lte=today,
        ).aggregate(total=Sum('montant'))['total'] or 0
    )

    # Nombre de paiements (tous statuts) ce mois
    nb_paiements_mois = Paiement.objects.filter(
        date_paiement__gte=month_start,
        date_paiement__lte=today,
    ).count()

    # Élèves en retard: montants exigibles (échéances dépassées) > payés + remises
    exigible_expr = (
        Case(
            When(date_echeance_inscription__lte=today, then=F('frais_inscription_du')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_1__lte=today, then=F('tranche_1_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_2__lte=today, then=F('tranche_2_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_3__lte=today, then=F('tranche_3_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
    )
    remises_expr = Coalesce(
        Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
        Value(0),
        output_field=DecimalField(max_digits=10, decimal_places=0),
    )
    # Les remises ne doivent compenser que le montant exigible à date (pas les échéances futures)
    remises_applicables = Least(remises_expr, exigible_expr)
    paye_effectif_expr = (
        F('frais_inscription_paye') + F('tranche_1_payee') + F('tranche_2_payee') + F('tranche_3_payee')
        + remises_applicables
    )
    retard_expr = ExpressionWrapper(exigible_expr - paye_effectif_expr, output_field=DecimalField(max_digits=10, decimal_places=0))
    eleves_retard_count = (
        EcheancierPaiement.objects
        .annotate(retard=retard_expr)
        .filter(retard__gt=0)
        .count()
    )

    # Paiements en attente
    en_attente_count = Paiement.objects.filter(statut='EN_ATTENTE').count()

    return {
        'total_paiements_mois': int(total_mois or 0),
        'nombre_paiements_mois': int(nb_paiements_mois or 0),
        'eleves_en_retard': int(eleves_retard_count or 0),
        'paiements_en_attente': int(en_attente_count or 0),
    }


@login_required
def tableau_bord_paiements(request):
    """Affiche le tableau de bord des paiements avec stats et listes utiles."""
    if not _template_exists('paiements/tableau_bord.html'):
        return HttpResponse('Tableau de bord paiements (template manquant)')

    stats = _compute_stats()

    # Paiements récents: derniers validés d'abord, sinon tout, sur 30 jours sinon fallback 20 derniers
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
    except Exception:
        today = date.today()

    try:
        from datetime import timedelta
        last_30 = today - timedelta(days=30)
    except Exception:
        last_30 = today

    paiements_recents_qs = (
        Paiement.objects
        .select_related('eleve', 'type_paiement', 'mode_paiement')
        .filter(date_paiement__gte=last_30)
        .order_by('-date_paiement', '-date_creation')
    )
    if paiements_recents_qs.count() == 0:
        paiements_recents_qs = (
            Paiement.objects
            .select_related('eleve', 'type_paiement', 'mode_paiement')
            .order_by('-date_paiement', '-date_creation')
        )
    paiements_recents = list(paiements_recents_qs[:20])

    # Top élèves en retard (montant de retard décroissant)
    exigible_expr = (
        Case(
            When(date_echeance_inscription__lte=today, then=F('frais_inscription_du')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_1__lte=today, then=F('tranche_1_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_2__lte=today, then=F('tranche_2_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_3__lte=today, then=F('tranche_3_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
    )
    remises_expr = Coalesce(
        Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
        Value(0),
        output_field=DecimalField(max_digits=10, decimal_places=0),
    )
    # Les remises ne compensent que les montants exigibles au jour J
    remises_applicables = Least(remises_expr, exigible_expr)
    paye_effectif_expr = (
        F('frais_inscription_paye') + F('tranche_1_payee') + F('tranche_2_payee') + F('tranche_3_payee')
        + remises_applicables
    )
    retard_expr = ExpressionWrapper(exigible_expr - paye_effectif_expr, output_field=DecimalField(max_digits=10, decimal_places=0))
    eleves_en_retard = (
        EcheancierPaiement.objects
        .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
        .annotate(retard_db=retard_expr)
        .filter(retard_db__gt=0)
        .order_by('-retard_db')[:10]
    )

    context = {
        'titre_page': 'Tableau de bord des paiements',
        'stats': stats,
        'paiements_recents': paiements_recents,
        'eleves_en_retard': eleves_en_retard,
    }
    return render(request, 'paiements/tableau_bord.html', context)

@login_required
def liste_paiements(request):
    """Liste des paiements avec recherche, filtres de statut et pagination.
    Paramètres GET:
      - q: texte de recherche (élève: nom/prénom/matricule, reçu, référence, observations)
      - statut: EN_ATTENTE | VALIDE | REJETE | REMBOURSE (optionnel)
      - page: numéro de page
    """
    titre_page = "Liste des paiements"
    q = (request.GET.get('q') or '').strip()
    statut = (request.GET.get('statut') or '').strip()
    page = request.GET.get('page') or 1

    # Base queryset avec relations chargées
    qs = (
        Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement')
    )

    # Filtre recherche plein texte simple
    if q:
        qs = qs.filter(
            Q(numero_recu__icontains=q)
            | Q(reference_externe__icontains=q)
            | Q(observations__icontains=q)
            | Q(eleve__nom__icontains=q)
            | Q(eleve__prenom__icontains=q)
            | Q(eleve__matricule__icontains=q)
        )

    # Appliquer filtre par statut (si fourni) pour que les totaux reflètent la liste courante
    if statut:
        qs = qs.filter(statut=statut)

    # Calcul des totaux dynamiques (adaptés aux filtres en place)
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
    except Exception:
        today = date.today()
    try:
        month_start = today.replace(day=1)
    except Exception:
        month_start = date(today.year, today.month, 1)

    qs_effectif = qs
    qs_non_annule = qs_effectif.exclude(statut='ANNULE')

    total_paiements = qs_non_annule.count()
    montant_total = int(qs_non_annule.aggregate(total=Sum('montant'))['total'] or 0)
    montant_total_valide = int(qs_non_annule.filter(statut='VALIDE').aggregate(total=Sum('montant'))['total'] or 0)

    en_attente_qs = qs_effectif.filter(statut='EN_ATTENTE')
    total_en_attente = en_attente_qs.count()
    montant_en_attente = int(en_attente_qs.aggregate(total=Sum('montant'))['total'] or 0)

    ce_mois_qs = qs_non_annule.filter(date_paiement__gte=month_start, date_paiement__lte=today)
    total_ce_mois = ce_mois_qs.count()
    montant_ce_mois = int(ce_mois_qs.aggregate(total=Sum('montant'))['total'] or 0)

    # Calculs supplémentaires: Dû scolarité net après remises + frais d'inscription (réels depuis l'échéancier)
    eleves_qs = Eleve.objects.all()
    if q:
        eleves_qs = eleves_qs.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(matricule__icontains=q)
            | Q(classe__nom__icontains=q) | Q(classe__ecole__nom__icontains=q)
            | Q(paiements__numero_recu__icontains=q) | Q(paiements__reference_externe__icontains=q)
            | Q(paiements__observations__icontains=q)
        ).distinct()

    eleves_count = eleves_qs.count() if q else Eleve.objects.count()

    eche_qs = EcheancierPaiement.objects.filter(eleve__in=eleves_qs)
    dues_sco_expr = (
        Coalesce(
            F('tranche_1_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_2_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_3_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
    )
    remises_expr = Coalesce(
        Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
        Value(0),
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    aggr_du = eche_qs.aggregate(
        dues_sco=Coalesce(
            Sum(dues_sco_expr, output_field=DecimalField(max_digits=12, decimal_places=0)),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        remises=remises_expr,
    )
    dues_sco_total = int(aggr_du.get('dues_sco') or 0)
    remises_total = int(aggr_du.get('remises') or 0)
    du_sco_net = max(dues_sco_total - remises_total, 0)
    frais_inscription_total = int(
        eche_qs.aggregate(
            total=Coalesce(
                Sum(F('frais_inscription_du'), output_field=DecimalField(max_digits=12, decimal_places=0)),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            )
        ).get('total')
        or 0
    )
    du_global_net = du_sco_net + frais_inscription_total

    # Détail par école/classe (filtre libre appliqué aux élèves)
    detail_qs = (
        eche_qs
        .values(
            'eleve__classe__ecole__id', 'eleve__classe__ecole__nom',
            'eleve__classe__id', 'eleve__classe__nom'
        )
        .annotate(
            eleves_count=Count('eleve', distinct=True),
            dues_sco_sum=Coalesce(
                Sum(dues_sco_expr, output_field=DecimalField(max_digits=12, decimal_places=0)),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
            remises_sum=remises_expr,
            frais_insc_sum=Coalesce(
                Sum(F('frais_inscription_du'), output_field=DecimalField(max_digits=12, decimal_places=0)),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
        )
        .order_by('eleve__classe__ecole__nom', 'eleve__classe__nom')
    )
    totaux_du_detail_classes = []
    for row in detail_qs:
        dues = int(row.get('dues_sco_sum') or 0)
        rem = int(row.get('remises_sum') or 0)
        net_sco = max(dues - rem, 0)
        cnt = int(row.get('eleves_count') or 0)
        insc = int(row.get('frais_insc_sum') or 0)
        tot = net_sco + insc
        totaux_du_detail_classes.append({
            'ecole_id': row.get('eleve__classe__ecole__id'),
            'ecole_nom': row.get('eleve__classe__ecole__nom'),
            'classe_id': row.get('eleve__classe__id'),
            'classe_nom': row.get('eleve__classe__nom'),
            'eleves_count': cnt,
            'du_sco_net': net_sco,
            'frais_inscription_total': insc,
            'du_global_net': tot,
        })

    # Pagination
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(page)

    context = {
        'titre_page': titre_page,
        'q': q,
        'statut': statut,
        'paiements': page_obj.object_list,
        'page_obj': page_obj,
        # Totaux pour l'UI (utilisés par _paiements_resultats.html)
        'totaux': {
            'total_paiements': int(total_paiements or 0),
            'montant_total': int(montant_total or 0),
            'montant_total_valide': int(montant_total_valide or 0),
            'total_en_attente': int(total_en_attente or 0),
            'montant_en_attente': int(montant_en_attente or 0),
            'total_ce_mois': int(total_ce_mois or 0),
            'montant_ce_mois': int(montant_ce_mois or 0),
        },
        'totaux_du': {
            'eleves_count': int(eleves_count or 0),
            'du_sco_net': int(du_sco_net or 0),
            'frais_inscription_total': int(frais_inscription_total or 0),
            'du_global_net': int(du_global_net or 0),
        },
        'totaux_du_detail_classes': totaux_du_detail_classes,
        # Alerte relance en haut du fragment (compte global des élèves en retard)
        'eleves_en_retard': _compute_stats().get('eleves_en_retard', 0),
    }

    # Réponse partielle pour les requêtes AJAX (utilisé par la recherche/pagination dynamique)
    try:
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    except Exception:
        is_ajax = False
    if is_ajax and _template_exists('paiements/_paiements_resultats.html'):
        return render(request, 'paiements/_paiements_resultats.html', context)

    template = 'paiements/liste_paiements.html' if _template_exists('paiements/liste_paiements.html') else None
    if template:
        return render(request, template, context)
    return HttpResponse('Liste des paiements')

@login_required
def detail_paiement(request, paiement_id:int):
    """Affiche le détail d'un paiement.

    Contexte pour `templates/paiements/detail_paiement.html`:
      - titre_page: str
      - paiement: instance `Paiement`
      - is_admin: bool
      - user_permissions: dict avec `can_validate_payments`
    """
    paiement = get_object_or_404(
        Paiement.objects.select_related(
            'eleve', 'type_paiement', 'mode_paiement',
            'eleve__classe', 'eleve__classe__ecole',
        ),
        pk=paiement_id,
    )

    # Préparer les informations de permissions utilisées dans le template
    perms_ctx = {
        'can_validate_payments': can_validate_payments(request.user) if request.user.is_authenticated else False,
    }

    # Total des remises appliquées sur ce paiement
    try:
        remises_total = (
            paiement.remises.aggregate(total=Sum('montant_remise')).get('total') or 0
        )
    except Exception:
        remises_total = 0

    context = {
        'titre_page': f"Détail du paiement #{paiement.id}",
        'paiement': paiement,
        'is_admin': user_is_admin(request.user) if request.user.is_authenticated else False,
        'user_permissions': perms_ctx,
        'remises_total': int(remises_total or 0),
    }
    return render(request, 'paiements/detail_paiement.html', context)

@login_required
def ajouter_paiement(request, eleve_id:int=None):
    """Créer un paiement.
    - GET: affiche le formulaire `templates/paiements/form_paiement.html`
    - POST: enregistre le paiement en statut EN_ATTENTE
    """
    titre_page = "Ajouter un paiement"
    action = "Enregistrer"

    eleve = None
    initial = {}
    if eleve_id:
        eleve = get_object_or_404(Eleve.objects.select_related('classe', 'classe__ecole'), pk=eleve_id)
        initial['eleve'] = eleve

    if request.method == 'POST':
        form = PaiementForm(request.POST)
        if form.is_valid():
            # Pré-valider la cohérence métier avant d'enregistrer
            paiement: Paiement = form.save(commit=False)

            # Récupérer/assurer l'échéancier de l'élève pour les contrôles
            try:
                ech = getattr(paiement.eleve, 'echeancier', None)
            except Exception:
                ech = None
            if not ech:
                try:
                    ech = ensure_echeancier_for_eleve(
                        paiement.eleve,
                        created_by=request.user if request.user.is_authenticated else None,
                    )
                except Exception:
                    ech = None

            # Si on ne parvient pas à obtenir un échéancier, on empêche un enregistrement potentiellement incohérent
            if not ech:
                messages.error(request, "Impossible de récupérer l'échéancier de l'élève. Réessayez ou créez-le d'abord.")
                return render(request, 'paiements/form_paiement.html', {
                    'titre_page': titre_page,
                    'action': action,
                    'form': form,
                    'eleve': eleve,
                })

            # 1) Bloquer si la tranche ciblée est déjà soldée
            type_nom = (getattr(paiement.type_paiement, 'nom', '') or '').strip().lower()
            try:
                fi_due = int(ech.frais_inscription_du or 0)
                fi_payee = int(ech.frais_inscription_paye or 0)
                t1_due = int(ech.tranche_1_due or 0)
                t1_payee = int(ech.tranche_1_payee or 0)
                t2_due = int(ech.tranche_2_due or 0)
                t2_payee = int(ech.tranche_2_payee or 0)
                t3_due = int(ech.tranche_3_due or 0)
                t3_payee = int(ech.tranche_3_payee or 0)
            except Exception:
                fi_due = fi_payee = t1_due = t1_payee = t2_due = t2_payee = t3_due = t3_payee = 0

            if 'inscription' in type_nom and fi_payee >= fi_due:
                messages.error(request, "L'inscription est déjà totalement payée pour cet élève. Aucune somme supplémentaire n'est autorisée pour l'inscription.")
                return render(request, 'paiements/form_paiement.html', {
                    'titre_page': titre_page,
                    'action': action,
                    'form': form,
                    'eleve': eleve,
                })
            if ('tranche 1' in type_nom or '1ère tranche' in type_nom or '1ere tranche' in type_nom) and t1_payee >= t1_due:
                messages.error(request, "La 1ère tranche est déjà totalement payée pour cet élève.")
                return render(request, 'paiements/form_paiement.html', {
                    'titre_page': titre_page,
                    'action': action,
                    'form': form,
                    'eleve': eleve,
                })
            if ('tranche 2' in type_nom or '2ème tranche' in type_nom or '2eme tranche' in type_nom) and t2_payee >= t2_due:
                messages.error(request, "La 2ème tranche est déjà totalement payée pour cet élève.")
                return render(request, 'paiements/form_paiement.html', {
                    'titre_page': titre_page,
                    'action': action,
                    'form': form,
                    'eleve': eleve,
                })
            if ('tranche 3' in type_nom or '3ème tranche' in type_nom or '3eme tranche' in type_nom) and t3_payee >= t3_due:
                messages.error(request, "La 3ème tranche est déjà totalement payée pour cet élève.")
                return render(request, 'paiements/form_paiement.html', {
                    'titre_page': titre_page,
                    'action': action,
                    'form': form,
                    'eleve': eleve,
                })

            # 2) Bloquer les sur-paiements par rapport au total annuel dû (incluant inscription + tranches)
            try:
                total_du = int((ech.frais_inscription_du or 0) + (ech.tranche_1_due or 0) + (ech.tranche_2_due or 0) + (ech.tranche_3_due or 0))
            except Exception:
                total_du = 0

            try:
                aggs = (
                    Paiement.objects
                    .filter(eleve=paiement.eleve)
                    .exclude(statut__in=['REJETE', 'REMBOURSE'])
                    .aggregate(sum_montant=Sum('montant'))
                )
                deja_saisi = int(aggs.get('sum_montant') or 0)
            except Exception:
                deja_saisi = 0

            try:
                remises_valides = (
                    Paiement.objects
                    .filter(eleve=paiement.eleve, statut='VALIDE')
                    .aggregate(total=Sum('remises__montant_remise'))
                    .get('total') or 0
                )
                remises_valides = int(remises_valides)
            except Exception:
                remises_valides = 0

            montant_soumis = int(paiement.montant or 0)
            restant_global = max(0, total_du - max(0, deja_saisi + remises_valides))
            if montant_soumis > restant_global:
                # Message précis avec le plafond autorisé restant
                try:
                    montant_autorise = max(0, restant_global)
                except Exception:
                    montant_autorise = 0
                messages.error(
                    request,
                    f"Montant trop élevé: le reste total à payer pour cet élève est de {montant_autorise:,} GNF. Veuillez saisir un montant inférieur ou égal.",
                )
                return render(request, 'paiements/form_paiement.html', {
                    'titre_page': titre_page,
                    'action': action,
                    'form': form,
                    'eleve': eleve,
                })

            # Si tout est cohérent, on peut enregistrer
            with transaction.atomic():
                # Attacher l'utilisateur créateur si connecté
                if request.user.is_authenticated:
                    paiement.cree_par = request.user
                # Statut par défaut reste EN_ATTENTE (défini dans le modèle)
                paiement.save()
                # Auto-création de l'échéancier s'il n'existe pas, puis synchro/validation
                try:
                    _auto_validate_echeancier_for_eleve(paiement.eleve)
                except Exception:
                    logging.getLogger(__name__).exception("Auto-validation échéancier après enregistrement paiement")
            # Notifications: reçu paiement (WhatsApp + SMS) et, si inscription, confirmation d'inscription
            try:
                send_payment_receipt(paiement.eleve, paiement)
                type_nom = (getattr(paiement.type_paiement, 'nom', '') or '').strip().lower()
                if 'inscription' in type_nom:
                    send_enrollment_confirmation(paiement.eleve, paiement)
            except Exception:
                logging.getLogger(__name__).exception("Erreur lors de l'envoi des notifications Twilio")
            messages.success(request, "Paiement enregistré avec succès.")
            # Rediriger vers la page échéancier de l'élève
            return redirect('paiements:echeancier_eleve', eleve_id=paiement.eleve_id)
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = PaiementForm(initial=initial)
        # Si l'élève est imposé, fixer la valeur initiale proprement
        if eleve:
            form.fields['eleve'].initial = eleve

    context = {
        'titre_page': titre_page,
        'action': action,
        'form': form,
        'eleve': eleve,
    }
    return render(request, 'paiements/form_paiement.html', context)

@login_required
def valider_paiement(request, paiement_id:int):
    """Valide un paiement en le passant au statut VALIDE.

    - Vérifie les permissions: admin ou can_validate_payments
    - Met à jour: statut, date_validation, valide_par, date_modification
    - Optionnel: tente d'allouer le paiement à l'échéancier si une fonction utilitaire existe
    - Notifie le responsable (WhatsApp/SMS) avec le reçu
    """
    paiement = get_object_or_404(Paiement, pk=paiement_id)

    if not request.user.is_authenticated or not (user_is_admin(request.user) or can_validate_payments(request.user)):
        messages.error(request, "Vous n'avez pas l'autorisation de valider ce paiement.")
        return redirect('paiements:detail_paiement', paiement_id=paiement.id)

    if paiement.statut == 'VALIDE':
        messages.info(request, "Ce paiement est déjà validé.")
        return redirect('paiements:detail_paiement', paiement_id=paiement.id)

    with transaction.atomic():
        paiement.statut = 'VALIDE'
        try:
            paiement.date_validation = timezone.now()
        except Exception:
            from django.utils import timezone as _tz
            paiement.date_validation = _tz.now()
        paiement.valide_par = request.user
        try:
            paiement.date_modification = timezone.now()
        except Exception:
            pass
        paiement.save()

        # Tenter l'allocation à l'échéancier si une fonction utilitaire existe
        try:
            _allocate_payment_to_echeancier  # type: ignore[name-defined]
        except NameError:
            _allocate_payment_to_echeancier = None  # type: ignore[assignment]
        try:
            if _allocate_payment_to_echeancier:
                _allocate_payment_to_echeancier(paiement)  # type: ignore[misc]
        except Exception:
            logging.getLogger(__name__).exception("Erreur lors de l'allocation du paiement à l'échéancier")

        # S'assurer que l'échéancier existe et synchroniser le statut (incl. EN_RETARD)
        try:
            ensure_echeancier_for_eleve(paiement.eleve, created_by=request.user if request.user.is_authenticated else None)
            _auto_validate_echeancier_for_eleve(paiement.eleve)
        except Exception:
            logging.getLogger(__name__).exception("Erreur ensure/auto-validate échéancier après validation du paiement")

    # Envoyer le reçu de paiement après validation
    try:
        send_payment_receipt(paiement.eleve, paiement)
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de l'envoi du reçu après validation")

    messages.success(request, "Paiement validé avec succès.")
    return redirect('paiements:detail_paiement', paiement_id=paiement.id)

@login_required
def relancer_eleve(request, eleve_id:int):
    """Crée une relance et envoie la notification (WhatsApp/SMS) au responsable.
    GET params optionnels:
      - canal: SMS | WHATSAPP (par défaut WHATSAPP)
      - message: texte personnalisé
    """
    eleve = get_object_or_404(Eleve.objects.select_related('classe'), pk=eleve_id)
    canal = (request.GET.get('canal') or 'WHATSAPP').upper()
    message_txt = (request.GET.get('message') or '').strip()

    # Solde estimé depuis l'échéancier
    try:
        echeancier = getattr(eleve, 'echeancier', None)
        solde_estime = echeancier.solde_restant if echeancier else 0
    except Exception:
        solde_estime = 0

    if not message_txt:
        message_txt = (
            f"Merci de régulariser les frais de scolarité. Élève: {eleve.nom_complet} ({eleve.matricule})."
        )

    with transaction.atomic():
        relance = Relance.objects.create(
            eleve=eleve,
            canal=canal if canal in {c for c, _ in Relance.CANAL_CHOICES} else 'AUTRE',
            message=message_txt,
            statut='ENREGISTREE',
            solde_estime=solde_estime or 0,
            cree_par=request.user if request.user.is_authenticated else None,
        )
    try:
        send_relance_notification(relance)
        messages.success(request, "Relance créée et notification envoyée.")
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de l'envoi de la relance Twilio")
        messages.warning(request, "Relance créée mais l'envoi de la notification a échoué.")

    # Redirection: détail élève si dispo, sinon liste paiements
    try:
        return redirect('eleves:detail', eleve_id=eleve.id)
    except Exception:
        return redirect('paiements:liste_paiements')

@login_required
def envoyer_notifs_retards(request):
    """Envoie des notifications de retard aux responsables des élèves avec solde > 0.
    Action manuelle: GET uniquement, simple résumé via messages.
    """
    if not request.user.is_authenticated:
        return HttpResponse(status=403)
    # Optionnel: restreindre aux admins/permissions
    if not (user_is_admin(request.user) or can_view_reports(request.user)):
        return HttpResponse(status=403)

    # Annoter montant de retard au niveau DB: (exigible - (payé + remises))
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
    except Exception:
        today = date.today()

    exigible_expr = (
        Case(
            When(date_echeance_inscription__lte=today, then=F('frais_inscription_du')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_1__lte=today, then=F('tranche_1_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_2__lte=today, then=F('tranche_2_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_3__lte=today, then=F('tranche_3_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
    )
    remises_expr = Coalesce(
        Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
        Value(0),
        output_field=DecimalField(max_digits=10, decimal_places=0),
    )
    # Limiter l'effet des remises au montant actuellement exigible
    remises_applicables = Least(remises_expr, exigible_expr)
    paye_effectif_expr = (
        F('frais_inscription_paye') + F('tranche_1_payee') + F('tranche_2_payee') + F('tranche_3_payee')
        + remises_applicables
    )
    retard_expr = ExpressionWrapper(exigible_expr - paye_effectif_expr, output_field=DecimalField(max_digits=10, decimal_places=0))
    qs = (
        EcheancierPaiement.objects.select_related('eleve', 'eleve__classe')
        .annotate(retard=retard_expr)
        .filter(retard__gt=0)
    )
    envoyes = 0
    for ech in qs[:500]:  # sécurité: batch max 500
        try:
            send_retard_notification(ech.eleve, ech.retard)
            envoyes += 1
        except Exception:
            logging.getLogger(__name__).exception("Échec envoi retard pour %s", getattr(ech.eleve, 'nom_complet', 'eleve'))
            continue
    messages.info(request, f"Notifications de retard envoyées: {envoyes} (sur {qs.count()} éligibles)")
    # Rediriger vers relances ou tableau de bord
    return redirect('paiements:liste_relances')

@login_required
def liste_relances(request):
    """Liste des relances avec filtres et pagination."""
    titre_page = "Liste des relances"
    q = (request.GET.get('q') or '').strip()
    canal = (request.GET.get('canal') or '').strip().upper()
    statut = (request.GET.get('statut') or '').strip().upper()

    qs = (
        Relance.objects.select_related('eleve', 'eleve__classe')
        .order_by('-date_creation')
    )
    if q:
        qs = qs.filter(
            Q(eleve__nom__icontains=q)
            | Q(eleve__prenom__icontains=q)
            | Q(eleve__matricule__icontains=q)
            | Q(message__icontains=q)
        )
    if canal:
        qs = qs.filter(canal=canal)
    if statut:
        qs = qs.filter(statut=statut)

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    context = {
        'titre_page': titre_page,
        'q': q,
        'canal': canal,
        'statut': statut,
        'page_obj': page_obj,
    }
    template = 'paiements/relances.html' if _template_exists('paiements/relances.html') else None
    if template:
        return render(request, template, context)
    return HttpResponse('Liste des relances')

@login_required
def echeancier_eleve(request, eleve_id:int):
    """Affiche l'échéancier et l'historique des paiements d'un élève.

    Contexte fourni au template `templates/paiements/echeancier_eleve.html`:
      - titre_page: Titre de la page
      - eleve: instance d'`Eleve`
      - echeancier: instance d'`EcheancierPaiement` (ou None si non créé)
      - paiements: queryset des `Paiement` liés à l'élève (ordonnés récents d'abord)
      - today: date du jour (timezone-aware localdate si dispo)
    """
    # Récupération de l'élève avec sa classe/école pour l'en-tête
    eleve = get_object_or_404(Eleve.objects.select_related('classe', 'classe__ecole'), pk=eleve_id)

    # Échéancier (peut ne pas exister encore)
    try:
        echeancier = getattr(eleve, 'echeancier', None)
    except Exception:
        echeancier = None

    # Historique des paiements (les plus récents d'abord)
    paiements = (
        Paiement.objects
        .select_related('type_paiement', 'mode_paiement')
        .filter(eleve=eleve)
        .order_by('-date_paiement', '-date_creation')
    )

    # Date du jour pour l'affichage des retards
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
    except Exception:
        today = date.today()

    context = {
        'titre_page': "Échéancier des paiements",
        'eleve': eleve,
        'echeancier': echeancier,
        'paiements': paiements,
        'today': today,
    }
    return render(request, 'paiements/echeancier_eleve.html', context)

def creer_echeancier(request, eleve_id:int):
    """Créer ou éditer l'échéancier d'un élève.

    - Si un échéancier existe déjà: redirige vers la page d'échéancier avec message.
    - GET: affiche `templates/paiements/form_echeancier.html` pré-rempli si possible par la grille tarifaire.
    - POST: valide et enregistre l'échéancier puis redirige vers la page d'échéancier de l'élève.
    """
    eleve = get_object_or_404(Eleve.objects.select_related('classe', 'classe__ecole'), pk=eleve_id)

    # Si un échéancier existe déjà, on informe et on redirige
    if getattr(eleve, 'echeancier', None):
        messages.info(request, "Un échéancier existe déjà pour cet élève.")
        return redirect('paiements:echeancier_eleve', eleve_id=eleve.id)

    # Pré-remplissage depuis la grille tarifaire si disponible
    initial = {}
    try:
        niveau = getattr(eleve.classe, 'niveau', None)
        ecole = getattr(eleve.classe, 'ecole', None)
        # Année scolaire préférée: celle de la classe de l'élève, sinon calcul par date
        today = date.today()
        annee_scolaire_def = f"{today.year}-{today.year+1}" if today.month >= 9 else f"{today.year-1}-{today.year}"
        annee_classe = getattr(eleve.classe, 'annee_scolaire', None)
        from eleves.models import GrilleTarifaire as _Grille
        grille = None
        # 1) Essayer l'année de la classe si présente
        if annee_classe:
            grille = _Grille.objects.filter(ecole=ecole, niveau=niveau, annee_scolaire=annee_classe).first()
            if grille is None:
                messages.info(request, f"Aucune grille trouvée pour l'année {annee_classe}. Recherche d'une autre année...")
        # 2) Sinon essayer l'année par défaut calculée (ou si 1) a échoué et diffère)
        if grille is None:
            if not annee_classe or (annee_classe and annee_classe != annee_scolaire_def):
                grille = _Grille.objects.filter(ecole=ecole, niveau=niveau, annee_scolaire=annee_scolaire_def).first()
                if grille and annee_classe and annee_classe != annee_scolaire_def:
                    messages.info(request, f"Utilisation de la grille {grille.annee_scolaire} (aucune pour {annee_classe}).")
        # 3) Fallback: prendre la plus récente disponible pour l'école/niveau
        if grille is None:
            grille = _Grille.objects.filter(ecole=ecole, niveau=niveau).order_by('-annee_scolaire').first()
            if grille:
                messages.warning(request, f"Grille exacte introuvable. Utilisation de la plus récente: {grille.annee_scolaire}.")

        if grille:
            initial.update({
                'annee_scolaire': grille.annee_scolaire,
                'frais_inscription_du': grille.frais_inscription,
                'tranche_1_due': grille.tranche_1,
                'tranche_2_due': grille.tranche_2,
                'tranche_3_due': grille.tranche_3,
            })
        # Proposer des dates d'échéance par défaut
        try:
            # Inscription: aujourd'hui, puis jalons (janvier/mars) pour les tranches
            from datetime import date as _d
            today_d = _d.today()
            initial.setdefault('date_echeance_inscription', today_d)
            # 15 janvier, 15 mars, 15 mai de l'année de fin de l'année scolaire (annee_debut + 1)
            annee_scol = (initial.get('annee_scolaire') or annee_scolaire_def)
            try:
                annee_debut = int(str(annee_scol).split('-')[0])
            except Exception:
                annee_debut = today_d.year
            annee_fin = annee_debut + 1
            initial.setdefault('date_echeance_tranche_1', _d(annee_fin, 1, 15))
            initial.setdefault('date_echeance_tranche_2', _d(annee_fin, 3, 15))
            # Dernière tranche: 15 mai
            initial.setdefault('date_echeance_tranche_3', _d(annee_fin, 5, 15))
        except Exception:
            pass
    except Exception:
        grille = None

    if request.method == 'POST':
        form = EcheancierForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                ech: EcheancierPaiement = form.save(commit=False)
                ech.eleve = eleve
                if request.user.is_authenticated:
                    ech.cree_par = request.user
                ech.save()
            messages.success(request, "Échéancier créé avec succès.")
            return redirect('paiements:echeancier_eleve', eleve_id=eleve.id)
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = EcheancierForm(initial=initial)

    context = {
        'titre_page': "Créer un échéancier",
        'eleve': eleve,
        'form': form,
        'grille': grille if 'grille' in locals() else None,
        'action': 'Créer',
    }
    return render(request, 'paiements/form_echeancier.html', context)

@login_required
def assurer_echeancier(request, eleve_id: int):
    """Assure la création automatique de l'échéancier si manquant, puis redirige vers la page échéancier.

    Utilise `ensure_echeancier_for_eleve()` pour créer silencieusement à partir de la grille tarifaire.
    """
    eleve = get_object_or_404(Eleve.objects.select_related('classe', 'classe__ecole'), pk=eleve_id)
    try:
        ensure_echeancier_for_eleve(
            eleve,
            created_by=request.user if getattr(request.user, 'is_authenticated', False) else None,
        )
        # Synchroniser le statut juste après
        _auto_validate_echeancier_for_eleve(eleve)
        messages.success(request, "Échéancier créé/mis à jour automatiquement.")
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de l'assurance de l'échéancier")
        messages.error(request, "Impossible de créer automatiquement l'échéancier. Veuillez réessayer ou le créer manuellement.")
    return redirect('paiements:echeancier_eleve', eleve_id=eleve.id)

@login_required
def valider_echeancier(request, eleve_id: int):
    """Valide/synchronise l'échéancier d'un élève sur la base des paiements validés.

    - POST requis pour exécuter l'action
    - Vérifie l'autorisation via `can_validate_payments`
    - Utilise `_auto_validate_echeancier_for_eleve` pour ajuster le statut et montants payés si nécessaire
    - Redirige ensuite vers la page `echeancier_eleve`
    """
    # Autorisation
    if not has_permission(request.user, 'peut_valider_paiements'):
        messages.error(request, "Vous n'avez pas l'autorisation de valider les échéanciers.")
        return redirect('paiements:echeancier_eleve', eleve_id=eleve_id)

    # Méthode HTTP
    if request.method != 'POST':
        messages.warning(request, "Action invalide: la validation doit être envoyée en POST.")
        return redirect('paiements:echeancier_eleve', eleve_id=eleve_id)

    eleve = get_object_or_404(Eleve.objects.select_related('classe', 'classe__ecole'), pk=eleve_id)
    try:
        with transaction.atomic():
            # S'assurer qu'un échéancier existe d'abord
            ensure_echeancier_for_eleve(
                eleve,
                created_by=request.user if getattr(request.user, 'is_authenticated', False) else None,
            )
            # Puis synchroniser/valider
            _auto_validate_echeancier_for_eleve(eleve)
        messages.success(request, "Échéancier validé et synchronisé avec les paiements.")
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de la validation/synchronisation de l'échéancier")
        messages.error(request, "Une erreur est survenue lors de la validation de l'échéancier.")
    # Nouveau flux: si l'élève a un paiement récent en attente sans remise, rediriger vers son détail
    try:
        paiement_en_attente = (
            Paiement.objects
            .filter(eleve=eleve, statut='EN_ATTENTE')
            .order_by('-date_paiement', '-date_creation', '-id')
            .first()
        )
    except Exception:
        paiement_en_attente = None

    if paiement_en_attente:
        try:
            nb_remises = paiement_en_attente.remises.count()
        except Exception:
            nb_remises = 0
        if (nb_remises or 0) == 0:
            messages.info(request, "Aucune remise appliquée: veuillez valider le paiement en attente.")
            return redirect('paiements:detail_paiement', paiement_id=paiement_en_attente.id)

    return redirect('paiements:echeancier_eleve', eleve_id=eleve.id)

def generer_recu_pdf(request, paiement_id:int):
    """Génère un reçu PDF téléchargeable pour un paiement validé.

    - Ajoute un filigrane via `ecole_moderne/pdf_utils.draw_logo_watermark`
    - Inclut les informations clés du paiement et de l'élève
    - Liste les remises appliquées et affiche le total des remises
    """
    paiement = get_object_or_404(
        Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement', 'eleve__classe', 'eleve__classe__ecole'),
        pk=paiement_id,
    )

    # Optionnel: n'autoriser le reçu que pour les paiements validés
    if getattr(paiement, 'statut', 'EN_ATTENTE') != 'VALIDE':
        messages.warning(request, "Le reçu n'est disponible que pour les paiements validés.")
        return redirect('paiements:detail_paiement', paiement_id=paiement.id)

    if canvas is None:
        return HttpResponse("La génération de PDF n'est pas disponible sur ce serveur (ReportLab manquant).", status=500)

    # Valider/synchroniser l'échéancier de l'élève avant génération du reçu
    try:
        with transaction.atomic():
            _auto_validate_echeancier_for_eleve(paiement.eleve)
    except Exception:
        logging.getLogger(__name__).exception("Validation automatique de l'échéancier avant reçu échouée")

    # Calcul total remises
    remises_total = paiement.remises.aggregate(total=Sum('montant_remise')).get('total') or 0

    # Préparer le buffer et le canvas
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Filigrane: toujours actif pour les reçus PDF (opacité optimisée définie dans pdf_utils)
    try:
        draw_logo_watermark(c, width, height)
    except Exception:
        pass

    # Mise en page simple
    left = 40
    top = height - 40
    line_h = 18

    def draw_line(text, x=left, y=None, bold=False):
        nonlocal top
        if y is None:
            y = top
        font_name = 'Helvetica-Bold' if bold else 'Helvetica'
        c.setFont(font_name, 11)
        c.drawString(x, y, text)
        top = y - line_h

    # Logo en en-tête (côté gauche)
    try:
        from django.contrib.staticfiles import finders
        logo_path = finders.find('logos/logo.png')
        
        if logo_path and ImageReader is not None:
            try:
                logo_img = ImageReader(logo_path)
                logo_w, logo_h = 80, 80
                c.drawImage(logo_img, left, top - logo_h, width=logo_w, height=logo_h, preserveAspectRatio=True, mask='auto')
                
                # Titre à côté du logo
                c.setFont('Helvetica-Bold', 18)
                c.drawString(left + logo_w + 20, top - 25, "REÇU DE PAIEMENT")
                
                # Nom de l'école sous le titre
                c.setFont('Helvetica-Bold', 12)
                ecole_nom = paiement.eleve.classe.ecole.nom if getattr(paiement.eleve.classe, 'ecole', None) else ""
                c.drawString(left + logo_w + 20, top - 45, ecole_nom)
                
                top -= (logo_h + 20)
            except Exception:
                # Fallback sans logo
                c.setFont('Helvetica-Bold', 18)
                c.drawString(left, top, "REÇU DE PAIEMENT")
                top -= 15
                c.setFont('Helvetica-Bold', 12)
                c.drawString(left, top, (paiement.eleve.classe.ecole.nom if getattr(paiement.eleve.classe, 'ecole', None) else ""))
                top -= 25
        else:
            # Fallback sans logo
            c.setFont('Helvetica-Bold', 18)
            c.drawString(left, top, "REÇU DE PAIEMENT")
            top -= 15
            c.setFont('Helvetica-Bold', 12)
            c.drawString(left, top, (paiement.eleve.classe.ecole.nom if getattr(paiement.eleve.classe, 'ecole', None) else ""))
            top -= 25
    except Exception:
        # Fallback en cas d'erreur
        c.setFont('Helvetica-Bold', 18)
        c.drawString(left, top, "REÇU DE PAIEMENT")
        top -= 15
        c.setFont('Helvetica-Bold', 12)
        c.drawString(left, top, (paiement.eleve.classe.ecole.nom if getattr(paiement.eleve.classe, 'ecole', None) else ""))
        top -= 25

    # Photo élève (en haut à droite si disponible) ou placeholder avec initiales si absente
    try:
        img_drawn = False
        img_w, img_h = 100, 100
        x_img = width - 40 - img_w
        y_img = height - 40 - img_h
        if ImageReader is not None:
            photo_path = getattr(getattr(paiement.eleve, 'photo', None), 'path', None)
            if photo_path and os.path.exists(photo_path):
                try:
                    img = ImageReader(photo_path)
                    c.drawImage(img, x_img, y_img, width=img_w, height=img_h, preserveAspectRatio=True, mask='auto')
                    img_drawn = True
                except Exception:
                    img_drawn = False
        if not img_drawn:
            # Dessiner un placeholder avec initiales
            nom_complet = str(getattr(paiement.eleve, 'nom_complet', '') or '').strip()
            initiales = ''.join([p[0].upper() for p in nom_complet.split()[:2]]) or 'E'
            c.setLineWidth(1)
            try:
                c.roundRect(x_img, y_img, img_w, img_h, 8)
            except Exception:
                c.rect(x_img, y_img, img_w, img_h)
            c.setFont('Helvetica-Bold', 24)
            c.drawCentredString(x_img + img_w/2, y_img + img_h/2 - 8, initiales)
            c.setFont('Helvetica', 8)
            c.drawCentredString(x_img + img_w/2, y_img + 6, "Pas de photo")
        # Afficher le nom de l'élève sous l'image/placeholder
        try:
            nom_aff = str(getattr(paiement.eleve, 'nom_complet', '') or '').strip()
            if nom_aff:
                c.setFont('Helvetica', 9)
                c.drawCentredString(x_img + img_w/2, y_img - 12, nom_aff)
        except Exception:
            pass
    except Exception:
        # En cas de problème avec le rendu de la photo/placeholder, ne pas bloquer la génération du reçu
        pass

    # Informations paiement
    draw_line(f"Numéro de reçu : {paiement.numero_recu}", bold=True)
    draw_line(f"Date de paiement : {paiement.date_paiement.strftime('%d/%m/%Y')}")
    draw_line(f"Type de paiement : {paiement.type_paiement.nom}")
    draw_line(f"Mode de paiement : {paiement.mode_paiement.nom}")
    if getattr(paiement, 'reference_externe', None):
        draw_line(f"Référence externe : {paiement.reference_externe}")
    if getattr(paiement, 'observations', None):
        # Limiter l'observation à une ligne raisonnable pour le reçu
        obs = str(paiement.observations).strip()
        if obs:
            draw_line(f"Observations : {obs}")
    draw_line(f"Montant : {str(f'{paiement.montant:,.0f}').replace(',', ' ')} GNF", bold=True)

    if remises_total and int(remises_total) > 0:
        draw_line(f"Total remises : -{str(f'{int(remises_total):,}').replace(',', ' ')} GNF")
    # Montant net (jamais négatif)
    montant_net = max(0, int(paiement.montant - (remises_total or 0)))
    draw_line(f"Montant net à payer : {str(f'{montant_net:,}').replace(',', ' ')} GNF", bold=True)

    # Affectation du paiement courant sur les tranches (simulation déterministe)
    # Objectif: montrer, pour CE reçu, quelle partie couvre Inscription/T1/T2/T3
    try:
        echeancier_for_alloc = getattr(paiement.eleve, 'echeancier', None)
    except Exception:
        echeancier_for_alloc = None
    if echeancier_for_alloc:
        try:
            # Restants initiaux égaux aux dus de l'échéancier
            rest_insc = int(echeancier_for_alloc.frais_inscription_du or 0)
            rest_t1 = int(echeancier_for_alloc.tranche_1_due or 0)
            rest_t2 = int(echeancier_for_alloc.tranche_2_due or 0)
            rest_t3 = int(echeancier_for_alloc.tranche_3_due or 0)

            # Parcourir tous les paiements validés (y compris celui-ci) dans l'ordre
            paiements_valides = (
                Paiement.objects
                .filter(eleve=paiement.eleve, statut='VALIDE')
                .order_by('date_paiement', 'date_creation', 'id')
            )

            allocations = {}
            for p in paiements_valides.iterator():
                # Couverture de ce paiement = montant + remises sur CE paiement
                try:
                    rem_p = p.remises.aggregate(total=Sum('montant_remise')).get('total') or 0
                except Exception:
                    rem_p = 0
                reste_a_repartir = max(0, int(p.montant) + int(rem_p))

                a_insc = a_t1 = a_t2 = a_t3 = 0
                if reste_a_repartir and rest_insc > 0:
                    a = min(rest_insc, reste_a_repartir)
                    a_insc = a
                    rest_insc -= a
                    reste_a_repartir -= a
                if reste_a_repartir and rest_t1 > 0:
                    a = min(rest_t1, reste_a_repartir)
                    a_t1 = a
                    rest_t1 -= a
                    reste_a_repartir -= a
                if reste_a_repartir and rest_t2 > 0:
                    a = min(rest_t2, reste_a_repartir)
                    a_t2 = a
                    rest_t2 -= a
                    reste_a_repartir -= a
                if reste_a_repartir and rest_t3 > 0:
                    a = min(rest_t3, reste_a_repartir)
                    a_t3 = a
                    rest_t3 -= a
                    reste_a_repartir -= a

                allocations[p.id] = (a_insc, a_t1, a_t2, a_t3)

            if allocations.get(paiement.id):
                top -= 6
                draw_line("Affectation du paiement", bold=True)
                a_insc, a_t1, a_t2, a_t3 = allocations[paiement.id]
                draw_line(f"Inscription: {str(f'{int(a_insc):,}').replace(',', ' ')} GNF")
                draw_line(f"1ère tranche: {str(f'{int(a_t1):,}').replace(',', ' ')} GNF")
                draw_line(f"2ème tranche: {str(f'{int(a_t2):,}').replace(',', ' ')} GNF")
                draw_line(f"3ème tranche: {str(f'{int(a_t3):,}').replace(',', ' ')} GNF")
        except Exception:
            pass

    # Élève
    top -= 6
    draw_line("Informations de l'élève", bold=True)
    draw_line(f"Nom : {paiement.eleve.nom_complet}")
    if getattr(paiement.eleve, 'matricule', None):
        draw_line(f"Matricule : {paiement.eleve.matricule}")
    if getattr(paiement.eleve, 'classe', None):
        draw_line(f"Classe : {paiement.eleve.classe}")

    # Échéances (si disponibles sur l'échéancier de l'élève)
    try:
        echeancier = getattr(paiement.eleve, 'echeancier', None)
    except Exception:
        echeancier = None
    if echeancier:
        top -= 6
        draw_line("Échéances", bold=True)
        try:
            def _fmt_amount(v):
                try:
                    return str(f"{int(v or 0):,}").replace(',', ' ')
                except Exception:
                    return str(v or 0)
            def _fmt_date(d):
                try:
                    return d.strftime('%d/%m/%Y') if d else ''
                except Exception:
                    return str(d) if d else ''
            # Inscription
            draw_line(f"Inscription: {_fmt_amount(echeancier.frais_inscription_du)} GNF - Échéance: {_fmt_date(echeancier.date_echeance_inscription)}")
            # Tranches
            draw_line(f"1ère tranche: {_fmt_amount(echeancier.tranche_1_due)} GNF - Échéance: {_fmt_date(echeancier.date_echeance_tranche_1)}")
            draw_line(f"2ème tranche: {_fmt_amount(echeancier.tranche_2_due)} GNF - Échéance: {_fmt_date(echeancier.date_echeance_tranche_2)}")
            draw_line(f"3ème tranche: {_fmt_amount(echeancier.tranche_3_due)} GNF - Échéance: {_fmt_date(echeancier.date_echeance_tranche_3)}")
        except Exception:
            pass

        # Restes à payer par tranche
        try:
            def _reste(due, paye):
                try:
                    return max(0, int((due or 0) - (paye or 0)))
                except Exception:
                    return 0
            # Calcul global basé sur les paiements validés: somme(montants) - somme(remises)
            try:
                total_du = int((echeancier.frais_inscription_du or 0) + (echeancier.tranche_1_due or 0) + (echeancier.tranche_2_due or 0) + (echeancier.tranche_3_due or 0))
            except Exception:
                total_du = 0

            try:
                aggs = (
                    Paiement.objects
                    .filter(eleve=paiement.eleve, statut='VALIDE')
                    .aggregate(sum_montant=Sum('montant'), sum_remises=Sum('remises__montant_remise'))
                )
                sum_montant = int(aggs.get('sum_montant') or 0)
                sum_remises = int(aggs.get('sum_remises') or 0)
            except Exception:
                sum_montant = 0
                sum_remises = 0

            # Calcul de la couverture: montants payés + remises validées (les remises couvrent une partie du dû)
            couverture_validee = max(0, int(sum_montant) + int(sum_remises))
            # Inclure le paiement courant s'il n'est pas encore validé (montant + remises sur ce reçu)
            try:
                couverture_courante = max(0, int(paiement.montant) + int(remises_total or 0))
            except Exception:
                couverture_courante = 0
            couverture_effective = couverture_validee + (couverture_courante if paiement.statut != 'VALIDE' else 0)
            tout_solde = (total_du <= couverture_effective)
            solde_global = max(0, int(total_du - couverture_effective))

            top -= 6
            # Solde global restant
            draw_line(f"Solde global restant : {str(f'{solde_global:,}').replace(',', ' ')} GNF", bold=True)
            draw_line("Restes à payer par tranche", bold=True)
            if tout_solde:
                r_insc = r_t1 = r_t2 = r_t3 = 0
            else:
                r_insc = _reste(echeancier.frais_inscription_du, echeancier.frais_inscription_paye)
                r_t1 = _reste(echeancier.tranche_1_due, echeancier.tranche_1_payee)
                r_t2 = _reste(echeancier.tranche_2_due, echeancier.tranche_2_payee)
                r_t3 = _reste(echeancier.tranche_3_due, echeancier.tranche_3_payee)
            draw_line(f"Inscription: {str(f'{r_insc:,}').replace(',', ' ')} GNF")
            draw_line(f"1ère tranche: {str(f'{r_t1:,}').replace(',', ' ')} GNF")
            draw_line(f"2ème tranche: {str(f'{r_t2:,}').replace(',', ' ')} GNF")
            draw_line(f"3ème tranche: {str(f'{r_t3:,}').replace(',', ' ')} GNF")
        except Exception:
            pass

    # Remises détaillées
    if remises_total and int(remises_total) > 0:
        top -= 6
        draw_line("Remises appliquées", bold=True)
        for pr in paiement.remises.select_related('remise').all():
            nom = getattr(pr.remise, 'nom', 'Remise')
            montant = str(f"{int(pr.montant_remise):,}").replace(',', ' ')
            draw_line(f"- {nom} : -{montant} GNF")

    # Bloc signatures
    top -= 20
    c.setFont('Helvetica-Bold', 11)
    c.drawString(left, top, "Signatures")
    top -= 16
    # Lignes de signature (caissier et responsable)
    sig_line_y = top
    c.setLineWidth(0.8)
    # Caissier à gauche
    c.line(left, sig_line_y, left + 200, sig_line_y)
    c.setFont('Helvetica', 10)
    c.drawString(left, sig_line_y - 14, "Caissier(e)")
    # Responsable à droite
    right_x = left + 260
    c.setLineWidth(0.8)
    c.line(right_x, sig_line_y, right_x + 200, sig_line_y)
    c.setFont('Helvetica', 10)
    c.drawString(right_x, sig_line_y - 14, "Responsable")

    # Pied de page
    c.setFont('Helvetica', 9)
    c.drawRightString(width - 40, 30, f"Généré le {timezone.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage()
    c.save()

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"Recu_{paiement.numero_recu}.pdf"
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_liste_paiements_excel(request):
    """Exporte en Excel la liste des paiements selon les filtres (q, statut).
    Colonnes: Élève, Classe, École, Type, Montant, Mode, Date, Statut, N° Reçu, Observations
    """
    q = (request.GET.get('q') or '').strip()
    statut = (request.GET.get('statut') or '').strip()

    # Construire le queryset cohérent avec la liste
    qs = (
        Paiement.objects
        .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole', 'type_paiement', 'mode_paiement')
        .exclude(statut='ANNULE')
        .order_by('-date_paiement', '-date_creation')
    )
    if q:
        qs = qs.filter(
            Q(numero_recu__icontains=q)
            | Q(reference_externe__icontains=q)
            | Q(observations__icontains=q)
            | Q(eleve__nom__icontains=q)
            | Q(eleve__prenom__icontains=q)
            | Q(eleve__matricule__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)

    # Créer le classeur
    wb = Workbook()
    ws = wb.active
    ws.title = 'Paiements'

    headers = [
        'Élève', 'Classe', 'École', 'Type', 'Montant (GNF)', 'Mode', 'Date', 'Statut', 'N° Reçu', 'Observations'
    ]
    ws.append(headers)

    # Styles
    header_fill = PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='DDDDDD')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border_all

    # Lignes
    row_idx = 2
    for p in qs.iterator():
        eleve_nom = f"{getattr(p.eleve, 'nom', '')} {getattr(p.eleve, 'prenom', '')}".strip()
        classe_nom = getattr(getattr(p.eleve, 'classe', None), 'nom', '')
        ecole_nom = getattr(getattr(getattr(p.eleve, 'classe', None), 'ecole', None), 'nom', '')
        type_nom = getattr(p.type_paiement, 'nom', '')
        mode_nom = getattr(p.mode_paiement, 'nom', '')
        date_val = getattr(p, 'date_paiement', None)
        statut_txt = getattr(p, 'statut', '')
        recu = getattr(p, 'numero_recu', '')
        obs = getattr(p, 'observations', '') or ''

        ws.cell(row=row_idx, column=1, value=eleve_nom)
        ws.cell(row=row_idx, column=2, value=classe_nom)
        ws.cell(row=row_idx, column=3, value=ecole_nom)

        ws.cell(row=row_idx, column=4, value=type_nom)
        montant_cell = ws.cell(row=row_idx, column=5, value=float(p.montant or 0))
        montant_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        ws.cell(row=row_idx, column=6, value=mode_nom)

        date_cell = ws.cell(row=row_idx, column=7, value=date_val)
        date_cell.number_format = 'DD/MM/YYYY'
        ws.cell(row=row_idx, column=8, value=statut_txt)
        ws.cell(row=row_idx, column=9, value=recu)
        ws.cell(row=row_idx, column=10, value=obs)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = border_all
            if col in (1, 2, 3, 4, 6, 8, 9, 10):
                ws.cell(row=row_idx, column=col).alignment = Alignment(vertical='top')

        row_idx += 1

    # Ajustement des largeurs de colonnes
    widths = [22, 14, 18, 18, 16, 14, 12, 12, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Ligne de total montant
    if row_idx > 2:
        total_label_cell = ws.cell(row=row_idx, column=4, value='Total:')
        total_label_cell.font = Font(bold=True)
        total_cell = ws.cell(row=row_idx, column=5, value=f"=SUM(E2:E{row_idx-1})")
        total_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        total_cell.font = Font(bold=True)

    # Réponse HTTP
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"paiements_{ts}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def rapport_remises(request):
    """Rapport des remises avec agrégations par élève et filtres période/recherche.

    Contexte pour `templates/paiements/rapport_remises.html`:
      - rows: liste de dicts avec paiements/élève et champs: nb_remises, total_remise
      - total_global: somme de toutes les remises listées
      - q, date_debut, date_fin: filtres saisis
    """
    titre_page = "Rapport des remises"
    q = (request.GET.get('q') or '').strip()
    date_debut = (request.GET.get('date_debut') or '').strip()
    date_fin = (request.GET.get('date_fin') or '').strip()

    # Base queryset sur les remises liées à des paiements validés
    rem_qs = PaiementRemise.objects.select_related('paiement', 'paiement__eleve')
    rem_qs = rem_qs.filter(paiement__statut='VALIDE')

    # Filtre période sur la date du paiement si fournie
    try:
        if date_debut:
            rem_qs = rem_qs.filter(paiement__date_paiement__gte=date_debut)
        if date_fin:
            rem_qs = rem_qs.filter(paiement__date_paiement__lte=date_fin)
    except Exception:
        # En cas de format invalide, ignorer silencieusement
        pass

    # Filtre recherche simple sur élève
    if q:
        rem_qs = rem_qs.filter(
            Q(paiement__eleve__nom__icontains=q)
            | Q(paiement__eleve__prenom__icontains=q)
            | Q(paiement__eleve__matricule__icontains=q)
            | Q(paiement__eleve__classe__nom__icontains=q)
        )

    # Agrégations par élève
    rows = (
        rem_qs
        .values(
            'paiement__eleve__id',
            'paiement__eleve__prenom',
            'paiement__eleve__nom',
            'paiement__eleve__matricule',
            'paiement__eleve__classe__nom',
        )
        .annotate(
            nb_remises=Count('id'),
            total_remise=Coalesce(Sum('montant_remise'), Value(0, output_field=DecimalField(max_digits=10, decimal_places=0)))
        )
        .order_by('-total_remise')
    )

    total_global = 0
    try:
        total_global = int(rem_qs.aggregate(s=Coalesce(Sum('montant_remise'), Value(0)))['s'] or 0)
    except Exception:
        total_global = 0

    context = {
        'titre_page': titre_page,
        'q': q,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'rows': rows,
        'total_global': total_global,
    }
    template = 'paiements/rapport_remises.html' if _template_exists('paiements/rapport_remises.html') else None
    if template:
        return render(request, template, context)
    return HttpResponse('Rapport remises')

@login_required
@user_passes_test(lambda u: u.is_staff or (hasattr(u, 'profil') and u.profil.role in ['ADMIN', 'COMPTABLE', 'DIRECTEUR']))
def liste_eleves_soldes(request):
    """Liste des élèves soldés en tenant compte des remises (hors frais d'inscription).

    Règles:
    - Frais d'inscription (30 000 GNF) non impactés par les remises.
    - Remises s'appliquent uniquement à la scolarité (tranches 1..3).
    - Élève considéré soldé si: net_du = inscription_du + max(tranches_du - remises_totales, 0) est payé.
    """
    from django.utils import timezone as _tz
    today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()

    # Forcer 2025-2026 comme année courante pour cohérence avec les données
    default_annee = "2025-2026"
    annee = (request.GET.get('annee') or default_annee).strip()
    ecole_id = (request.GET.get('ecole_id') or '').strip()
    classe_id = (request.GET.get('classe_id') or '').strip()
    q = (request.GET.get('q') or '').strip()

    # Base queryset
    qs = (
        EcheancierPaiement.objects
        .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    )

    # Restreindre à l'année scolaire sélectionnée
    try:
        qs = qs.filter(annee_scolaire=annee)
    except Exception:
        pass

    # Filtres école/classe
    if ecole_id:
        qs = qs.filter(eleve__classe__ecole_id=ecole_id)
    if classe_id:
        qs = qs.filter(eleve__classe_id=classe_id)
    if q:
        qs = qs.filter(
            Q(eleve__nom__icontains=q) | Q(eleve__prenom__icontains=q) | Q(eleve__matricule__icontains=q)
        )

    # Expressions de calcul
    dues_sco = (
        Coalesce(
            F('tranche_1_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_2_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_3_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
    )
    paye_total = (
        Coalesce(
            F('frais_inscription_paye'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_1_payee'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_2_payee'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_3_payee'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
    )
    # Déterminer la période de l'année scolaire pour restreindre les remises aux paiements de l'année
    try:
        annee_debut = int(annee.split('-')[0])
        periode_debut = date(annee_debut, 9, 1)
        periode_fin = date(annee_debut + 1, 8, 31)
    except Exception:
        # Fallback simple si parsing échoue: limiter à l'année civile courante
        periode_debut = date(today.year, 1, 1)
        periode_fin = date(today.year, 12, 31)
    # Important: caper la fin de période à aujourd'hui pour éviter une fin future
    try:
        if periode_fin > today:
            periode_fin = today
    except Exception:
        pass

    remises_total = Coalesce(
        Sum(
            'eleve__paiements__remises__montant_remise',
            filter=(
                Q(eleve__paiements__statut='VALIDE') &
                Q(eleve__paiements__date_paiement__gte=periode_debut) &
                Q(eleve__paiements__date_paiement__lte=periode_fin)
            ),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    # Paiements validés sur la période (même s'ils ne sont pas encore alloués dans l'échéancier)
    paiements_valides_total = Coalesce(
        Sum(
            'eleve__paiements__montant',
            filter=(
                Q(eleve__paiements__statut='VALIDE') &
                Q(eleve__paiements__date_paiement__gte=periode_debut) &
                Q(eleve__paiements__date_paiement__lte=periode_fin)
            ),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    net_sco_du = Greatest(
        Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
        ExpressionWrapper(dues_sco - remises_total, output_field=DecimalField(max_digits=12, decimal_places=0))
    )
    net_du = ExpressionWrapper(
        Coalesce(
            F('frais_inscription_du'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + net_sco_du,
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    # Choisir le meilleur indicateur de "payé":
    # - prioriser la somme réelle des paiements validés (puisque l'allocation peut être différée)
    # - fallback sur les champs agrégés de l'échéancier
    paye_effectif = Greatest(
        Coalesce(paiements_valides_total, Value(0, output_field=DecimalField(max_digits=12, decimal_places=0))),
        Coalesce(paye_total, Value(0, output_field=DecimalField(max_digits=12, decimal_places=0))),
    )
    solde_calc = ExpressionWrapper(net_du - paye_effectif, output_field=DecimalField(max_digits=12, decimal_places=0))

    qs = qs.annotate(
        total_du_calc=net_du,
        total_paye_calc=paye_effectif,
        solde_calcule=solde_calc,
        total_remises_calc=remises_total,
    ).order_by('eleve__classe__nom', 'eleve__nom', 'eleve__prenom')

    # Élèves soldés: solde <= 0
    qs_soldes = qs.filter(solde_calcule__lte=0)

    # Totaux
    aggr = qs_soldes.aggregate(
        du=Coalesce(
            Sum('total_du_calc', output_field=DecimalField(max_digits=12, decimal_places=0)),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        paye=Coalesce(
            Sum('total_paye_calc', output_field=DecimalField(max_digits=12, decimal_places=0)),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        solde=Coalesce(
            Sum('solde_calcule', output_field=DecimalField(max_digits=12, decimal_places=0)),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        remises=Coalesce(
            Sum('total_remises_calc', output_field=DecimalField(max_digits=12, decimal_places=0)),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
    )

    # Pagination
    paginator = Paginator(qs_soldes, 25)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    # Options d'écoles/classes
    ecoles_qs = []
    try:
        from eleves.models import Ecole
        ecoles_qs = Ecole.objects.all().order_by('nom')
    except Exception:
        ecoles_qs = []
    classes = Classe.objects.select_related('ecole').all().order_by('ecole__nom', 'nom')

    context = {
        'annee': annee,
        'annees_options': [annee],
        'ecoles': ecoles_qs,
        'classes': classes,
        'ecole_id': ecole_id,
        'classe_id': classe_id,
        'q': q,
        'page_obj': page_obj,
        'totaux': {
            'du': int(aggr['du'] or 0),
            'paye': int(aggr['paye'] or 0),
            'solde': int(aggr['solde'] or 0),
            'remises': int(aggr['remises'] or 0),
        },
        'periode_debut': periode_debut,
        'periode_fin': periode_fin,
    }
    template = 'paiements/eleves_soldes.html' if _template_exists('paiements/eleves_soldes.html') else None
    if template:
        return render(request, template, context)
    return HttpResponse('Élèves soldés')

@login_required
def ajax_eleve_info(request):
    """Retourne des informations élève + échéancier pour le formulaire paiement.
    Attend un paramètre `matricule` (GET). Utilisé par `templates/paiements/form_paiement.html`.
    """
    matricule = request.GET.get('matricule') or request.POST.get('matricule')
    if not matricule:
        return JsonResponse({'success': False, 'error': 'Matricule requis.'}, status=400)

    try:
        eleve = Eleve.objects.select_related('classe', 'classe__ecole').get(matricule__iexact=matricule)
    except Eleve.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Élève introuvable.'}, status=404)

    # Construire la réponse
    # Sécuriser l'accès à l'URL de la photo (FieldFile.url peut lever une exception si vide)
    photo_url = ''
    try:
        photo_field = getattr(eleve, 'photo', None)
        if photo_field and getattr(photo_field, 'name', ''):
            photo_url = photo_field.url
    except Exception:
        photo_url = ''

    data = {
        'success': True,
        'eleve': {
            'id': eleve.id,
            'matricule': getattr(eleve, 'matricule', ''),
            'nom': getattr(eleve, 'nom', ''),
            'prenom': getattr(eleve, 'prenom', ''),
            'classe': getattr(eleve.classe, 'nom', '') if getattr(eleve, 'classe', None) else '',
            'ecole': getattr(eleve.classe.ecole, 'nom', '') if getattr(eleve, 'classe', None) and getattr(eleve.classe, 'ecole', None) else '',
            'photo_url': photo_url,
        },
        'echeancier': None,
        'has_echeancier': False,
    }

    # Échéancier (si présent)
    try:
        echeancier = getattr(eleve, 'echeancier', None)
    except Exception:
        echeancier = None

    if echeancier:
        data['echeancier'] = {
            'inscription_du': int(echeancier.frais_inscription_du or 0),
            'inscription_paye': int(echeancier.frais_inscription_paye or 0),
            'tranche_1_du': int(echeancier.tranche_1_due or 0),
            'tranche_1_paye': int(echeancier.tranche_1_payee or 0),
            'tranche_2_du': int(echeancier.tranche_2_due or 0),
            'tranche_2_paye': int(echeancier.tranche_2_payee or 0),
            'tranche_3_du': int(echeancier.tranche_3_due or 0),
            'tranche_3_paye': int(echeancier.tranche_3_payee or 0),
            'total_du': int(echeancier.total_du or 0),
            'total_paye': int(echeancier.total_paye or 0),
            'reste_a_payer': int((echeancier.total_du or 0) - (echeancier.total_paye or 0)),
        }
        data['has_echeancier'] = True

    return JsonResponse(data)

@login_required
def ajax_classes_par_ecole(request):
    return JsonResponse({'ok': True, 'classes': []})

@login_required
def ajax_statistiques_paiements(request):
    """Endpoint AJAX minimal pour statistiques paiements.
    Fourni pour satisfaire le routage; peut être enrichi ultérieurement.
    """
    try:
        total = Paiement.objects.count()
        montant_total = int(Paiement.objects.aggregate(total=Sum('montant'))['total'] or 0)
    except Exception:
        total = 0
        montant_total = 0
    return JsonResponse({'success': True, 'total': total, 'montant_total': montant_total})

@login_required
@require_http_methods(["GET", "POST"])
def ajax_calculer_remise(request):
    """Calcule un aperçu de remise. Implémentation basique pour compatibilité UI.
    Accepte un paramètre 'montant' et retourne le même montant en sortie par défaut.
    """
    montant_raw = request.GET.get('montant') or request.POST.get('montant') or '0'
    try:
        montant = int(float(str(montant_raw).replace(' ', '').replace(',', '.')))
    except Exception:
        montant = 0
    return JsonResponse({
        'success': True,
        'montant_initial': montant,
        'montant_apres_remise': montant,
        'details': [],
    })

@login_required
@can_apply_discounts
def appliquer_remise_paiement(request, paiement_id:int):
    """Affiche et traite le formulaire d'application de remises pour un paiement."""
    paiement = get_object_or_404(
        Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement'),
        pk=paiement_id,
    )

    # Seuls les paiements en attente peuvent être modifiés
    if getattr(paiement, 'statut', 'EN_ATTENTE') != 'EN_ATTENTE':
        messages.warning(request, "Seuls les paiements en attente peuvent recevoir des remises.")
        return redirect('paiements:detail_paiement', paiement_id=paiement.id)

    # Base scolarité = T1+T2+T3 (hors inscription)
    base_scolarite = 0
    try:
        ech = getattr(paiement.eleve, 'echeancier', None)
        if ech:
            base_scolarite = int((ech.tranche_1_due or 0) + (ech.tranche_2_due or 0) + (ech.tranche_3_due or 0))
        if not base_scolarite:
            # Fallback via grille tarifaire de la classe
            try:
                from eleves.models import GrilleTarifaire as _Grille
                classe = getattr(paiement.eleve, 'classe', None)
                ecole = getattr(classe, 'ecole', None)
                niveau = getattr(classe, 'niveau', None)
                annee = getattr(classe, 'annee_scolaire', None)
                grille = None
                if ecole and niveau and annee:
                    grille = _Grille.objects.filter(ecole=ecole, niveau=niveau, annee_scolaire=annee).first()
                if not grille and ecole and niveau:
                    grille = _Grille.objects.filter(ecole=ecole, niveau=niveau).order_by('-annee_scolaire').first()
                if grille:
                    base_scolarite = int((grille.tranche_1 or 0) + (grille.tranche_2 or 0) + (grille.tranche_3 or 0))
            except Exception:
                pass
    except Exception:
        base_scolarite = 0

    if request.method == 'POST':
        form = PaiementRemiseForm(request.POST, paiement=paiement)
        if form.is_valid():
            remises = form.cleaned_data.get('remises') or []
            pct_str = form.cleaned_data.get('pourcentage_scolarite') or ''
            try:
                pct_value = int(pct_str) if str(pct_str).isdigit() else 0
            except Exception:
                pct_value = 0
            # Si aucune remise n'est sélectionnée, ne rien modifier et afficher une erreur
            if not remises and pct_value <= 0:
                messages.error(request, "Aucune remise sélectionnée. Aucune modification n'a été effectuée.")
                try:
                    remises_existantes = list(paiement.remises.select_related('remise').all())
                except Exception:
                    remises_existantes = []
                context = {
                    'paiement': paiement,
                    'form': form,
                    'remises_existantes': remises_existantes,
                    'base_scolarite': int(base_scolarite or 0),
                }
                return render(request, 'paiements/appliquer_remise.html', context)

            # pourcentage_scolarite est un aperçu UI, on ne le persiste pas ici faute de modèle dédié
            with transaction.atomic():
                # Remplacer les remises existantes par la sélection
                PaiementRemise.objects.filter(paiement=paiement).delete()
                created = 0
                for remise in remises:
                    try:
                        montant_remise = remise.calculer_remise(paiement.montant)
                    except Exception:
                        montant_remise = 0
                    PaiementRemise.objects.create(
                        paiement=paiement,
                        remise=remise,
                        montant_remise=montant_remise,
                    )
                    created += 1

                # Appliquer également la remise scolarité (%) si choisie
                if pct_value > 0:
                    from datetime import date
                    annee = paiement.date_paiement.year
                    # Chercher une remise existante "Remise scolarité X%" active et couvrant la date
                    nom_remise = f"Remise scolarité {pct_value}%"
                    remise_pct = RemiseReduction.objects.filter(
                        nom=nom_remise,
                        type_remise='POURCENTAGE',
                        valeur=pct_value,
                        actif=True,
                        date_debut__lte=paiement.date_paiement,
                        date_fin__gte=paiement.date_paiement,
                    ).first()
                    if not remise_pct:
                        # Créer une remise "technique" pour l'année en cours
                        remise_pct = RemiseReduction.objects.create(
                            nom=nom_remise,
                            type_remise='POURCENTAGE',
                            valeur=pct_value,
                            motif='AUTRE',
                            description="Remise scolarité variable (technique)",
                            date_debut=date(annee, 1, 1),
                            date_fin=date(annee, 12, 31),
                            actif=True,
                        )
                    # 3% s'applique sur base scolarité (T1+T2+T3), pas sur le montant du paiement
                    try:
                        montant_remise_pct = (base_scolarite * pct_value) / 100
                    except Exception:
                        montant_remise_pct = (paiement.montant * pct_value) / 100
                    # Ne jamais dépasser le montant du paiement
                    try:
                        from decimal import Decimal as _D
                        montant_remise_pct = min(_D(montant_remise_pct), _D(paiement.montant))
                    except Exception:
                        try:
                            montant_remise_pct = min(float(montant_remise_pct), float(paiement.montant))
                        except Exception:
                            pass
                    PaiementRemise.objects.create(
                        paiement=paiement,
                        remise=remise_pct,
                        montant_remise=montant_remise_pct,
                    )
                    created += 1
            messages.success(request, f"Remises appliquées: {created}.")
            return redirect('paiements:detail_paiement', paiement_id=paiement.id)
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire de remises.")
    else:
        form = PaiementRemiseForm(paiement=paiement)

    # Remises déjà liées au paiement (pour affichage et cases cochées)
    try:
        remises_existantes = list(paiement.remises.select_related('remise').all())
    except Exception:
        remises_existantes = []

    context = {
        'paiement': paiement,
        'form': form,
        'remises_existantes': remises_existantes,
        'base_scolarite': int(base_scolarite or 0),
    }
    return render(request, 'paiements/appliquer_remise.html', context)

@login_required
def calculateur_remise(request):
    return HttpResponse('Calculateur de remise (placeholder)')

@login_required
@can_apply_discounts
def annuler_remise_paiement(request, paiement_id:int, remise_id:int=None):
    """Annule les remises appliquées à un paiement.

    - Si remise_id est fourni: supprime uniquement cette remise
    - Sinon: supprime toutes les remises du paiement
    """
    paiement = get_object_or_404(Paiement, pk=paiement_id)
    try:
        if remise_id:
            PaiementRemise.objects.filter(paiement=paiement, id=remise_id).delete()
            messages.success(request, "Remise supprimée.")
        else:
            PaiementRemise.objects.filter(paiement=paiement).delete()
            messages.success(request, "Toutes les remises de ce paiement ont été supprimées.")
    except Exception:
        messages.error(request, "Impossible d'annuler la remise.")
    return redirect('paiements:detail_paiement', paiement_id=paiement.id)

@login_required
def export_paiements_periode_excel(request):
    """Exporte les paiements entre deux dates (du, au) en Excel.
    Paramètres: ?du=YYYY-MM-DD&au=YYYY-MM-DD&statut=VALIDE|EN_ATTENTE|... (optionnel)
    """
    du = request.GET.get('du')
    au = request.GET.get('au')
    statut = (request.GET.get('statut') or '').strip()

    qs = Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement')
    # Filtres période
    try:
        if du:
            qs = qs.filter(date_paiement__gte=du)
        if au:
            qs = qs.filter(date_paiement__lte=au)
    except Exception:
        pass
    if statut:
        qs = qs.filter(statut=statut)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Paiements'
    headers = ['Élève', 'Matricule', 'Classe', 'École', 'Type', 'Montant', 'Mode', 'Date', 'Statut', 'N° Reçu']
    ws.append(headers)
    for p in qs.order_by('date_paiement', 'id'):
        ws.append([
            f"{getattr(p.eleve, 'nom', '')} {getattr(p.eleve, 'prenom', '')}",
            getattr(p.eleve, 'matricule', ''),
            getattr(getattr(p.eleve, 'classe', None), 'nom', ''),
            getattr(getattr(getattr(p.eleve, 'classe', None), 'ecole', None), 'nom', ''),
            getattr(p.type_paiement, 'nom', ''),
            int(p.montant or 0),
            getattr(p.mode_paiement, 'nom', ''),
            getattr(p, 'date_paiement', None).strftime('%Y-%m-%d') if getattr(p, 'date_paiement', None) else '',
            p.statut,
            p.numero_recu or '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = 'paiements_periode.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def rapport_retards(request):
    """Rapport des élèves en retard de paiement (montant exigible > payé+remises).
    Filtres: ?classe_id=&ecole_id=&du=&au=
    """
    from django.utils import timezone as _tz
    today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()

    exigible_expr = (
        Case(
            When(date_echeance_inscription__lte=today, then=F('frais_inscription_du')),
            default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_1__lte=today, then=F('tranche_1_due')),
            default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_2__lte=today, then=F('tranche_2_due')),
            default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_3__lte=today, then=F('tranche_3_due')),
            default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
        )
    )
    remises_expr = Coalesce(
        Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
        Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    remises_applicables = Least(remises_expr, exigible_expr)
    paye_effectif_expr = (
        F('frais_inscription_paye') + F('tranche_1_payee') + F('tranche_2_payee') + F('tranche_3_payee') + remises_applicables
    )
    retard_expr = ExpressionWrapper(exigible_expr - paye_effectif_expr, output_field=DecimalField(max_digits=12, decimal_places=0))

    qs = (EcheancierPaiement.objects
          .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
          .annotate(retard=retard_expr)
          .filter(retard__gt=0)
          .order_by('-retard'))

    context = {'titre_page': 'Rapport des retards', 'items': qs}
    if _template_exists('rapports/liste_rapports.html'):
        return render(request, 'rapports/liste_rapports.html', context)
    return HttpResponse(f"Retards: {qs.count()} élèves en retard")

@login_required
def rapport_encaissements(request):
    """Rapport des encaissements entre ?du=&au=, somme et décompte par statut."""
    du = request.GET.get('du')
    au = request.GET.get('au')
    qs = Paiement.objects.all()
    try:
        if du:
            qs = qs.filter(date_paiement__gte=du)
        if au:
            qs = qs.filter(date_paiement__lte=au)
    except Exception:
        pass
    total = int(qs.aggregate(total=Sum('montant'))['total'] or 0)
    par_statut = list(qs.values('statut').annotate(count=Count('id'), somme=Coalesce(Sum('montant'), Value(0))).order_by('statut'))
    context = {'titre_page': 'Rapport des encaissements', 'total': total, 'par_statut': par_statut}
    if _template_exists('rapports/tableau_bord.html'):
        return render(request, 'rapports/tableau_bord.html', context)
    return JsonResponse({'total': total, 'par_statut': par_statut})

@login_required
def api_paiements_list(request):
    """API JSON: liste des paiements avec filtres simples (?q=&statut=&limit=)."""
    q = (request.GET.get('q') or '').strip()
    statut = (request.GET.get('statut') or '').strip()
    try:
        limit = int(request.GET.get('limit') or 50)
    except Exception:
        limit = 50
    qs = Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement')
    if q:
        qs = qs.filter(
            Q(numero_recu__icontains=q) | Q(reference_externe__icontains=q) | Q(observations__icontains=q)
            | Q(eleve__nom__icontains=q) | Q(eleve__prenom__icontains=q) | Q(eleve__matricule__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)
    data = []
    for p in qs.order_by('-date_paiement', '-id')[:limit]:
        data.append({
            'id': p.id,
            'eleve': {
                'id': getattr(p.eleve, 'id', None),
                'matricule': getattr(p.eleve, 'matricule', ''),
                'nom': getattr(p.eleve, 'nom', ''),
                'prenom': getattr(p.eleve, 'prenom', ''),
            },
            'type': getattr(p.type_paiement, 'nom', ''),
            'mode': getattr(p.mode_paiement, 'nom', ''),
            'montant': int(p.montant or 0),
            'date': getattr(p, 'date_paiement', None).strftime('%Y-%m-%d') if getattr(p, 'date_paiement', None) else None,
            'statut': p.statut,
            'numero_recu': p.numero_recu,
        })
    return JsonResponse({'results': data})

@login_required
def api_paiement_detail(request, pk:int):
    """API JSON: détail d'un paiement"""
    p = get_object_or_404(Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement'), pk=pk)
    data = {
        'id': p.id,
        'eleve': {
            'id': getattr(p.eleve, 'id', None),
            'matricule': getattr(p.eleve, 'matricule', ''),
            'nom': getattr(p.eleve, 'nom', ''),
            'prenom': getattr(p.eleve, 'prenom', ''),
        },
        'type': getattr(p.type_paiement, 'nom', ''),
        'mode': getattr(p.mode_paiement, 'nom', ''),
        'montant': int(p.montant or 0),
        'date': getattr(p, 'date_paiement', None).strftime('%Y-%m-%d') if getattr(p, 'date_paiement', None) else None,
        'statut': p.statut,
        'numero_recu': p.numero_recu,
        'remises_total': int(p.remises.aggregate(total=Sum('montant_remise')).get('total') or 0) if hasattr(p, 'remises') else 0,
    }
    return JsonResponse(data)

def _template_exists(path:str)->bool:
    """Utilitaire léger: détecte si un template existe dans le chargeur Django."""
    try:
        from django.template.loader import get_template
        get_template(path)
        return True
    except Exception:
        return False
