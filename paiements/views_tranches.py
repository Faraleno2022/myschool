from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.db.models import Sum
from django.utils import timezone
from datetime import date, datetime
from decimal import Decimal

from eleves.models import Classe
from paiements.models import Paiement
from utilisateurs.utils import user_is_admin, user_school
from rapports.utils import _draw_header_and_watermark

# ReportLab
# ReportLab: fera l'objet d'un import différé dans la vue PDF


def _annee_vers_dates(annee_scolaire: str):
    try:
        deb, fin = annee_scolaire.split('-')
        an_deb = int(deb)
        an_fin = int(fin)
        return an_deb, an_fin
    except Exception:
        # Fallback année en cours selon rentrée (Septembre)
        today = timezone.now().date()
        y = today.year
        if today.month >= 9:
            return y, y + 1
        return y - 1, y


@login_required
def export_tranches_par_classe_pdf(request):
    """Export PDF des tranches par classe avec logo entête et filigrane.

    Filtres GET:
    - ecole: id de l'école
    - classe: id de la classe
    - annee_scolaire: ex '2024-2025'

    Respecte la séparation par école pour les non-admins.
    """
    # Contrôle d'accès: Admin ou Comptable uniquement
    is_admin = user_is_admin(request.user)
    is_comptable = False
    try:
        if hasattr(request.user, 'profil'):
            is_comptable = (getattr(request.user.profil, 'role', None) == 'COMPTABLE')
    except Exception:
        is_comptable = False
    if not (is_admin or is_comptable):
        return HttpResponseForbidden("Accès refusé: vous n'avez pas l'autorisation d'exporter ce rapport.")

    # Lecture et validation des paramètres
    raw_ecole = (request.GET.get('ecole') or '').strip()
    raw_classe = (request.GET.get('classe') or request.GET.get('classe_id') or '').strip()
    annee_scolaire = (request.GET.get('annee_scolaire') or '').strip()

    def parse_int(value):
        try:
            return int(value)
        except Exception:
            return None

    ecole_id = parse_int(raw_ecole) if raw_ecole else None
    classe_id = parse_int(raw_classe) if raw_classe else None

    # Scope classes
    classes = Classe.objects.select_related('ecole').all()
    ecole_user = user_school(request.user)
    restreindre = not user_is_admin(request.user) and ecole_user is not None
    if restreindre:
        classes = classes.filter(ecole=ecole_user)
    elif ecole_id:
        classes = classes.filter(ecole_id=ecole_id)
    if classe_id:
        classes = classes.filter(id=classe_id)

    # Anti-abus: limiter le nombre de classes exportées en une requête
    classes = classes.order_by('ecole__nom', 'niveau', 'nom')[:200]

    # Préparer réponse PDF
    response = HttpResponse(content_type='application/pdf')
    suffix = datetime.now().strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="tranches_par_classe_{suffix}.pdf"'

    # Import différé de ReportLab pour éviter les erreurs si non installé
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab n'est pas installé. Veuillez exécuter: pip install reportlab", status=500)

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=60, bottomMargin=30
    )
    elements = []
    styles = getSampleStyleSheet()
    cell = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=9)

    titre = 'Tranches par classe'
    if annee_scolaire:
        titre += f" – Année {annee_scolaire}"
    elements.append(Paragraph(titre, styles['Title']))
    elements.append(Spacer(1, 0.5*cm))

    header = [
        'Élève', 'Inscription payée', 'Tranche 1 payée', 'Tranche 2 payée', 'Tranche 3 payée',
        'Total dû', 'Total payé', 'Reste'
    ]

    def P(x):
        return Paragraph(str(x or ''), cell)

    # Parcours des classes
    for classe in classes:
        # Titre de la classe
        titre_classe = f"Classe: {classe.nom} – {getattr(classe.ecole, 'nom', '')}"
        elements.append(Paragraph(titre_classe, styles['Heading2']))
        elements.append(Spacer(1, 0.2*cm))

        data = [header]

        # Élèves de la classe
        # Utiliser le related_name défini sur Eleve.classe = 'eleves'
        eleves = getattr(classe, 'eleves', None)
        if eleves is not None:
            eleves = eleves.all().order_by('nom', 'prenom')
        else:
            eleves = []

        for e in eleves:
            # Tenter via échéancier s'il existe
            eche = getattr(e, 'echeancier', None)
            insc = t1 = t2 = t3 = Decimal('0')
            total_du = total_paye = reste = Decimal('0')

            if eche is not None and (not annee_scolaire or eche.annee_scolaire == annee_scolaire):
                insc = eche.frais_inscription_paye or 0
                t1 = eche.tranche_1_payee or 0
                t2 = eche.tranche_2_payee or 0
                t3 = eche.tranche_3_payee or 0
                total_du = (eche.frais_inscription_du or 0) + (eche.tranche_1_due or 0) + (eche.tranche_2_due or 0) + (eche.tranche_3_due or 0)
                total_paye = (insc or 0) + (t1 or 0) + (t2 or 0) + (t3 or 0)
                reste = (total_du or 0) - (total_paye or 0)
            else:
                # Fallback: somme depuis Paiement validé
                paiements = Paiement.objects.filter(eleve=e, statut='VALIDE')
                if annee_scolaire:
                    an_deb, an_fin = _annee_vers_dates(annee_scolaire)
                    start = date(an_deb, 9, 1)
                    end = date(an_fin, 8, 31)
                    paiements = paiements.filter(date_paiement__range=(start, end))
                insc = paiements.filter(type_paiement__nom__icontains='inscription').aggregate(total=Sum('montant'))['total'] or 0
                t1 = paiements.filter(type_paiement__nom__icontains='tranche 1').aggregate(total=Sum('montant'))['total'] or 0
                t2 = paiements.filter(type_paiement__nom__icontains='tranche 2').aggregate(total=Sum('montant'))['total'] or 0
                t3 = paiements.filter(type_paiement__nom__icontains='tranche 3').aggregate(total=Sum('montant'))['total'] or 0
                # Sans échéancier, on ne connaît pas le dû exact; on met 0 et reste = 0
                total_du = Decimal('0')
                total_paye = (insc or 0) + (t1 or 0) + (t2 or 0) + (t3 or 0)
                reste = Decimal('0')

            data.append([
                P(getattr(e, 'nom_complet', f"{e.nom} {e.prenoms}")),
                f"{insc:,}".replace(',', ' '),
                f"{t1:,}".replace(',', ' '),
                f"{t2:,}".replace(',', ' '),
                f"{t3:,}".replace(',', ' '),
                f"{total_du:,}".replace(',', ' '),
                f"{total_paye:,}".replace(',', ' '),
                f"{reste:,}".replace(',', ' '),
            ])

        # Construire la table pour la classe
        col_widths = [5.5*cm, 3*cm, 3*cm, 3*cm, 3*cm, 3*cm, 3*cm, 3*cm]
        table = Table(data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 8),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ('TOPPADDING', (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.6*cm))

    # Construire le document avec en-tête + filigrane logo
    doc.build(elements, onFirstPage=_draw_header_and_watermark, onLaterPages=_draw_header_and_watermark)
    return response

@login_required
def export_tranches_par_classe_excel(request):
    """Export Excel (XLSX) des tranches par classe: Élève, Inscription payée, Tranche 1, Tranche 2, Tranche 3, Total dû, Total payé, Reste.

    Filtres GET facultatifs: ecole, classe/classe_id, annee_scolaire.
    Respecte la séparation par école pour non-admin.
    """
    # Contrôle d'accès
    is_admin = user_is_admin(request.user)
    is_comptable = False
    try:
        if hasattr(request.user, 'profil'):
            is_comptable = (getattr(request.user.profil, 'role', None) == 'COMPTABLE')
    except Exception:
        is_comptable = False
    if not (is_admin or is_comptable):
        return HttpResponseForbidden("Accès refusé: vous n'avez pas l'autorisation d'exporter ce rapport.")

    # Import openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse("OpenPyXL n'est pas installé. Veuillez exécuter: pip install openpyxl", status=500)

    raw_ecole = (request.GET.get('ecole') or '').strip()
    raw_classe = (request.GET.get('classe') or request.GET.get('classe_id') or '').strip()
    annee_scolaire = (request.GET.get('annee_scolaire') or '').strip()

    def parse_int(value):
        try:
            return int(value)
        except Exception:
            return None

    ecole_id = parse_int(raw_ecole) if raw_ecole else None
    classe_id = parse_int(raw_classe) if raw_classe else None

    classes = Classe.objects.select_related('ecole').all()
    ecole_user = user_school(request.user)
    restreindre = not user_is_admin(request.user) and ecole_user is not None
    if restreindre:
        classes = classes.filter(ecole=ecole_user)
    elif ecole_id:
        classes = classes.filter(ecole_id=ecole_id)
    if classe_id:
        classes = classes.filter(id=classe_id)
    classes = classes.order_by('ecole__nom', 'niveau', 'nom')[:200]

    wb = Workbook()
    ws_index = wb.active
    ws_index.title = 'Index'
    ws_index.append(['Tranches par classe', f"Année: {annee_scolaire}" if annee_scolaire else ''])
    ws_index.append(['Écoles / Classes listées:'])

    headers = ['Élève', 'Inscription payée', 'Tranche 1 payée', 'Tranche 2 payée', 'Tranche 3 payée', 'Total dû', 'Total payé', 'Reste']

    from django.db.models import Sum
    from datetime import date
    from decimal import Decimal

    def annee_vers_dates(annee_scolaire: str):
        try:
            deb, fin = annee_scolaire.split('-')
            an_deb = int(deb)
            an_fin = int(fin)
            return an_deb, an_fin
        except Exception:
            today = timezone.now().date()
            y = today.year
            if today.month >= 9:
                return y, y + 1
            return y - 1, y

    for idx, classe in enumerate(classes, start=1):
        sheet_name = f"{classe.nom[:25]}"  # Limite Excel <=31
        ws = wb.create_sheet(title=sheet_name)
        ws.append([f"Classe: {classe.nom} – {getattr(classe.ecole, 'nom', '')}"])
        ws.append(headers)

        eleves_mgr = getattr(classe, 'eleves', None)
        eleves = eleves_mgr.all().order_by('nom', 'prenom') if eleves_mgr is not None else []

        for e in eleves:
            eche = getattr(e, 'echeancier', None)
            insc = t1 = t2 = t3 = Decimal('0')
            total_du = total_paye = reste = Decimal('0')
            if eche is not None and (not annee_scolaire or eche.annee_scolaire == annee_scolaire):
                insc = eche.frais_inscription_paye or 0
                t1 = eche.tranche_1_payee or 0
                t2 = eche.tranche_2_payee or 0
                t3 = eche.tranche_3_payee or 0
                total_du = (eche.frais_inscription_du or 0) + (eche.tranche_1_due or 0) + (eche.tranche_2_due or 0) + (eche.tranche_3_due or 0)
                total_paye = (insc or 0) + (t1 or 0) + (t2 or 0) + (t3 or 0)
                reste = (total_du or 0) - (total_paye or 0)
            else:
                paiements = Paiement.objects.filter(eleve=e, statut='VALIDE')
                if annee_scolaire:
                    an_deb, an_fin = annee_vers_dates(annee_scolaire)
                    start = date(an_deb, 9, 1)
                    end = date(an_fin, 8, 31)
                    paiements = paiements.filter(date_paiement__range=(start, end))
                insc = paiements.filter(type_paiement__nom__icontains='inscription').aggregate(total=Sum('montant'))['total'] or 0
                t1 = paiements.filter(type_paiement__nom__icontains='tranche 1').aggregate(total=Sum('montant'))['total'] or 0
                t2 = paiements.filter(type_paiement__nom__icontains='tranche 2').aggregate(total=Sum('montant'))['total'] or 0
                t3 = paiements.filter(type_paiement__nom__icontains='tranche 3').aggregate(total=Sum('montant'))['total'] or 0
                total_du = Decimal('0')
                total_paye = (insc or 0) + (t1 or 0) + (t2 or 0) + (t3 or 0)
                reste = Decimal('0')

            ws.append([
                getattr(e, 'nom_complet', f"{e.prenom} {e.nom}"),
                int(insc), int(t1), int(t2), int(t3), int(total_du), int(total_paye), int(reste)
            ])

        # Ajuster largeur colonnes simple
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 22 if col == 1 else 16

        # Index line
        ws_index.append([getattr(classe.ecole, 'nom', ''), classe.nom, sheet_name])

    # Supprimer la feuille par défaut si vide
    if ws_index.max_row == 2:
        ws_index.append(['Aucune classe'])

    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    resp = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    suffix = datetime.now().strftime('%Y%m%d')
    filename = f'tranches_par_classe_{suffix}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
